import os
import sqlite3
import re
import time
import datetime
import requests
import zipfile
import calendar
from io import BytesIO
from functools import wraps
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

import pandas as pd
import razorpay
from flask import (
    Flask, render_template, request, redirect, url_for,
    session, flash, send_file, jsonify
)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from reportlab.lib.pagesizes import A4, letter

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "").strip()
if not app.secret_key:
    raise RuntimeError("SECRET_KEY environment variable is not configured.")

# =========================
# PAYMENT / SUBSCRIPTION CONFIG
# =========================
# Live/Test keys .env or hosting environment variables se set karna hai.
# Keys ko code me hardcode nahi karna.
RAZORPAY_KEY_ID = os.getenv("RAZORPAY_KEY_ID", "").strip()
RAZORPAY_KEY_SECRET = os.getenv("RAZORPAY_KEY_SECRET", os.getenv("RAZORPAY_SECRET", "")).strip()

razorpay_client = (
    razorpay.Client(auth=(RAZORPAY_KEY_ID, RAZORPAY_KEY_SECRET))
    if RAZORPAY_KEY_ID and RAZORPAY_KEY_SECRET
    else None
)

PAYMENTS_ENABLED = os.getenv("PAYMENTS_ENABLED", "false").strip().lower() in ["1", "true", "yes", "on"]
DEMO_MODE = False

# Final rule: 1–5 employees all features free. Above 5 employees yearly subscription required.
FREE_EMPLOYEE_LIMIT = 5

ADMIN_USERNAMES = [
    "smarthireai5"
]

DB_NAME = "payroll_pro.db"
UPLOAD_FOLDER = "uploads"
PAYSLIP_FOLDER = "payslips"
INVOICE_FOLDER = "subscription_invoices"
COMPANY_ASSET_FOLDER = os.path.join(UPLOAD_FOLDER, "company_assets")

# Common invoice counter DB path.
# Server par isko /var/lib/smarthire/invoice_counter.db set karenge,
# taaki Payroll Pro aur CMPF dono same invoice sequence use karein.
COMMON_INVOICE_DB = os.getenv("COMMON_INVOICE_DB", "invoice_counter.db")
INVOICE_PREFIX = os.getenv("INVOICE_PREFIX", "INV")

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(PAYSLIP_FOLDER, exist_ok=True)
os.makedirs(INVOICE_FOLDER, exist_ok=True)
os.makedirs(COMPANY_ASSET_FOLDER, exist_ok=True)


# ---------------------------
# DATABASE
# ---------------------------
def get_db():
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    return conn


def safe_add_column(cur, table_name, column_name, column_type):
    cur.execute(f"PRAGMA table_info({table_name})")
    existing_columns = [col[1] for col in cur.fetchall()]

    if column_name not in existing_columns:
        cur.execute(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {column_type}")

def clean_text(value, default=""):
    if pd.isna(value):
        return default
    value = str(value).strip()
    if value.lower() in ["nan", "none", "null"]:
        return default
    return value


def clean_float(value, default=0):
    try:
        if pd.isna(value) or value == "":
            return default
        return float(value)
    except Exception:
        return default

def init_db():
    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
    CREATE TABLE IF NOT EXISTS companies (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_name TEXT NOT NULL
    )
    """)

    safe_add_column(cur, "companies", "address", "TEXT DEFAULT ''")
    safe_add_column(cur, "companies", "email", "TEXT DEFAULT ''")
    safe_add_column(cur, "companies", "phone", "TEXT DEFAULT ''")
    safe_add_column(cur, "companies", "overtime_multiplier", "REAL DEFAULT 1")
    safe_add_column(cur, "companies", "working_days_policy", "TEXT DEFAULT 'attendance'")
    safe_add_column(cur, "companies", "logo_path", "TEXT DEFAULT ''")
    safe_add_column(cur, "companies", "authorized_signature_path", "TEXT DEFAULT ''")
    safe_add_column(cur, "companies", "authorized_signatory", "TEXT DEFAULT ''")
    safe_add_column(cur, "companies", "authorized_designation", "TEXT DEFAULT ''")


    cur.execute("""
    CREATE TABLE IF NOT EXISTS compliance_settings (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER NOT NULL UNIQUE,

        pf_employee_rate REAL DEFAULT 12,
        pf_employer_rate REAL DEFAULT 12,
        pf_wage_ceiling REAL DEFAULT 15000,
        pf_max_deduction REAL DEFAULT 1800,

        esic_employee_rate REAL DEFAULT 0.75,
        esic_employer_rate REAL DEFAULT 3.25,
        esic_wage_limit REAL DEFAULT 21000,

        gratuity_rate REAL DEFAULT 4.81,
        bonus_rate REAL DEFAULT 8.33,

        tds_enabled INTEGER DEFAULT 0,

        salary_days_policy TEXT DEFAULT 'attendance',
        custom_salary_days REAL DEFAULT 30,
        count_weekly_off_paid INTEGER DEFAULT 1,
        count_paid_leave_paid INTEGER DEFAULT 1,
        count_holiday_paid INTEGER DEFAULT 1,
        deduct_lop INTEGER DEFAULT 1,

        festival_bonus_enabled INTEGER DEFAULT 0,
        festival_bonus_month INTEGER DEFAULT 10,

        updated_at TEXT,

        FOREIGN KEY(company_id) REFERENCES companies(id)
    )
""")

    compliance_columns = {
    "salary_days_policy": "TEXT DEFAULT 'attendance'",
    "custom_salary_days": "REAL DEFAULT 30",
    "count_weekly_off_paid": "INTEGER DEFAULT 1",
    "count_paid_leave_paid": "INTEGER DEFAULT 1",
    "count_holiday_paid": "INTEGER DEFAULT 1",
    "deduct_lop": "INTEGER DEFAULT 1",

    "festival_bonus_enabled": "INTEGER DEFAULT 0",
    "festival_bonus_month": "INTEGER DEFAULT 10",
    "bonus_min_service_days": "INTEGER DEFAULT 30",
    "bonus_prorata_enabled": "INTEGER DEFAULT 1"
}

    cur.execute("PRAGMA table_info(compliance_settings)")
    existing_columns = [col[1] for col in cur.fetchall()]

    for column_name, column_type in compliance_columns.items():
        if column_name not in existing_columns:
           cur.execute(f"""
            ALTER TABLE compliance_settings
            ADD COLUMN {column_name} {column_type}
        """)

    safe_add_column(cur, "compliance_settings", "salary_days_policy", "TEXT DEFAULT 'attendance'")
    safe_add_column(cur, "compliance_settings", "custom_salary_days", "REAL DEFAULT 30")
    safe_add_column(cur, "compliance_settings", "count_weekly_off_paid", "INTEGER DEFAULT 1")
    safe_add_column(cur, "compliance_settings", "count_paid_leave_paid", "INTEGER DEFAULT 1")
    safe_add_column(cur, "compliance_settings", "count_holiday_paid", "INTEGER DEFAULT 1")
    safe_add_column(cur, "compliance_settings", "deduct_lop", "INTEGER DEFAULT 1")

    safe_add_column(cur, "compliance_settings", "festival_bonus_enabled", "INTEGER DEFAULT 0")
    safe_add_column(cur, "compliance_settings", "festival_bonus_month", "INTEGER DEFAULT 10")

    safe_add_column(cur, "compliance_settings", "bonus_min_service_days", "INTEGER DEFAULT 30")
    safe_add_column(cur, "compliance_settings", "bonus_prorata_enabled", "INTEGER DEFAULT 1")

    cur.execute("""
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER NOT NULL,
        full_name TEXT NOT NULL,
        username TEXT NOT NULL UNIQUE,
        password_hash TEXT NOT NULL,
        FOREIGN KEY(company_id) REFERENCES companies(id)
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS employees (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER NOT NULL,
        emp_code TEXT NOT NULL,
        employee_name TEXT NOT NULL,
        role TEXT NOT NULL,
        department TEXT,
        gender TEXT DEFAULT 'male',
        monthly_salary REAL NOT NULL,
        tax_regime TEXT DEFAULT 'new',
        other_annual_deductions REAL DEFAULT 0,
        special_allowance REAL DEFAULT 0,
        UNIQUE(company_id, emp_code),
        FOREIGN KEY(company_id) REFERENCES companies(id)
    )
    """)

    # Extra employee details for payslip
    safe_add_column(cur, "employees", "uan_no", "TEXT DEFAULT ''")
    safe_add_column(cur, "employees", "esic_no", "TEXT DEFAULT ''")
    safe_add_column(cur, "employees", "bank_name", "TEXT DEFAULT ''")
    safe_add_column(cur, "employees", "account_no", "TEXT DEFAULT ''")
    safe_add_column(cur, "employees", "ifsc_code", "TEXT DEFAULT ''")

    cur.execute("""
    CREATE TABLE IF NOT EXISTS attendance (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER NOT NULL,
        emp_code TEXT NOT NULL,
        month TEXT NOT NULL,
        working_days INTEGER NOT NULL,
        present_days INTEGER NOT NULL,
        overtime_hours REAL DEFAULT 0,
        bonus REAL DEFAULT 0,
        manual_deduction REAL DEFAULT 0,
        FOREIGN KEY(company_id) REFERENCES companies(id)
    )
    """)

    safe_add_column(cur, "attendance", "weekly_off", "REAL DEFAULT 0")
    safe_add_column(cur, "attendance", "paid_leave", "REAL DEFAULT 0")
    safe_add_column(cur, "attendance", "holiday", "REAL DEFAULT 0")
    safe_add_column(cur, "attendance", "lop_days", "REAL DEFAULT 0")
    safe_add_column(cur, "attendance", "paid_days", "REAL DEFAULT 0")

    cur.execute("""
    CREATE TABLE IF NOT EXISTS leave_requests (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER NOT NULL,
        emp_code TEXT NOT NULL,
        leave_type TEXT NOT NULL,
        start_date TEXT NOT NULL,
        end_date TEXT NOT NULL,
        total_days REAL DEFAULT 0,
        reason TEXT,
        status TEXT DEFAULT 'Pending',
        admin_remark TEXT,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    )
""")

    cur.execute("""
    CREATE TABLE IF NOT EXISTS leave_balances (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER NOT NULL,
        emp_code TEXT NOT NULL,
        casual_leave REAL DEFAULT 0,
        sick_leave REAL DEFAULT 0,
        paid_leave REAL DEFAULT 0,
        used_leave REAL DEFAULT 0,
        updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(company_id, emp_code)
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS leave_policy_settings (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER NOT NULL UNIQUE,
        casual_leave_limit REAL DEFAULT 6,
        sick_leave_limit REAL DEFAULT 6,
        paid_leave_limit REAL DEFAULT 12,
        updated_at TEXT DEFAULT CURRENT_TIMESTAMP
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS payroll_history (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER NOT NULL,
        emp_code TEXT NOT NULL,
        employee_name TEXT NOT NULL,
        role TEXT NOT NULL,
        department TEXT,
        gender TEXT,
        month TEXT NOT NULL,
        monthly_salary REAL NOT NULL,
        basic REAL NOT NULL,
        da REAL NOT NULL,
        hra REAL NOT NULL,
        special_allowance REAL NOT NULL,
        other_allowance REAL NOT NULL,
        gross REAL NOT NULL,
        esi_employee REAL NOT NULL,
        professional_tax REAL NOT NULL,
        pf_employee REAL NOT NULL,
        lwf_employee REAL NOT NULL,
        tds REAL NOT NULL,
        manual_deduction REAL NOT NULL,
        total_deductions REAL NOT NULL,
        esi_employer REAL NOT NULL,
        pf_employer REAL NOT NULL,
        gratuity REAL NOT NULL,
        bonus_ctc REAL NOT NULL,
        festival_bonus REAL NOT NULL,
        lwf_employer REAL NOT NULL,
        total_contributions REAL NOT NULL,
        net_pay REAL NOT NULL,
        monthly_ctc REAL NOT NULL,
        annual_ctc REAL NOT NULL,
        overtime_hours REAL NOT NULL,
        overtime_amount REAL NOT NULL,
        created_at TEXT NOT NULL,
        FOREIGN KEY(company_id) REFERENCES companies(id)
    )
    """)

    safe_add_column(cur, "payroll_history", "run_id", "TEXT")
    safe_add_column(cur, "payroll_history", "is_current", "INTEGER DEFAULT 1")

    cur.execute("""
    CREATE TABLE IF NOT EXISTS full_final_settlements (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER NOT NULL,
        emp_code TEXT NOT NULL,
        employee_name TEXT,
        role TEXT,
        department TEXT,
        last_working_day TEXT NOT NULL,
        settlement_month TEXT NOT NULL,

    monthly_salary REAL DEFAULT 0,
    paid_days REAL DEFAULT 0,
    earned_salary REAL DEFAULT 0,

    leave_balance REAL DEFAULT 0,
    leave_encashment REAL DEFAULT 0,
    bonus_payable REAL DEFAULT 0,
    gratuity_payable REAL DEFAULT 0,
    other_earnings REAL DEFAULT 0,

    notice_recovery REAL DEFAULT 0,
    loan_recovery REAL DEFAULT 0,
    advance_recovery REAL DEFAULT 0,
    other_deductions REAL DEFAULT 0,

    total_earnings REAL DEFAULT 0,
    total_deductions REAL DEFAULT 0,
    final_payable REAL DEFAULT 0,

    reason TEXT,
    remarks TEXT,
    created_at TEXT NOT NULL,

    FOREIGN KEY(company_id) REFERENCES companies(id)
)
""")

    cur.execute("""
    CREATE TABLE IF NOT EXISTS subscriptions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER NOT NULL,
        plan_name TEXT NOT NULL,
        status TEXT NOT NULL,
        start_date TEXT,
        end_date TEXT,
        FOREIGN KEY(company_id) REFERENCES companies(id)
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS payments (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER NOT NULL,
        amount REAL NOT NULL,
        payment_id TEXT,
        order_id TEXT,
        status TEXT,
        created_at TEXT NOT NULL,
        FOREIGN KEY(company_id) REFERENCES companies(id)
    )
    """)

    safe_add_column(cur, "payments", "user_id", "INTEGER")

    cur.execute("""
    CREATE TABLE IF NOT EXISTS subscription_invoices (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER NOT NULL,
        user_id INTEGER,
        product_type TEXT NOT NULL DEFAULT 'PAYROLL_PRO',
        invoice_no TEXT NOT NULL UNIQUE,
        plan_id TEXT,
        plan_name TEXT,
        amount REAL DEFAULT 0,
        discount REAL DEFAULT 0,
        grand_total REAL DEFAULT 0,
        payment_id TEXT UNIQUE,
        order_id TEXT,
        payment_mode TEXT DEFAULT 'Razorpay',
        status TEXT DEFAULT 'PAID',
        invoice_date TEXT,
        subscription_start TEXT,
        subscription_end TEXT,
        pdf_path TEXT,
        created_at TEXT NOT NULL,
        FOREIGN KEY(company_id) REFERENCES companies(id)
    )
    """)


    conn.commit()
    conn.close()


def ensure_leave_tables():
    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS leave_policy (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_id INTEGER NOT NULL,
            casual_leave_limit REAL DEFAULT 6,
            sick_leave_limit REAL DEFAULT 6,
            paid_leave_limit REAL DEFAULT 12,
            created_at TEXT,
            UNIQUE(company_id)
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS leave_balances (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_id INTEGER NOT NULL,
            emp_code TEXT NOT NULL,
            casual_leave REAL DEFAULT 6,
            sick_leave REAL DEFAULT 6,
            paid_leave REAL DEFAULT 12,
            used_leave REAL DEFAULT 0,
            UNIQUE(company_id, emp_code)
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS leave_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_id INTEGER NOT NULL,
            emp_code TEXT NOT NULL,
            leave_type TEXT NOT NULL,
            start_date TEXT NOT NULL,
            end_date TEXT NOT NULL,
            total_days REAL DEFAULT 0,
            reason TEXT,
            status TEXT DEFAULT 'Pending',
            created_at TEXT
        )
    """)

    conn.commit()
    conn.close()


def add_leave_payroll_columns():
    conn = get_db()
    cur = conn.cursor()

    cur.execute("PRAGMA table_info(payroll_history)")
    existing_columns = [col["name"] for col in cur.fetchall()]

    new_columns = {
        "paid_leave_days": "REAL DEFAULT 0",
        "lwp_days": "REAL DEFAULT 0",
        "lwp_deduction": "REAL DEFAULT 0",
        "payable_days": "REAL DEFAULT 0"
    }

    for column_name, column_type in new_columns.items():
        if column_name not in existing_columns:
            cur.execute(f"""
                ALTER TABLE payroll_history
                ADD COLUMN {column_name} {column_type}
            """)

    conn.commit()
    conn.close()


def add_payment_order_id_column():
    conn = get_db()
    cur = conn.cursor()

    cur.execute("PRAGMA table_info(payments)")
    existing_columns = [col["name"] for col in cur.fetchall()]

    if "order_id" not in existing_columns:
        cur.execute("""
            ALTER TABLE payments
            ADD COLUMN order_id TEXT
        """)

    conn.commit()
    conn.close()



# ---------------------------
# SUBSCRIPTION INVOICE HELPERS
# ---------------------------
def ensure_common_invoice_db():
    folder = os.path.dirname(COMMON_INVOICE_DB)
    if folder:
        os.makedirs(folder, exist_ok=True)

    conn = sqlite3.connect(COMMON_INVOICE_DB, timeout=30)
    cur = conn.cursor()
    cur.execute("PRAGMA busy_timeout = 10000")
    cur.execute("""
        CREATE TABLE IF NOT EXISTS invoice_sequences (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            year INTEGER NOT NULL UNIQUE,
            last_number INTEGER NOT NULL DEFAULT 0,
            updated_at TEXT NOT NULL
        )
    """)
    conn.commit()
    conn.close()


def get_next_global_invoice_no():
    ensure_common_invoice_db()

    year = datetime.datetime.now().year
    now_text = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    conn = sqlite3.connect(COMMON_INVOICE_DB, timeout=30, isolation_level=None)
    cur = conn.cursor()
    cur.execute("PRAGMA busy_timeout = 10000")

    try:
        # BEGIN IMMEDIATE prevents duplicate invoice numbers during same-time payments.
        cur.execute("BEGIN IMMEDIATE")

        cur.execute("""
            SELECT last_number
            FROM invoice_sequences
            WHERE year = ?
        """, (year,))

        row = cur.fetchone()

        if row:
            next_number = int(row[0] or 0) + 1
            cur.execute("""
                UPDATE invoice_sequences
                SET last_number = ?, updated_at = ?
                WHERE year = ?
            """, (next_number, now_text, year))
        else:
            next_number = 1
            cur.execute("""
                INSERT INTO invoice_sequences (year, last_number, updated_at)
                VALUES (?, ?, ?)
            """, (year, next_number, now_text))

        conn.commit()
        return f"{INVOICE_PREFIX}-{year}-{next_number:03d}"

    except Exception:
        conn.rollback()
        raise

    finally:
        conn.close()


def invoice_clean(value, default="-"):
    value = "" if value is None else str(value).strip()
    if not value or value.lower() in ["none", "nan", "null"]:
        return default
    return value


def wrap_pdf_text(value, max_chars=80):
    words = invoice_clean(value, "").split()
    if not words:
        return ["-"]

    lines = []
    current = ""

    for word in words:
        if len(current) + len(word) + 1 <= max_chars:
            current = (current + " " + word).strip()
        else:
            lines.append(current)
            current = word

    if current:
        lines.append(current)

    return lines or ["-"]


def draw_wrapped_pdf_text(c, value, x, y, max_chars=80, line_height=11, font="Helvetica", size=9):
    c.setFont(font, size)
    for line in wrap_pdf_text(value, max_chars):
        c.drawString(x, y, line)
        y -= line_height
    return y


def generate_subscription_invoice_pdf(invoice_id):
    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            si.*,
            c.company_name,
            c.address AS customer_address,
            c.email AS customer_email,
            c.phone AS customer_phone,
            u.full_name AS user_full_name,
            u.username AS user_email
        FROM subscription_invoices si
        LEFT JOIN companies c ON c.id = si.company_id
        LEFT JOIN users u ON u.id = si.user_id
        WHERE si.id = ?
    """, (invoice_id,))

    invoice = cur.fetchone()
    conn.close()

    if not invoice:
        raise ValueError("Invoice record not found")

    os.makedirs(INVOICE_FOLDER, exist_ok=True)

    invoice_no = invoice_clean(invoice["invoice_no"], f"invoice_{invoice_id}")
    safe_invoice_no = "".join(ch if ch.isalnum() or ch in ["-", "_"] else "_" for ch in invoice_no)
    pdf_path = invoice["pdf_path"] or os.path.join(INVOICE_FOLDER, f"{safe_invoice_no}.pdf")

    c = canvas.Canvas(pdf_path, pagesize=letter)
    width, height = letter

    left = 42
    right = width - 42

    # Header ko top border se safe distance par rakha hai.
    # Pehle height - 42 tha, jisse PDF viewer me top border text ko touch/cut kar sakta tha.
    top = height - 70

    # Border
    c.setStrokeColorRGB(0, 0, 0)
    c.rect(30, 30, width - 60, height - 60, fill=0, stroke=1)

    # Header
    c.setFont("Helvetica-Bold", 18)
    c.drawString(left, top, "SmartHire AI")

    c.setFont("Helvetica-Bold", 14)
    c.drawRightString(right, top, "SUBSCRIPTION INVOICE / RECEIPT")

    c.setFont("Helvetica", 9)
    y = top - 18
    c.drawString(left, y, "AI-Powered Payroll & HR Software")
    y -= 12
    c.drawString(left, y, "Nagpur, Maharashtra, India")
    y -= 12
    c.drawString(left, y, "Email: info@smarthireai.in | Website: www.smarthireai.in")

    # Invoice details box
    box_y = top - 92
    c.rect(left, box_y - 72, right - left, 72, fill=0, stroke=1)

    c.setFont("Helvetica-Bold", 9)
    c.drawString(left + 10, box_y - 18, "Invoice No:")
    c.drawString(left + 10, box_y - 36, "Invoice Date:")
    c.drawString(left + 10, box_y - 54, "Status:")

    c.setFont("Helvetica", 9)
    c.drawString(left + 95, box_y - 18, invoice_clean(invoice["invoice_no"]))
    c.drawString(left + 95, box_y - 36, invoice_clean(invoice["invoice_date"]))
    c.drawString(left + 95, box_y - 54, invoice_clean(invoice["status"], "PAID").upper())

    c.setFont("Helvetica-Bold", 9)
    c.drawString(width / 2 + 20, box_y - 18, "Product:")
    c.drawString(width / 2 + 20, box_y - 36, "Payment Mode:")
    c.drawString(width / 2 + 20, box_y - 54, "Valid Till:")

    c.setFont("Helvetica", 9)
    c.drawString(width / 2 + 115, box_y - 18, "SmartHireAI Payroll Pro")
    c.drawString(width / 2 + 115, box_y - 36, invoice_clean(invoice["payment_mode"], "Razorpay"))
    c.drawString(width / 2 + 115, box_y - 54, invoice_clean(invoice["subscription_end"]))

    # Bill To
    y = box_y - 105
    c.setFont("Helvetica-Bold", 10)
    c.drawString(left, y, "Bill To:")
    y -= 16

    c.setFont("Helvetica-Bold", 9)
    c.drawString(left, y, invoice_clean(invoice["company_name"], "Customer Company"))
    y -= 12

    c.setFont("Helvetica", 9)
    y = draw_wrapped_pdf_text(c, invoice_clean(invoice["customer_address"], ""), left, y, 80, 11, "Helvetica", 9)

    customer_email = invoice_clean(invoice["customer_email"], invoice_clean(invoice["user_email"], ""))
    customer_phone = invoice_clean(invoice["customer_phone"], "")

    if customer_email not in ["", "-"]:
        c.drawString(left, y, f"Email: {customer_email}")
        y -= 12

    if customer_phone not in ["", "-"]:
        c.drawString(left, y, f"Phone: {customer_phone}")
        y -= 12

    # Item table
    table_top = y - 20
    c.rect(left, table_top - 105, right - left, 105, fill=0, stroke=1)

    c.setFont("Helvetica-Bold", 9)
    c.drawString(left + 10, table_top - 18, "Description")
    c.drawRightString(right - 10, table_top - 18, "Amount")

    c.line(left, table_top - 28, right, table_top - 28)

    c.setFont("Helvetica", 9)
    description = f"SmartHireAI Payroll Pro Software - Yearly Subscription ({invoice_clean(invoice['plan_name'])})"
    draw_wrapped_pdf_text(c, description, left + 10, table_top - 46, 72, 11, "Helvetica", 9)

    amount = float(invoice["grand_total"] or invoice["amount"] or 0)
    c.drawRightString(right - 10, table_top - 46, f"Rs. {amount:,.2f}")

    c.line(left, table_top - 78, right, table_top - 78)

    c.setFont("Helvetica-Bold", 9)
    c.drawRightString(right - 130, table_top - 96, "Grand Total:")
    c.drawRightString(right - 10, table_top - 96, f"Rs. {amount:,.2f}")

    # Razorpay details
    y = table_top - 135
    c.setFont("Helvetica-Bold", 10)
    c.drawString(left, y, "Payment Details")
    y -= 16

    c.setFont("Helvetica", 9)
    c.drawString(left, y, f"Razorpay Payment ID: {invoice_clean(invoice['payment_id'])}")
    y -= 13
    c.drawString(left, y, f"Razorpay Order ID: {invoice_clean(invoice['order_id'])}")
    y -= 13
    c.drawString(left, y, f"Subscription Start: {invoice_clean(invoice['subscription_start'])}")
    y -= 13
    c.drawString(left, y, f"Subscription End: {invoice_clean(invoice['subscription_end'])}")

    # Note
    y -= 30
    c.setFont("Helvetica-Oblique", 8.5)
    c.drawString(left, y, "Note: GST is not charged in this invoice. This is a software subscription payment receipt/invoice.")

    # Footer
    c.setFont("Helvetica", 8)
    c.drawCentredString(width / 2, 50, "Thank you for choosing SmartHireAI Payroll Pro.")

    c.save()

    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        UPDATE subscription_invoices
        SET pdf_path = ?
        WHERE id = ?
    """, (pdf_path, invoice_id))
    conn.commit()
    conn.close()

    return pdf_path


def create_subscription_invoice_record(
    cur,
    company_id,
    user_id,
    plan_id,
    plan_name,
    amount,
    payment_id,
    order_id,
    start_date,
    end_date
):
    invoice_no = get_next_global_invoice_no()
    invoice_date = start_date.strftime("%Y-%m-%d")
    created_at = start_date.strftime("%Y-%m-%d %H:%M:%S")

    safe_invoice_no = "".join(ch if ch.isalnum() or ch in ["-", "_"] else "_" for ch in invoice_no)
    pdf_path = os.path.join(INVOICE_FOLDER, f"{safe_invoice_no}.pdf")

    cur.execute("""
        INSERT INTO subscription_invoices (
            company_id,
            user_id,
            product_type,
            invoice_no,
            plan_id,
            plan_name,
            amount,
            discount,
            grand_total,
            payment_id,
            order_id,
            payment_mode,
            status,
            invoice_date,
            subscription_start,
            subscription_end,
            pdf_path,
            created_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        company_id,
        user_id,
        "PAYROLL_PRO",
        invoice_no,
        plan_id,
        plan_name,
        amount,
        0,
        amount,
        payment_id,
        order_id,
        "Razorpay",
        "PAID",
        invoice_date,
        start_date.strftime("%Y-%m-%d"),
        end_date.strftime("%Y-%m-%d"),
        pdf_path,
        created_at
    ))

    return cur.lastrowid



# ---------------------------
# HELPERS
# ---------------------------
def rupee(value):
    return int(round(float(value or 0)))


def money_str(value):
    return str(int(round(float(value or 0))))


def month_only(payroll_month):
    if isinstance(payroll_month, str) and "-" in payroll_month:
        return payroll_month.split("-")[1]
    return str(payroll_month)


def current_company_id():
    return session.get("company_id")


def is_admin_user():
    user_id = session.get("user_id")

    if not user_id:
        return False

    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
        SELECT username
        FROM users
        WHERE id = ?
    """, (user_id,))

    user = cur.fetchone()
    conn.close()

    if not user:
        return False

    username = str(user["username"] or "").strip().lower()

    # ADMIN_USERNAMES list se admin check
    admin_usernames = [str(admin_username).strip().lower() for admin_username in ADMIN_USERNAMES]

    # Local / owner testing ke liye direct allowed users
    owner_usernames = [
        "sai.enterprises7310@gmail.com",
        "admin",
        "sanjay",
        "mansi international"
    ]

    return username in admin_usernames or username in owner_usernames


def create_error_report(row_errors, filename="upload_errors.xlsx"):
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    df_errors = pd.DataFrame({"Error": row_errors})
    file_path = os.path.join(UPLOAD_FOLDER, filename)
    df_errors.to_excel(file_path, index=False)
    return file_path



# ---------------------------
# YEARLY SUBSCRIPTION PLANS
# ---------------------------
YEARLY_PRICING_PLANS = {
    "yearly_6_25": {
        "plan_id": "yearly_6_25",
        "name": "Yearly Starter",
        "price": 9999,
        "duration_days": 365,
        "employee_from": 6,
        "employee_to": 25,
        "label": "6–25 Employees · 1 Year",
        "button_text": "Start Yearly Plan - ₹9,999",
        "badge": "Starter"
    },
    "yearly_26_50": {
        "plan_id": "yearly_26_50",
        "name": "Yearly Growth",
        "price": 14999,
        "duration_days": 365,
        "employee_from": 26,
        "employee_to": 50,
        "label": "26–50 Employees · 1 Year",
        "button_text": "Start Yearly Plan - ₹14,999",
        "badge": "Growth"
    },
    "yearly_51_100": {
        "plan_id": "yearly_51_100",
        "name": "Yearly Business",
        "price": 19999,
        "duration_days": 365,
        "employee_from": 51,
        "employee_to": 100,
        "label": "51–100 Employees · 1 Year",
        "button_text": "Start Yearly Plan - ₹19,999",
        "badge": "Business"
    },
    "yearly_101_200": {
        "plan_id": "yearly_101_200",
        "name": "Yearly Professional",
        "price": 24999,
        "duration_days": 365,
        "employee_from": 101,
        "employee_to": 200,
        "label": "101–200 Employees · 1 Year",
        "button_text": "Start Yearly Plan - ₹24,999",
        "badge": "Professional"
    },
    "yearly_201_400": {
        "plan_id": "yearly_201_400",
        "name": "Yearly Enterprise",
        "price": 29999,
        "duration_days": 365,
        "employee_from": 201,
        "employee_to": 400,
        "label": "201–400 Employees · 1 Year",
        "button_text": "Start Yearly Plan - ₹29,999",
        "badge": "Enterprise"
    },
    "yearly_401_plus": {
        "plan_id": "yearly_401_plus",
        "name": "Yearly Unlimited",
        "price": 39999,
        "duration_days": 365,
        "employee_from": 401,
        "employee_to": None,
        "label": "401+ Employees · 1 Year",
        "button_text": "Start Yearly Plan - ₹39,999",
        "badge": "Unlimited"
    }
}


def get_subscription_pricing_plans():
    return YEARLY_PRICING_PLANS


def get_subscription_plan(plan_id):
    plan_id = str(plan_id or "").strip().lower()
    return YEARLY_PRICING_PLANS.get(plan_id)


def get_subscription_plan_by_name(plan_name):
    """Return the configured yearly plan matching a stored subscription name."""
    normalized_name = str(plan_name or "").strip().lower()

    if not normalized_name:
        return None

    for plan in YEARLY_PRICING_PLANS.values():
        if str(plan.get("name", "")).strip().lower() == normalized_name:
            return plan

    return None


def subscription_plan_supports_employee_count(plan, employee_count):
    """Check whether the employee count belongs to the selected plan slab."""
    if not plan:
        return False

    employee_count = int(employee_count or 0)
    employee_from = int(plan.get("employee_from") or 0)
    employee_to = plan.get("employee_to")

    if employee_count < employee_from:
        return False

    if employee_to is None:
        return True

    return employee_count <= int(employee_to)


def get_active_subscription_row(company_id=None):
    """Return the latest non-expired active subscription row for a company."""
    if company_id is None:
        company_id = current_company_id()

    if not company_id:
        return None

    conn = get_db()
    cur = conn.cursor()
    today = datetime.datetime.now().strftime("%Y-%m-%d")

    cur.execute("""
        SELECT id, company_id, plan_name, status, start_date, end_date
        FROM subscriptions
        WHERE company_id = ?
          AND LOWER(COALESCE(status, '')) = 'active'
          AND end_date IS NOT NULL
          AND TRIM(end_date) != ''
          AND date(end_date) >= date(?)
        ORDER BY date(end_date) DESC, id DESC
        LIMIT 1
    """, (company_id, today))

    row = cur.fetchone()
    conn.close()
    return row


def validate_subscription_plan_for_company(selected_plan, company_id):
    """
    Allow the client to purchase any configured yearly paid plan.

    The Free Plan remains available up to FREE_EMPLOYEE_LIMIT employees.
    Paid-plan buttons are not locked by the current employee count; the
    customer selects the plan they want to purchase.
    """
    if not selected_plan:
        return False, "Invalid yearly plan selected."

    if not company_id:
        return False, "Company not found. Please login again."

    return True, ""

def get_company_employee_count(company_id=None):
    if company_id is None:
        company_id = current_company_id()

    if not company_id:
        return 0

    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
        SELECT COUNT(*) AS count
        FROM employees
        WHERE company_id = ?
    """, (company_id,))

    row = cur.fetchone()
    conn.close()

    return int((row["count"] if row else 0) or 0)


def get_recommended_yearly_plan(employee_count):
    employee_count = int(employee_count or 0)

    if employee_count <= FREE_EMPLOYEE_LIMIT:
        return None

    for plan in YEARLY_PRICING_PLANS.values():
        employee_from = int(plan["employee_from"])
        employee_to = plan["employee_to"]

        if employee_to is None and employee_count >= employee_from:
            return plan

        if employee_to is not None and employee_from <= employee_count <= int(employee_to):
            return plan

    return YEARLY_PRICING_PLANS["yearly_401_plus"]


def ensure_subscription_valid():
    conn = get_db()
    cur = conn.cursor()
    today = datetime.datetime.now().strftime("%Y-%m-%d")

    cur.execute("""
        UPDATE subscriptions
        SET status = 'expired'
        WHERE status = 'active'
        AND end_date IS NOT NULL
        AND end_date < ?
    """, (today,))

    conn.commit()
    conn.close()


def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        ensure_subscription_valid()
        return fn(*args, **kwargs)
    return wrapper


def get_active_plan():
    company_id = current_company_id()

    default_free_plan = {
        "is_pro": False,
        "plan": "FREE",
        "status": "free",
        "start_date": "-",
        "end_date": "-"
    }

    if not company_id:
        return default_free_plan

    conn = None

    try:
        conn = get_db()
        cur = conn.cursor()

        today = datetime.datetime.now().strftime("%Y-%m-%d")

        cur.execute("""
            SELECT
                plan_name,
                status,
                start_date,
                end_date
            FROM subscriptions
            WHERE company_id = ?
              AND LOWER(COALESCE(status, '')) = 'active'
              AND end_date IS NOT NULL
              AND TRIM(end_date) != ''
              AND date(end_date) >= date(?)
            ORDER BY date(end_date) DESC, id DESC
            LIMIT 1
        """, (company_id, today))

        sub = cur.fetchone()

        if not sub:
            return default_free_plan

        plan_name = str(sub["plan_name"] or "PRO").strip().upper()

        # Safety: only real paid/pro plans should unlock PRO features.
        # If accidentally FREE/LIFETIME FREE is saved as active, do not unlock.
        free_plan_names = ["FREE", "FREE PLAN", "LIFETIME FREE", "BASIC"]

        if plan_name in free_plan_names:
            return default_free_plan

        return {
            "is_pro": True,
            "plan": plan_name,
            "status": sub["status"] or "active",
            "start_date": sub["start_date"] or "-",
            "end_date": sub["end_date"] or "-"
        }

    except Exception:
        # Subscription check fail hua to safe side: FREE access.
        return default_free_plan

    finally:
        if conn:
            try:
                conn.close()
            except Exception:
                pass


def is_campaign_free_mode():
    """
    Launch campaign mode.
    When ON, all PRO/download features are temporarily free for users.
    Later, turn this OFF from Render environment variable.
    """
    return os.environ.get("CAMPAIGN_FREE_MODE", "off").lower() == "on"

@app.context_processor
def inject_layout_plan():
    try:
        if session.get("user_id"):
            active_plan = get_active_plan()
            return {
                "layout_active_plan": active_plan
            }
    except Exception:
        pass

    return {
        "layout_active_plan": {
            "is_pro": False,
            "plan": "FREE",
            "status": "free",
            "start_date": "-",
            "end_date": "-"
        }
    }

def get_compliance_settings(company_id):
    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
        SELECT *
        FROM compliance_settings
        WHERE company_id = ?
    """, (company_id,))

    settings = cur.fetchone()

    if not settings:
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        cur.execute("""
            INSERT INTO compliance_settings (
                company_id,

                pf_employee_rate,
                pf_employer_rate,
                pf_wage_ceiling,
                pf_max_deduction,

                esic_employee_rate,
                esic_employer_rate,
                esic_wage_limit,

                gratuity_rate,
                bonus_rate,
                tds_enabled,

                salary_days_policy,
                custom_salary_days,
                count_weekly_off_paid,
                count_paid_leave_paid,
                count_holiday_paid,
                deduct_lop,

                festival_bonus_enabled,
                festival_bonus_month,
                bonus_min_service_days,
                bonus_prorata_enabled,

                updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            company_id,

            12,
            12,
            15000,
            1800,

            0.75,
            3.25,
            21000,

            4.81,
            8.33,
            0,

            "attendance",
            30,
            1,
            1,
            1,
            1,

            0,
            10,
            30,
            1,

            now
        ))

        conn.commit()

        cur.execute("""
            SELECT *
            FROM compliance_settings
            WHERE company_id = ?
        """, (company_id,))

        settings = cur.fetchone()

    conn.close()
    return settings


def is_pro_user():
    if is_campaign_free_mode():
        return True

    if is_admin_user():
        return True

    active_plan = get_active_plan()

    if active_plan and active_plan.get("is_pro") is True:
        return True

    return False


def require_pro_feature(message="Upgrade to yearly plan to use this feature."):
    if is_campaign_free_mode():
        return True

    if is_admin_user():
        return True

    employee_count = get_company_employee_count()
    active_plan = get_active_plan()

    if active_plan and active_plan.get("is_pro") is True:
        plan_config = get_subscription_plan_by_name(active_plan.get("plan"))

        # Legacy/custom paid plans are preserved as unlimited to avoid breaking access.
        if not plan_config:
            return True

        employee_to = plan_config.get("employee_to")

        if employee_to is None or employee_count <= int(employee_to):
            return True

        recommended_plan = get_recommended_yearly_plan(employee_count)
        recommended_text = ""

        if recommended_plan:
            recommended_text = (
                f" Please activate {recommended_plan['name']} at "
                f"₹{recommended_plan['price']}/year."
            )

        flash(
            f"Your current {plan_config['name']} supports up to {employee_to} employees. "
            f"Your workspace currently has {employee_count} employees.{recommended_text}",
            "warning"
        )
        return False

    # 1–5 employees tak all features free.
    if employee_count <= FREE_EMPLOYEE_LIMIT:
        return True

    recommended_plan = get_recommended_yearly_plan(employee_count)
    plan_text = ""
    if recommended_plan:
        plan_text = f" Recommended plan: {recommended_plan['name']} - ₹{recommended_plan['price']}/year."

    flash(
        f"{message} Free plan allows up to {FREE_EMPLOYEE_LIMIT} employees. "
        f"Your workspace has {employee_count} employees.{plan_text}",
        "warning"
    )
    return False


def get_employee_limit():
    # Builder/admin account remains unlimited.
    if is_admin_user():
        return None

    active_plan = get_active_plan()

    if active_plan and active_plan.get("is_pro") is True:
        plan_config = get_subscription_plan_by_name(active_plan.get("plan"))

        # Preserve older/custom paid plans as unlimited.
        if not plan_config:
            return None

        employee_to = plan_config.get("employee_to")
        return None if employee_to is None else int(employee_to)

    return FREE_EMPLOYEE_LIMIT


def can_add_employee():
    # Builder/admin account remains unlimited.
    if is_admin_user():
        return True, ""

    company_id = current_company_id()

    if not company_id:
        return False, "Company not found. Please login again."

    employee_count = get_company_employee_count(company_id)
    active_plan = get_active_plan()

    if active_plan and active_plan.get("is_pro") is True:
        employee_limit = get_employee_limit()

        if employee_limit is None or employee_count < employee_limit:
            return True, ""

        next_employee_count = employee_count + 1
        recommended_plan = get_recommended_yearly_plan(next_employee_count)
        plan_text = ""

        if recommended_plan:
            plan_text = (
                f" Please activate {recommended_plan['name']} at "
                f"₹{recommended_plan['price']}/year."
            )

        return False, (
            f"Your current yearly plan allows up to {employee_limit} employees. "
            f"Upgrade the plan before adding employee number {next_employee_count}.{plan_text}"
        )

    if employee_count >= FREE_EMPLOYEE_LIMIT:
        recommended_plan = get_recommended_yearly_plan(employee_count + 1)
        plan_text = ""
        if recommended_plan:
            plan_text = f" Recommended plan: {recommended_plan['name']} - ₹{recommended_plan['price']}/year."

        return False, (
            f"Free plan allows up to {FREE_EMPLOYEE_LIMIT} employees only. "
            f"Please activate a yearly subscription to add more employees.{plan_text}"
        )

    return True, ""



# ---------------------------
# COMPLIANCE RULES
# ---------------------------
def calculate_professional_tax_maharashtra(gross_salary, gender, payroll_month):
    gross_salary = float(gross_salary or 0)
    gender = str(gender or "male").strip().lower()
    month = month_only(payroll_month)

    # Female employee Maharashtra PT rule
    if gender == "female":
        if gross_salary <= 25000:
            return 0
        return 300 if month == "02" else 200

    # Male employee Maharashtra PT rule
    if gross_salary <= 7500:
        return 0

    if gross_salary <= 10000:
        return 175

    # Male salary above 10000
    return 300 if month == "02" else 200


def calculate_lwf_maharashtra(payroll_month_mm):
    payroll_month_mm = month_only(payroll_month_mm)
    if payroll_month_mm in ["06", "12"]:
        return {"employee": 25.0, "employer": 75.0}
    return {"employee": 0.0, "employer": 0.0}


def calculate_bonus_logic(basic, payroll_month, bonus_rate=0.0833):
    basic = float(basic or 0)

    # Bonus is part of CTC as per compliance setting
    bonus_ctc = rupee(basic * bonus_rate)

    # Regular monthly salary me bonus payout nahi.
    # Festival/Diwali payout separately control kar sakte ho.
    festival_bonus = 0

    return bonus_ctc, festival_bonus


# ============================================================
# SMART HIRE PAYROLL - COMPLIANCE SETTINGS FINAL UPDATE
#
# Replace the existing /compliance-settings route with this
# complete block.
#
# Existing helpers used:
# - get_db()
# - current_company_id()
# - get_compliance_settings(company_id)
# - require_pro_feature(...)
# - login_required
#
# Required import:
# import datetime
# ============================================================


def ensure_compliance_settings_columns():
    """
    Safe migration for older databases.
    Existing compliance data is never deleted.
    """
    conn = get_db()
    cur = conn.cursor()

    cur.execute("PRAGMA table_info(compliance_settings)")

    existing_columns = {
        row["name"] if hasattr(row, "keys") else row[1]
        for row in cur.fetchall()
    }

    required_columns = {
        "festival_bonus_enabled": "INTEGER DEFAULT 0",
        "festival_bonus_month": "INTEGER DEFAULT 10",
        "bonus_min_service_days": "INTEGER DEFAULT 30",
        "bonus_prorata_enabled": "INTEGER DEFAULT 1",
        "updated_at": "TEXT"
    }

    for column_name, column_definition in required_columns.items():
        if column_name not in existing_columns:
            cur.execute(
                f"""
                ALTER TABLE compliance_settings
                ADD COLUMN {column_name} {column_definition}
                """
            )

    conn.commit()
    conn.close()


@app.route("/compliance-settings", methods=["GET", "POST"])
@login_required
def compliance_settings():
    if not require_pro_feature(
        "Upgrade to the yearly plan to manage Compliance Settings."
    ):
        return redirect(url_for("pricing"))

    company_id = current_company_id()

    if not company_id:
        flash(
            "Company not found. Please login again.",
            "danger"
        )
        return redirect(url_for("login"))

    ensure_compliance_settings_columns()

    # Creates the default row when this company has no settings yet.
    get_compliance_settings(company_id)

    allowed_policies = {
        "attendance": "Attendance Based",
        "fixed_26": "Fixed 26 Days",
        "fixed_30": "Fixed 30 Days",
        "calendar": "Calendar Days",
        "custom": "Custom Salary Days"
    }

    month_choices = [
        (1, "January"),
        (2, "February"),
        (3, "March"),
        (4, "April"),
        (5, "May"),
        (6, "June"),
        (7, "July"),
        (8, "August"),
        (9, "September"),
        (10, "October"),
        (11, "November"),
        (12, "December")
    ]

    def checkbox_value(name):
        value = str(
            request.form.get(name, "")
        ).strip().lower()

        return 1 if value in {
            "1",
            "on",
            "true",
            "yes"
        } else 0

    def parse_float(field_name, label, default, errors):
        raw_value = request.form.get(field_name, "")

        if raw_value is None or str(raw_value).strip() == "":
            return float(default)

        try:
            return float(raw_value)
        except (TypeError, ValueError):
            errors.append(
                f"{label} must be a valid number."
            )
            return float(default)

    def parse_int(field_name, label, default, errors):
        raw_value = request.form.get(field_name, "")

        if raw_value is None or str(raw_value).strip() == "":
            return int(default)

        try:
            numeric_value = float(raw_value)

            if not numeric_value.is_integer():
                raise ValueError

            return int(numeric_value)

        except (TypeError, ValueError):
            errors.append(
                f"{label} must be a whole number."
            )
            return int(default)

    if request.method == "POST":
        errors = []

        # PF SETTINGS
        pf_employee_rate = parse_float(
            "pf_employee_rate",
            "PF Employee Rate",
            12,
            errors
        )

        pf_employer_rate = parse_float(
            "pf_employer_rate",
            "PF Employer Rate",
            12,
            errors
        )

        pf_wage_ceiling = parse_float(
            "pf_wage_ceiling",
            "PF Wage Ceiling",
            15000,
            errors
        )

        pf_max_deduction = parse_float(
            "pf_max_deduction",
            "PF Maximum Deduction",
            1800,
            errors
        )

        # ESIC SETTINGS
        esic_employee_rate = parse_float(
            "esic_employee_rate",
            "ESIC Employee Rate",
            0.75,
            errors
        )

        esic_employer_rate = parse_float(
            "esic_employer_rate",
            "ESIC Employer Rate",
            3.25,
            errors
        )

        esic_wage_limit = parse_float(
            "esic_wage_limit",
            "ESIC Wage Limit",
            21000,
            errors
        )

        # OTHER STATUTORY SETTINGS
        gratuity_rate = parse_float(
            "gratuity_rate",
            "Gratuity Rate",
            4.81,
            errors
        )

        bonus_rate = parse_float(
            "bonus_rate",
            "Bonus Rate",
            8.33,
            errors
        )

        tds_enabled = checkbox_value("tds_enabled")

        # SALARY DAYS POLICY
        salary_days_policy = str(
            request.form.get(
                "salary_days_policy",
                "attendance"
            )
        ).strip().lower()

        if salary_days_policy not in allowed_policies:
            errors.append(
                "Please select a valid Salary Days Policy."
            )
            salary_days_policy = "attendance"

        custom_salary_days = parse_float(
            "custom_salary_days",
            "Custom Salary Days",
            30,
            errors
        )

        count_weekly_off_paid = checkbox_value(
            "count_weekly_off_paid"
        )

        count_paid_leave_paid = checkbox_value(
            "count_paid_leave_paid"
        )

        count_holiday_paid = checkbox_value(
            "count_holiday_paid"
        )

        deduct_lop = checkbox_value("deduct_lop")

        # FESTIVAL BONUS SETTINGS
        festival_bonus_enabled = checkbox_value(
            "festival_bonus_enabled"
        )

        festival_bonus_month = parse_int(
            "festival_bonus_month",
            "Festival Bonus Month",
            10,
            errors
        )

        bonus_min_service_days = parse_int(
            "bonus_min_service_days",
            "Bonus Minimum Service Days",
            30,
            errors
        )

        bonus_prorata_enabled = checkbox_value(
            "bonus_prorata_enabled"
        )

        # VALIDATION
        percentage_fields = {
            "PF Employee Rate": pf_employee_rate,
            "PF Employer Rate": pf_employer_rate,
            "ESIC Employee Rate": esic_employee_rate,
            "ESIC Employer Rate": esic_employer_rate,
            "Gratuity Rate": gratuity_rate,
            "Bonus Rate": bonus_rate
        }

        for label, value in percentage_fields.items():
            if value < 0:
                errors.append(f"{label} cannot be negative.")
            elif value > 100:
                errors.append(
                    f"{label} cannot be more than 100%."
                )

        amount_fields = {
            "PF Wage Ceiling": pf_wage_ceiling,
            "PF Maximum Deduction": pf_max_deduction,
            "ESIC Wage Limit": esic_wage_limit
        }

        for label, value in amount_fields.items():
            if value < 0:
                errors.append(f"{label} cannot be negative.")
            elif value > 10000000:
                errors.append(
                    f"{label} amount is too high. Please check."
                )

        if pf_employee_rate > 0 and pf_wage_ceiling <= 0:
            errors.append(
                "PF Wage Ceiling must be greater than 0 "
                "when PF Employee Rate is enabled."
            )

        if pf_employee_rate > 0 and pf_max_deduction <= 0:
            errors.append(
                "PF Maximum Deduction must be greater than 0 "
                "when PF Employee Rate is enabled."
            )

        if (
            (
                esic_employee_rate > 0
                or esic_employer_rate > 0
            )
            and esic_wage_limit <= 0
        ):
            errors.append(
                "ESIC Wage Limit must be greater than 0 "
                "when ESIC rates are enabled."
            )

        if salary_days_policy == "custom":
            if custom_salary_days <= 0:
                errors.append(
                    "Custom Salary Days must be greater than 0."
                )
            elif custom_salary_days > 31:
                errors.append(
                    "Custom Salary Days cannot be more than 31."
                )
        else:
            custom_salary_days = 30

        if not 1 <= festival_bonus_month <= 12:
            errors.append(
                "Festival Bonus Month must be between 1 and 12."
            )

        if bonus_min_service_days < 0:
            errors.append(
                "Bonus Minimum Service Days cannot be negative."
            )
        elif bonus_min_service_days > 3650:
            errors.append(
                "Bonus Minimum Service Days is too high. "
                "Please check."
            )

        if errors:
            flash(
                " ".join(dict.fromkeys(errors)),
                "danger"
            )
            return redirect(url_for("compliance_settings"))

        conn = get_db()
        cur = conn.cursor()

        try:
            now = (
                datetime.datetime.now()
                .strftime("%Y-%m-%d %H:%M:%S")
            )

            cur.execute("""
                UPDATE compliance_settings
                SET
                    pf_employee_rate = ?,
                    pf_employer_rate = ?,
                    pf_wage_ceiling = ?,
                    pf_max_deduction = ?,

                    esic_employee_rate = ?,
                    esic_employer_rate = ?,
                    esic_wage_limit = ?,

                    gratuity_rate = ?,
                    bonus_rate = ?,
                    tds_enabled = ?,

                    salary_days_policy = ?,
                    custom_salary_days = ?,
                    count_weekly_off_paid = ?,
                    count_paid_leave_paid = ?,
                    count_holiday_paid = ?,
                    deduct_lop = ?,

                    festival_bonus_enabled = ?,
                    festival_bonus_month = ?,
                    bonus_min_service_days = ?,
                    bonus_prorata_enabled = ?,

                    updated_at = ?

                WHERE company_id = ?
            """, (
                pf_employee_rate,
                pf_employer_rate,
                pf_wage_ceiling,
                pf_max_deduction,

                esic_employee_rate,
                esic_employer_rate,
                esic_wage_limit,

                gratuity_rate,
                bonus_rate,
                tds_enabled,

                salary_days_policy,
                custom_salary_days,
                count_weekly_off_paid,
                count_paid_leave_paid,
                count_holiday_paid,
                deduct_lop,

                festival_bonus_enabled,
                festival_bonus_month,
                bonus_min_service_days,
                bonus_prorata_enabled,

                now,
                company_id
            ))

            conn.commit()

            flash(
                "Compliance, salary policy and bonus settings "
                "updated successfully.",
                "success"
            )

            return redirect(url_for("compliance_settings"))

        except Exception as error:
            conn.rollback()

            flash(
                "Error saving compliance settings: "
                + str(error),
                "danger"
            )

            return redirect(url_for("compliance_settings"))

        finally:
            conn.close()

    conn = get_db()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT *
            FROM compliance_settings
            WHERE company_id = ?
            LIMIT 1
        """, (company_id,))

        settings = cur.fetchone()

        if not settings:
            flash(
                "Compliance settings could not be loaded.",
                "danger"
            )
            return redirect(url_for("dashboard"))

    except Exception as error:
        flash(
            "Error loading compliance settings: "
            + str(error),
            "danger"
        )
        return redirect(url_for("dashboard"))

    finally:
        conn.close()

    return render_template(
        "compliance_settings.html",
        settings=settings,
        salary_policy_options=allowed_policies,
        month_choices=month_choices
    )


# ---------------------------
# AUTH
# ---------------------------
@app.route("/home")
def home():
    if session.get("user_id"):
        return redirect(url_for("dashboard"))

    return render_template("landing.html")


@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        company_name = request.form["company_name"].strip()
        company_address = request.form.get("company_address", "").strip()
        company_email = request.form.get("company_email", "").strip()
        company_phone = request.form.get("company_phone", "").strip()

        full_name = request.form["full_name"].strip()
        username = request.form["username"].strip()
        password = request.form["password"].strip()

        conn = get_db()
        cur = conn.cursor()

        try:
            cur.execute("""
                INSERT INTO companies (company_name, address, email, phone)
                VALUES (?, ?, ?, ?)
            """, (company_name, company_address, company_email, company_phone))

            company_id = cur.lastrowid

            cur.execute("""
                INSERT INTO users (company_id, full_name, username, password_hash)
                VALUES (?, ?, ?, ?)
            """, (company_id, full_name, username, generate_password_hash(password)))

            conn.commit()
            flash("Registration successful. Please login.")
            return redirect(url_for("login"))

        except Exception as e:
            flash(f"Registration failed: {e}")

        finally:
            conn.close()

    return render_template("register.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"].strip()
        password = request.form["password"].strip()

        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT * FROM users WHERE username = ?", (username,))
        user = cur.fetchone()
        conn.close()

        if user and check_password_hash(user["password_hash"], password):
            session["user_id"] = user["id"]
            session["company_id"] = user["company_id"]
            session["username"] = user["username"]
            return redirect(url_for("dashboard"))

        flash("Invalid username or password.")

    return render_template("login.html")



def save_company_asset(file_storage, company_id, asset_type):
    """
    Saves company logo/signature in uploads/company_assets.
    Allowed: png, jpg, jpeg.
    Returns relative file path or empty string.
    """
    if not file_storage or not file_storage.filename:
        return ""

    allowed_ext = {"png", "jpg", "jpeg"}
    original_name = secure_filename(file_storage.filename)
    ext = original_name.rsplit(".", 1)[-1].lower() if "." in original_name else ""

    if ext not in allowed_ext:
        raise ValueError("Only PNG, JPG and JPEG files are allowed for logo/signature.")

    os.makedirs(COMPANY_ASSET_FOLDER, exist_ok=True)

    filename = f"company_{company_id}_{asset_type}.{ext}"
    file_path = os.path.join(COMPANY_ASSET_FOLDER, filename)
    file_storage.save(file_path)

    return file_path




def get_payslip_company_assets(company_id):
    """
    Payslip PDF ke liye company logo/signature direct DB se fetch karta hai.
    Isse payslip query me column miss hone par bhi logo/signature PDF me aayega.
    """
    defaults = {
        "company_name": "",
        "company_address": "",
        "company_email": "",
        "company_phone": "",
        "logo_path": "",
        "authorized_signature_path": "",
        "authorized_signatory": "",
        "authorized_designation": ""
    }

    try:
        conn = get_db()
        cur = conn.cursor()

        cur.execute("""
            SELECT
                COALESCE(company_name, '') AS company_name,
                COALESCE(address, '') AS company_address,
                COALESCE(email, '') AS company_email,
                COALESCE(phone, '') AS company_phone,
                COALESCE(logo_path, '') AS logo_path,
                COALESCE(authorized_signature_path, '') AS authorized_signature_path,
                COALESCE(authorized_signatory, '') AS authorized_signatory,
                COALESCE(authorized_designation, '') AS authorized_designation
            FROM companies
            WHERE id = ?
        """, (company_id,))

        row = cur.fetchone()
        conn.close()

        if row:
            return dict(row)

    except Exception as e:
        print("get_payslip_company_assets failed:", e)

    return defaults


def draw_payslip_company_header(c, row, width, height):
    """
    Payroll Pro payslip header: company logo + company details.
    Logo/signature data direct companies table se fetch hota hai.
    """
    def clean(value, default=""):
        value = "" if value is None else str(value).strip()
        if not value or value.lower() in ["nan", "none", "null"]:
            return default
        return value

    def short_text(value, max_len=75):
        value = clean(value, "")
        if len(value) > max_len:
            return value[:max_len - 3] + "..."
        return value

    try:
        company_id = row["company_id"]
    except Exception:
        company_id = current_company_id()

    assets = get_payslip_company_assets(company_id)

    company_name = clean(assets.get("company_name"), clean(row["company_name"] if "company_name" in row.keys() else "", "SmartHire Payroll"))
    company_address = clean(assets.get("company_address"), clean(row["company_address"] if "company_address" in row.keys() else "", ""))
    company_email = clean(assets.get("company_email"), clean(row["company_email"] if "company_email" in row.keys() else "", ""))
    company_phone = clean(assets.get("company_phone"), clean(row["company_phone"] if "company_phone" in row.keys() else "", ""))
    logo_path = clean(assets.get("logo_path"), "")

    header_top = height - 28
    header_bottom = height - 112

    c.setStrokeColorRGB(0, 0, 0)
    c.setFillColorRGB(1, 1, 1)
    c.rect(35, header_bottom, width - 70, header_top - header_bottom, fill=1, stroke=1)

    x_text = 48

    if logo_path and os.path.exists(logo_path):
        try:
            img = ImageReader(logo_path)
            c.drawImage(
                img,
                48,
                header_bottom + 15,
                width=58,
                height=50,
                preserveAspectRatio=True,
                mask="auto"
            )
            x_text = 118
        except Exception as e:
            print("Logo draw failed:", e)
            x_text = 48

    c.setFillColorRGB(0, 0, 0)
    c.setFont("Helvetica-Bold", 15)
    c.drawString(x_text, header_top - 24, short_text(company_name.upper(), 42))

    c.setFont("Helvetica-Bold", 9)
    c.drawRightString(width - 48, header_top - 24, "PAYSLIP")

    c.setFont("Helvetica", 8)
    y = header_top - 40

    if company_address:
        c.drawString(x_text, y, short_text(company_address, 72))
        y -= 11

    contact_parts = []
    if company_email:
        contact_parts.append(company_email)
    if company_phone:
        contact_parts.append(company_phone)

    if contact_parts:
        c.drawString(x_text, y, short_text(" | ".join(contact_parts), 72))



@app.route("/company-asset/<asset_type>")
@login_required
def company_asset(asset_type):
    company_id = current_company_id()

    if not company_id:
        return "Company not found", 404

    allowed_assets = {
        "logo": "logo_path",
        "signature": "authorized_signature_path"
    }

    if asset_type not in allowed_assets:
        return "Invalid asset", 404

    conn = get_db()
    cur = conn.cursor()

    cur.execute(f"""
        SELECT {allowed_assets[asset_type]} AS asset_path
        FROM companies
        WHERE id = ?
    """, (company_id,))

    row = cur.fetchone()
    conn.close()

    if not row:
        return "Asset not found", 404

    asset_path = str(row["asset_path"] or "").strip()

    if not asset_path or not os.path.exists(asset_path):
        return "Asset not found", 404

    return send_file(asset_path)


# ============================================================
# PAYROLL PRO - UPDATED COMPANY PROFILE
# Company details + PT/PAN/GST + logo + authorized signature
# ============================================================

def ensure_payroll_pro_company_profile_columns():
    """
    companies table me Payroll Pro company profile ke optional columns
    safely add karta hai. Existing data delete nahi hota.
    """
    conn = get_db()
    cur = conn.cursor()

    try:
        cur.execute("PRAGMA table_info(companies)")
        existing_columns = {
            row["name"] if hasattr(row, "keys") else row[1]
            for row in cur.fetchall()
        }

        columns_to_add = {
            "pt_registration_no": "TEXT DEFAULT ''",
            "pan_no": "TEXT DEFAULT ''",
            "gst_no": "TEXT DEFAULT ''",
            "support_email": "TEXT DEFAULT ''",
            "support_phone": "TEXT DEFAULT ''",
            "logo_path": "TEXT DEFAULT ''",
            "authorized_signature_path": "TEXT DEFAULT ''",
            "authorized_signatory": "TEXT DEFAULT ''",
            "authorized_designation": "TEXT DEFAULT ''",
            "overtime_multiplier": "REAL DEFAULT 1",
            "working_days_policy": "TEXT DEFAULT 'attendance'"
        }

        for column_name, column_definition in columns_to_add.items():
            if column_name not in existing_columns:
                cur.execute(
                    f"ALTER TABLE companies "
                    f"ADD COLUMN {column_name} {column_definition}"
                )

        conn.commit()

    finally:
        conn.close()


@app.route("/company-profile-asset/<asset_type>")
@login_required
def company_profile_asset(asset_type):
    """
    Current logged-in company ka logo/signature safely browser me show karta hai.
    """
    company_id = current_company_id()

    if not company_id:
        return "Company not found", 404

    asset_columns = {
        "logo": "logo_path",
        "signature": "authorized_signature_path"
    }

    column_name = asset_columns.get(str(asset_type).strip().lower())

    if not column_name:
        return "Invalid asset type", 404

    conn = get_db()
    cur = conn.cursor()

    try:
        cur.execute(
            f"""
            SELECT COALESCE({column_name}, '') AS asset_path
            FROM companies
            WHERE id = ?
            LIMIT 1
            """,
            (company_id,)
        )

        row = cur.fetchone()

    finally:
        conn.close()

    if not row:
        return "Asset not found", 404

    asset_path = str(row["asset_path"] or "").strip()

    if not asset_path or not os.path.exists(asset_path):
        return "Asset not found", 404

    return send_file(asset_path)


@app.route("/company-profile", methods=["GET", "POST"])
@login_required
def company_profile():
    ensure_payroll_pro_company_profile_columns()

    company_id = current_company_id()

    if not company_id:
        flash("Company not found. Please login again.", "danger")
        return redirect(url_for("login"))

    def clean_text(value, default=""):
        value = str(value or "").strip()

        if not value or value.lower() in ["nan", "none", "null"]:
            return default

        return value

    def valid_email(value):
        value = clean_text(value)

        if not value:
            return True

        return (
            "@" in value
            and "." in value.split("@")[-1]
            and " " not in value
        )

    conn = get_db()
    cur = conn.cursor()

    if request.method == "POST":
        try:
            company_name = clean_text(
                request.form.get("company_name")
            )
            address = clean_text(
                request.form.get("address")
            )
            email = clean_text(
                request.form.get("email")
            ).lower()
            phone = clean_text(
                request.form.get("phone")
            )

            support_email = clean_text(
                request.form.get("support_email")
            ).lower()
            support_phone = clean_text(
                request.form.get("support_phone")
            )

            pt_registration_no = clean_text(
                request.form.get("pt_registration_no")
            ).upper()
            pan_no = clean_text(
                request.form.get("pan_no")
            ).upper()
            gst_no = clean_text(
                request.form.get("gst_no")
            ).upper()

            authorized_signatory = clean_text(
                request.form.get("authorized_signatory")
            )
            authorized_designation = clean_text(
                request.form.get("authorized_designation"),
                "HR / Authorized Signatory"
            )

            try:
                overtime_multiplier = float(
                    request.form.get("overtime_multiplier", 1)
                )
            except Exception:
                overtime_multiplier = 1

            if overtime_multiplier not in [1, 2]:
                overtime_multiplier = 1

            working_days_policy = clean_text(
                request.form.get(
                    "working_days_policy",
                    "attendance"
                ),
                "attendance"
            ).lower()

            allowed_working_days_policies = [
                "attendance",
                "fixed_26",
                "fixed_30",
                "calendar"
            ]

            if working_days_policy not in allowed_working_days_policies:
                working_days_policy = "attendance"

            errors = []

            if not company_name:
                errors.append("Company name is required.")

            if len(company_name) > 150:
                errors.append("Company name is too long.")

            if len(address) > 500:
                errors.append("Company address is too long.")

            if not valid_email(email):
                errors.append("Please enter a valid company email.")

            if not valid_email(support_email):
                errors.append("Please enter a valid support email.")

            if len(email) > 150:
                errors.append("Company email is too long.")

            if len(support_email) > 150:
                errors.append("Support email is too long.")

            if len(phone) > 30:
                errors.append("Company phone number is too long.")

            if len(support_phone) > 30:
                errors.append("Support phone number is too long.")

            if pan_no and len(pan_no) != 10:
                errors.append("PAN number must contain 10 characters.")

            if gst_no and len(gst_no) != 15:
                errors.append("GST number must contain 15 characters.")

            if errors:
                flash(" ".join(errors), "danger")
                return redirect(url_for("company_profile"))

            # Existing asset paths
            cur.execute("""
                SELECT
                    COALESCE(logo_path, '') AS logo_path,
                    COALESCE(
                        authorized_signature_path,
                        ''
                    ) AS authorized_signature_path
                FROM companies
                WHERE id = ?
                LIMIT 1
            """, (company_id,))

            existing = cur.fetchone()

            logo_path = (
                str(existing["logo_path"] or "").strip()
                if existing
                else ""
            )
            authorized_signature_path = (
                str(
                    existing["authorized_signature_path"]
                    or ""
                ).strip()
                if existing
                else ""
            )

            logo_file = request.files.get("company_logo")
            signature_file = request.files.get(
                "authorized_signature"
            )

            if logo_file and logo_file.filename:
                logo_path = save_company_asset(
                    logo_file,
                    company_id,
                    "logo"
                )

            if signature_file and signature_file.filename:
                authorized_signature_path = save_company_asset(
                    signature_file,
                    company_id,
                    "signature"
                )

            cur.execute("""
                UPDATE companies
                SET
                    company_name = ?,
                    address = ?,
                    email = ?,
                    phone = ?,

                    support_email = ?,
                    support_phone = ?,

                    pt_registration_no = ?,
                    pan_no = ?,
                    gst_no = ?,

                    authorized_signatory = ?,
                    authorized_designation = ?,

                    overtime_multiplier = ?,
                    working_days_policy = ?,

                    logo_path = ?,
                    authorized_signature_path = ?
                WHERE id = ?
            """, (
                company_name,
                address,
                email,
                phone,

                support_email,
                support_phone,

                pt_registration_no,
                pan_no,
                gst_no,

                authorized_signatory,
                authorized_designation,

                overtime_multiplier,
                working_days_policy,

                logo_path,
                authorized_signature_path,

                company_id
            ))

            conn.commit()

            session["company_name"] = company_name
            session["company_email"] = email
            session["company_phone"] = phone

            flash(
                "Company profile updated successfully.",
                "success"
            )
            return redirect(url_for("company_profile"))

        except ValueError as e:
            conn.rollback()
            flash(str(e), "danger")
            return redirect(url_for("company_profile"))

        except Exception as e:
            conn.rollback()
            print("Company profile update error:", e)
            flash(
                f"Error updating company profile: {str(e)}",
                "danger"
            )
            return redirect(url_for("company_profile"))

        finally:
            conn.close()

    try:
        cur.execute("""
            SELECT
                company_name,
                address,
                email,
                phone,

                COALESCE(support_email, '') AS support_email,
                COALESCE(support_phone, '') AS support_phone,

                COALESCE(
                    pt_registration_no,
                    ''
                ) AS pt_registration_no,
                COALESCE(pan_no, '') AS pan_no,
                COALESCE(gst_no, '') AS gst_no,

                COALESCE(
                    authorized_signatory,
                    ''
                ) AS authorized_signatory,
                COALESCE(
                    authorized_designation,
                    ''
                ) AS authorized_designation,

                COALESCE(
                    overtime_multiplier,
                    1
                ) AS overtime_multiplier,
                COALESCE(
                    working_days_policy,
                    'attendance'
                ) AS working_days_policy,

                COALESCE(logo_path, '') AS logo_path,
                COALESCE(
                    authorized_signature_path,
                    ''
                ) AS authorized_signature_path

            FROM companies
            WHERE id = ?
            LIMIT 1
        """, (company_id,))

        company = cur.fetchone()

        if not company:
            flash(
                "Company profile not found. Please login again.",
                "danger"
            )
            return redirect(url_for("login"))

        logo_exists = bool(
            str(company["logo_path"] or "").strip()
            and os.path.exists(
                str(company["logo_path"]).strip()
            )
        )
        signature_exists = bool(
            str(
                company["authorized_signature_path"]
                or ""
            ).strip()
            and os.path.exists(
                str(
                    company["authorized_signature_path"]
                ).strip()
            )
        )

        return render_template(
            "company_profile.html",
            company=company,
            company_logo_url=(
                url_for(
                    "company_profile_asset",
                    asset_type="logo"
                )
                if logo_exists
                else ""
            ),
            authorized_signature_url=(
                url_for(
                    "company_profile_asset",
                    asset_type="signature"
                )
                if signature_exists
                else ""
            )
        )

    except Exception as e:
        print("Company profile load error:", e)
        flash(
            f"Error loading company profile: {str(e)}",
            "danger"
        )
        return redirect(url_for("dashboard"))

    finally:
        conn.close()


# ---------------------------
# DASHBOARD / PAYMENT
# ---------------------------
@app.route("/")
@login_required
def dashboard():
    company_id = current_company_id()
    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
        SELECT company_name
        FROM companies
        WHERE id = ?
    """, (company_id,))
    company = cur.fetchone()

    if not company:
        conn.close()
        session.clear()
        flash("Company record not found. Please login again.")
        return redirect(url_for("login"))

    cur.execute("""
        SELECT COUNT(*) AS count
        FROM employees
        WHERE company_id = ?
    """, (company_id,))
    employee_count = cur.fetchone()["count"] or 0

    cur.execute("""
        SELECT COUNT(*) AS count
        FROM attendance
        WHERE company_id = ?
    """, (company_id,))
    attendance_count = cur.fetchone()["count"] or 0

    cur.execute("""
        SELECT COUNT(*) AS pending_leaves
        FROM leave_requests
        WHERE company_id = ?
          AND status = 'Pending'
    """, (company_id,))
    pending_leaves = cur.fetchone()["pending_leaves"] or 0

    # Latest payroll month
    cur.execute("""
        SELECT month
        FROM payroll_history
        WHERE company_id = ?
          AND is_current = 1
        GROUP BY month
        ORDER BY month DESC
        LIMIT 1
    """, (company_id,))
    latest_month_row = cur.fetchone()
    latest_payroll_month = latest_month_row["month"] if latest_month_row else "-"

    # Payroll count - latest/current records only
    cur.execute("""
        SELECT COUNT(*) AS count
        FROM payroll_history
        WHERE company_id = ?
          AND is_current = 1
    """, (company_id,))
    payroll_count = cur.fetchone()["count"] or 0

    # Overall current payroll totals
    cur.execute("""
        SELECT
            COALESCE(SUM(gross), 0) AS total_gross,
            COALESCE(SUM(overtime_amount), 0) AS total_overtime_amount,
            COALESCE(SUM(festival_bonus), 0) AS total_festival_bonus,
            COALESCE(SUM(bonus_ctc), 0) AS total_bonus_ctc,

            COALESCE(SUM(total_deductions), 0) AS total_deductions,
            COALESCE(SUM(net_pay), 0) AS total_net_pay,

            COALESCE(SUM(pf_employer), 0) AS total_pf_employer,
            COALESCE(SUM(esi_employer), 0) AS total_esi_employer,
            COALESCE(SUM(gratuity), 0) AS total_gratuity,
            COALESCE(SUM(lwf_employer), 0) AS total_lwf_employer,

            COALESCE(SUM(monthly_ctc), 0) AS total_monthly_ctc,
            COALESCE(SUM(annual_ctc), 0) AS total_annual_ctc
        FROM payroll_history
        WHERE company_id = ?
          AND is_current = 1
    """, (company_id,))
    payroll_totals = cur.fetchone()

    total_gross = round(float(payroll_totals["total_gross"] or 0))
    total_overtime_amount = round(float(payroll_totals["total_overtime_amount"] or 0))
    total_festival_bonus = round(float(payroll_totals["total_festival_bonus"] or 0))
    total_bonus_ctc = round(float(payroll_totals["total_bonus_ctc"] or 0))

    total_deductions = round(float(payroll_totals["total_deductions"] or 0))
    total_net_pay = round(float(payroll_totals["total_net_pay"] or 0))

    total_pf_employer = round(float(payroll_totals["total_pf_employer"] or 0))
    total_esi_employer = round(float(payroll_totals["total_esi_employer"] or 0))
    total_gratuity = round(float(payroll_totals["total_gratuity"] or 0))
    total_lwf_employer = round(float(payroll_totals["total_lwf_employer"] or 0))

    total_employer_cost = round(
        total_pf_employer
        + total_esi_employer
        + total_gratuity
        + total_lwf_employer
    )

    total_monthly_ctc = round(float(payroll_totals["total_monthly_ctc"] or 0))
    total_annual_ctc = round(float(payroll_totals["total_annual_ctc"] or 0))

    # Latest month payroll totals
    latest_month_gross = 0
    latest_month_net_pay = 0
    latest_month_ctc = 0
    latest_month_employees = 0

    if latest_payroll_month != "-":
        cur.execute("""
            SELECT
                COUNT(*) AS employees,
                COALESCE(SUM(gross), 0) AS gross,
                COALESCE(SUM(net_pay), 0) AS net_pay,
                COALESCE(SUM(monthly_ctc), 0) AS monthly_ctc
            FROM payroll_history
            WHERE company_id = ?
              AND month = ?
              AND is_current = 1
        """, (company_id, latest_payroll_month))

        latest_totals = cur.fetchone()

        latest_month_employees = latest_totals["employees"] or 0
        latest_month_gross = round(float(latest_totals["gross"] or 0))
        latest_month_net_pay = round(float(latest_totals["net_pay"] or 0))
        latest_month_ctc = round(float(latest_totals["monthly_ctc"] or 0))

    # Subscription / active plan
    active_plan = get_active_plan()

    cur.execute("""
        SELECT plan_name, status, start_date, end_date
        FROM subscriptions
        WHERE company_id = ?
          AND status = 'active'
        ORDER BY id DESC
        LIMIT 1
    """, (company_id,))
    subscription = cur.fetchone()

    if subscription:
        plan_name = subscription["plan_name"] or active_plan["plan"]
        subscription_status = subscription["status"] or "active"
        subscription_end_date = subscription["end_date"] or active_plan["end_date"] or "-"
    else:
        plan_name = active_plan["plan"]
        subscription_status = active_plan.get("status", "free")
        subscription_end_date = active_plan["end_date"] or "-"

    # Chart data - last 12 payroll months
    cur.execute("""
        SELECT
            month,
            COALESCE(SUM(gross), 0) AS gross,
            COALESCE(SUM(net_pay), 0) AS net_pay,
            COALESCE(SUM(total_deductions), 0) AS deductions,
            COALESCE(SUM(monthly_ctc), 0) AS monthly_ctc
        FROM payroll_history
        WHERE company_id = ?
          AND is_current = 1
        GROUP BY month
        ORDER BY month ASC
        LIMIT 12
    """, (company_id,))
    chart_rows = cur.fetchall()

    chart_labels = []
    chart_gross = []
    chart_net_pay = []
    chart_deductions = []
    chart_monthly_ctc = []

    for row in chart_rows:
        chart_labels.append(row["month"])
        chart_gross.append(round(float(row["gross"] or 0)))
        chart_net_pay.append(round(float(row["net_pay"] or 0)))
        chart_deductions.append(round(float(row["deductions"] or 0)))
        chart_monthly_ctc.append(round(float(row["monthly_ctc"] or 0)))

    conn.close()

    return render_template(
        "dashboard.html",
        company_name=company["company_name"],

        employee_count=employee_count,
        attendance_count=attendance_count,
        payroll_count=payroll_count,
        pending_leaves=pending_leaves,

        latest_payroll_month=latest_payroll_month,
        latest_month_employees=latest_month_employees,
        latest_month_gross=latest_month_gross,
        latest_month_net_pay=latest_month_net_pay,
        latest_month_ctc=latest_month_ctc,

        total_gross=total_gross,
        total_overtime_amount=total_overtime_amount,
        total_festival_bonus=total_festival_bonus,
        total_bonus_ctc=total_bonus_ctc,

        total_deductions=total_deductions,
        total_net_pay=total_net_pay,

        total_pf_employer=total_pf_employer,
        total_esi_employer=total_esi_employer,
        total_gratuity=total_gratuity,
        total_lwf_employer=total_lwf_employer,
        total_employer_cost=total_employer_cost,

        total_monthly_ctc=total_monthly_ctc,
        total_annual_ctc=total_annual_ctc,

        chart_labels=chart_labels,
        chart_gross=chart_gross,
        chart_net_pay=chart_net_pay,
        chart_deductions=chart_deductions,
        chart_monthly_ctc=chart_monthly_ctc,

        active_plan=active_plan,
        plan_name=plan_name,
        subscription_status=subscription_status,
        subscription_end_date=subscription_end_date,

        now=datetime.datetime.now()
    )


@app.route("/pricing")
@login_required
def pricing():
    company_id = current_company_id()

    plans = get_subscription_pricing_plans()
    active_plan = get_active_plan()
    employee_count = get_company_employee_count(company_id)
    recommended_plan = get_recommended_yearly_plan(employee_count)

    conn = get_db()
    cur = conn.cursor()

    today = datetime.datetime.now().strftime("%Y-%m-%d")

    cur.execute("""
        SELECT plan_name, status, start_date, end_date
        FROM subscriptions
        WHERE company_id = ?
          AND status = 'active'
          AND date(end_date) >= date(?)
        ORDER BY date(end_date) DESC, id DESC
        LIMIT 1
    """, (company_id, today))

    active_subscription = cur.fetchone()
    conn.close()

    return render_template(
        "pricing.html",
        razorpay_key_id=RAZORPAY_KEY_ID,
        plans=plans,
        active_subscription=active_subscription,
        active_plan=active_plan,
        payments_enabled=PAYMENTS_ENABLED,
        employee_count=employee_count,
        free_employee_limit=FREE_EMPLOYEE_LIMIT,
        recommended_plan=recommended_plan
    )


@app.route("/create-order", methods=["POST"])
@login_required
def create_order():
    if not PAYMENTS_ENABLED:
        return jsonify({
            "status": "failed",
            "message": "Online payment is currently unavailable. Please contact SmartHireAI for activation."
        }), 403

    if not razorpay_client:
        return jsonify({
            "status": "failed",
            "message": "Online payment is currently unavailable. Please contact SmartHireAI support."
        }), 503

    company_id = current_company_id()

    if not company_id:
        return jsonify({
            "status": "failed",
            "message": "Company not found. Please login again."
        }), 400

    data = request.get_json(silent=True) or {}
    plan_id = str(data.get("plan_id", "")).strip().lower()
    selected_plan = get_subscription_plan(plan_id)

    if not selected_plan:
        return jsonify({
            "status": "failed",
            "message": "Invalid yearly plan selected."
        }), 400

    plan_allowed, validation_message = validate_subscription_plan_for_company(
        selected_plan,
        company_id
    )

    if not plan_allowed:
        return jsonify({
            "status": "failed",
            "message": validation_message
        }), 400

    amount_paise = int(selected_plan["price"]) * 100

    try:
        receipt = f"pp_{company_id}_{int(time.time())}"[:40]

        order = razorpay_client.order.create({
            "amount": amount_paise,
            "currency": "INR",
            "payment_capture": 1,
            "receipt": receipt,
            "notes": {
                "product_type": "PAYROLL_PRO",
                "company_id": str(company_id),
                "plan_id": selected_plan["plan_id"],
                "plan_name": selected_plan["name"]
            }
        })

        order["key_id"] = RAZORPAY_KEY_ID
        order["plan_id"] = selected_plan["plan_id"]
        order["plan_name"] = selected_plan["name"]
        order["amount_rupees"] = selected_plan["price"]

        return jsonify(order)

    except Exception as e:
        print("Razorpay order creation failed:", e)
        return jsonify({
            "status": "failed",
            "message": "Payment order could not be created. Please try again or contact support."
        }), 500


@app.route("/payment-success", methods=["POST"])
@login_required
def payment_success():
    if not PAYMENTS_ENABLED:
        return "Online payment is currently unavailable.", 403

    if not razorpay_client:
        return "Online payment service is not configured.", 503

    user_id = session.get("user_id")
    company_id = current_company_id()

    if not company_id:
        return "Company not found. Please login again.", 400

    payment_id = request.form.get("razorpay_payment_id", "").strip()
    order_id = request.form.get("razorpay_order_id", "").strip()
    signature = request.form.get("razorpay_signature", "").strip()
    plan_id = request.form.get("plan_id", "").strip().lower()

    selected_plan = get_subscription_plan(plan_id)

    if not selected_plan:
        return "Invalid yearly plan selected", 400

    if not payment_id or not order_id or not signature:
        return "Payment verification details missing", 400

    # Razorpay signature verification.
    try:
        razorpay_client.utility.verify_payment_signature({
            "razorpay_order_id": order_id,
            "razorpay_payment_id": payment_id,
            "razorpay_signature": signature
        })
    except Exception:
        return "Payment signature verification failed", 400

    # Do not trust amount, plan or company details received from the browser.
    try:
        razorpay_order = razorpay_client.order.fetch(order_id)
        razorpay_payment = razorpay_client.payment.fetch(payment_id)
    except Exception as e:
        print("Razorpay verification fetch failed:", e)
        return "Payment details could not be verified with Razorpay", 400

    expected_amount_paise = int(selected_plan["price"]) * 100
    order_notes = razorpay_order.get("notes") or {}

    verification_errors = []

    if str(razorpay_order.get("id") or "") != order_id:
        verification_errors.append("order mismatch")

    if int(razorpay_order.get("amount") or 0) != expected_amount_paise:
        verification_errors.append("order amount mismatch")

    if str(razorpay_order.get("currency") or "").upper() != "INR":
        verification_errors.append("order currency mismatch")

    if str(order_notes.get("product_type") or "").upper() != "PAYROLL_PRO":
        verification_errors.append("product mismatch")

    if str(order_notes.get("company_id") or "") != str(company_id):
        verification_errors.append("company mismatch")

    if str(order_notes.get("plan_id") or "").strip().lower() != plan_id:
        verification_errors.append("plan mismatch")

    if str(razorpay_payment.get("id") or "") != payment_id:
        verification_errors.append("payment mismatch")

    if str(razorpay_payment.get("order_id") or "") != order_id:
        verification_errors.append("payment order mismatch")

    if int(razorpay_payment.get("amount") or 0) != expected_amount_paise:
        verification_errors.append("payment amount mismatch")

    if str(razorpay_payment.get("currency") or "").upper() != "INR":
        verification_errors.append("payment currency mismatch")

    if str(razorpay_payment.get("status") or "").lower() != "captured":
        verification_errors.append("payment is not captured")

    if verification_errors:
        print("Payment verification failed:", ", ".join(verification_errors))
        return "Payment verification failed", 400

    plan_allowed, validation_message = validate_subscription_plan_for_company(
        selected_plan,
        company_id
    )

    if not plan_allowed:
        return validation_message, 400

    amount = selected_plan["price"]
    plan_name = selected_plan["name"]
    duration_days = int(selected_plan["duration_days"])

    payment_time = datetime.datetime.now()
    subscription_start = payment_time

    conn = get_db()
    cur = conn.cursor()

    try:
        # Razorpay payment IDs are globally unique.
        cur.execute("""
            SELECT id, company_id
            FROM payments
            WHERE payment_id = ?
            ORDER BY id DESC
            LIMIT 1
        """, (payment_id,))

        existing_payment = cur.fetchone()

        if existing_payment:
            if int(existing_payment["company_id"] or 0) == int(company_id):
                return "success", 200
            return "Payment reference is already associated with another company", 409

        # Preserve unused days when a customer renews before the current expiry date.
        cur.execute("""
            SELECT end_date
            FROM subscriptions
            WHERE company_id = ?
              AND LOWER(COALESCE(status, '')) = 'active'
              AND end_date IS NOT NULL
              AND TRIM(end_date) != ''
            ORDER BY date(end_date) DESC, id DESC
            LIMIT 1
        """, (company_id,))

        current_subscription = cur.fetchone()

        if current_subscription and current_subscription["end_date"]:
            try:
                current_end_date = datetime.datetime.strptime(
                    current_subscription["end_date"],
                    "%Y-%m-%d"
                ).date()

                if current_end_date >= payment_time.date():
                    next_start_date = current_end_date + datetime.timedelta(days=1)
                    subscription_start = datetime.datetime.combine(
                        next_start_date,
                        payment_time.time().replace(microsecond=0)
                    )
            except (TypeError, ValueError):
                subscription_start = payment_time

        subscription_end = subscription_start + datetime.timedelta(days=duration_days)

        # Deactivate previous subscription records before activating the new one.
        cur.execute("""
            UPDATE subscriptions
            SET status = 'inactive'
            WHERE company_id = ?
              AND status = 'active'
        """, (company_id,))

        cur.execute("""
            INSERT INTO subscriptions (
                company_id,
                plan_name,
                status,
                start_date,
                end_date
            )
            VALUES (?, ?, ?, ?, ?)
        """, (
            company_id,
            plan_name,
            "active",
            subscription_start.strftime("%Y-%m-%d"),
            subscription_end.strftime("%Y-%m-%d")
        ))

        cur.execute("""
            INSERT INTO payments (
                company_id,
                user_id,
                amount,
                payment_id,
                order_id,
                status,
                created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            company_id,
            user_id,
            amount,
            payment_id,
            order_id,
            "success",
            payment_time.strftime("%Y-%m-%d %H:%M:%S")
        ))

        invoice_id = create_subscription_invoice_record(
            cur=cur,
            company_id=company_id,
            user_id=user_id,
            plan_id=plan_id,
            plan_name=plan_name,
            amount=amount,
            payment_id=payment_id,
            order_id=order_id,
            start_date=subscription_start,
            end_date=subscription_end
        )

        conn.commit()

        try:
            generate_subscription_invoice_pdf(invoice_id)
        except Exception as invoice_error:
            # Payment and subscription remain saved. The download route can regenerate the PDF.
            print("Invoice PDF generation failed:", invoice_error)

        return "success", 200

    except Exception as e:
        conn.rollback()
        print("Payment activation failed:", e)
        return "Payment activation failed. Please contact support with your payment reference.", 500

    finally:
        conn.close()


@app.route("/payments")
@login_required
def payments():
    company_id = current_company_id()

    if not company_id:
        flash("Company not found. Please login again.", "danger")
        return redirect(url_for("login"))

    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            p.id,
            p.amount,
            p.payment_id,
            p.order_id,
            p.status,
            p.created_at,
            si.id AS invoice_id,
            si.invoice_no,
            si.status AS invoice_status
        FROM payments p
        LEFT JOIN subscription_invoices si
            ON si.company_id = p.company_id
           AND si.payment_id = p.payment_id
        WHERE p.company_id = ?
        ORDER BY p.id DESC
    """, (company_id,))

    data = cur.fetchall()

    total_payments = len(data)

    successful_payments = 0
    pending_payments = 0
    failed_payments = 0
    total_paid_amount = 0

    for row in data:
        status = str(row["status"] or "").lower()
        amount = float(row["amount"] or 0)

        if status in ["paid", "success", "successful", "captured"]:
            successful_payments += 1
            total_paid_amount += amount
        elif status in ["pending", "created", "authorized"]:
            pending_payments += 1
        else:
            failed_payments += 1

    total_paid_amount = round(total_paid_amount)

    active_plan = get_active_plan()

    today = datetime.datetime.now().strftime("%Y-%m-%d")

    cur.execute("""
        SELECT plan_name, status, start_date, end_date
        FROM subscriptions
        WHERE company_id = ?
          AND status = 'active'
          AND date(end_date) >= date(?)
        ORDER BY date(end_date) DESC, id DESC
        LIMIT 1
    """, (company_id, today))

    active_subscription = cur.fetchone()

    conn.close()

    return render_template(
        "payments.html",
        data=data,
        total_paid_amount=total_paid_amount,
        total_payments=total_payments,
        successful_payments=successful_payments,
        pending_payments=pending_payments,
        failed_payments=failed_payments,
        active_plan=active_plan,
        active_subscription=active_subscription,
        campaign_free_mode=is_campaign_free_mode()
    )



@app.route("/download-invoice/<int:invoice_id>")
@login_required
def download_invoice(invoice_id):
    company_id = current_company_id()

    if not company_id:
        flash("Company not found. Please login again.", "danger")
        return redirect(url_for("login"))

    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, invoice_no, pdf_path
        FROM subscription_invoices
        WHERE id = ?
          AND company_id = ?
    """, (invoice_id, company_id))

    invoice = cur.fetchone()
    conn.close()

    if not invoice:
        flash("Invoice not found.", "danger")
        return redirect(url_for("payments"))

    pdf_path = invoice["pdf_path"]

    if not pdf_path or not os.path.exists(pdf_path):
        try:
            pdf_path = generate_subscription_invoice_pdf(invoice_id)
        except Exception as e:
            flash(f"Invoice PDF could not be generated: {str(e)}", "danger")
            return redirect(url_for("payments"))

    invoice_no = invoice_clean(invoice["invoice_no"], f"invoice_{invoice_id}")
    download_name = f"{invoice_no}.pdf"

    return send_file(
        pdf_path,
        as_attachment=True,
        download_name=download_name
    )



# ============================================================
# PAYROLL PRO - EMPLOYEE MASTER PERSONAL / WHATSAPP FIELDS
# Adds:
#   mobile_no, aadhaar_no, pan_no, address, email_id
#
# All five fields are OPTIONAL during employee upload.
# mobile_no will later be used for WhatsApp payslip sending.
#
# Required existing imports:
# import os
# import re
# import pandas as pd
# from flask import request, redirect, url_for, flash, session, send_file
# from werkzeug.utils import secure_filename
# ============================================================


def ensure_employee_personal_columns():
    """
    employees table me new optional columns safely add karta hai.
    Existing employee data delete nahi hota.
    """
    conn = get_db()
    cur = conn.cursor()

    try:
        cur.execute("PRAGMA table_info(employees)")
        existing_columns = {
            row["name"] if hasattr(row, "keys") else row[1]
            for row in cur.fetchall()
        }

        columns_to_add = {
            "mobile_no": "TEXT DEFAULT ''",
            "aadhaar_no": "TEXT DEFAULT ''",
            "pan_no": "TEXT DEFAULT ''",
            "address": "TEXT DEFAULT ''",
            "email_id": "TEXT DEFAULT ''"
        }

        for column_name, column_definition in columns_to_add.items():
            if column_name not in existing_columns:
                cur.execute(
                    f"ALTER TABLE employees "
                    f"ADD COLUMN {column_name} {column_definition}"
                )

        conn.commit()

    finally:
        conn.close()


def clean_employee_text_value(value, default=""):
    """
    Excel numeric-looking text jaise 9876543210.0 ko clean text banata hai.
    """
    if value is None:
        return default

    try:
        if pd.isna(value):
            return default
    except Exception:
        pass

    text = str(value).strip()

    if not text or text.lower() in ["nan", "none", "null"]:
        return default

    if text.endswith(".0") and text[:-2].replace("-", "").isdigit():
        text = text[:-2]

    return text


def clean_employee_digits(value):
    text = clean_employee_text_value(value)
    return "".join(ch for ch in text if ch.isdigit())


def normalize_employee_mobile(value):
    """
    Indian employee mobile ko 10-digit format me save karta hai.
    91 prefix diya ho to remove karke last 10 digits save hoti hain.
    """
    digits = clean_employee_digits(value)

    if not digits:
        return ""

    if len(digits) == 12 and digits.startswith("91"):
        digits = digits[2:]

    if len(digits) == 11 and digits.startswith("0"):
        digits = digits[1:]

    return digits


def valid_employee_email(value):
    email = clean_employee_text_value(value).lower()

    if not email:
        return True

    return (
        "@" in email
        and "." in email.split("@")[-1]
        and " " not in email
    )


@app.route("/upload-employees", methods=["GET", "POST"])
@login_required
def upload_employees():
    ensure_employee_personal_columns()

    if request.method == "POST":
        company_id = current_company_id()

        if not company_id:
            flash("Company not found. Please login again.", "danger")
            return redirect(url_for("login"))

        if "file" not in request.files:
            flash("Please select a file.", "warning")
            return redirect(url_for("upload_employees"))

        file = request.files["file"]

        if not file or file.filename == "":
            flash("Please select a file.", "warning")
            return redirect(url_for("upload_employees"))

        filename = file.filename.lower()

        if not (
            filename.endswith(".csv")
            or filename.endswith(".xlsx")
        ):
            flash(
                "Only CSV or Excel (.xlsx) file allowed.",
                "danger"
            )
            return redirect(url_for("upload_employees"))

        path = os.path.join(
            UPLOAD_FOLDER,
            secure_filename(file.filename)
        )
        file.save(path)

        conn = None

        try:
            if filename.endswith(".csv"):
                df = pd.read_csv(
                    path,
                    dtype=str,
                    keep_default_na=False
                )
            else:
                df = pd.read_excel(
                    path,
                    engine="openpyxl",
                    dtype=str,
                    keep_default_na=False
                )

            # Clean column names
            df.columns = [
                str(col).strip().lower()
                .replace(" ", "_")
                .replace("-", "_")
                for col in df.columns
            ]

            # Common alternate headings support
            column_aliases = {
                "employee_code": "emp_code",
                "employee_id": "emp_code",
                "emp_id": "emp_code",

                "name": "employee_name",
                "emp_name": "employee_name",

                "designation": "role",

                "mobile": "mobile_no",
                "mobile_number": "mobile_no",
                "whatsapp_no": "mobile_no",
                "whatsapp_number": "mobile_no",
                "phone_number": "mobile_no",
                "contact_no": "mobile_no",

                "aadhar_no": "aadhaar_no",
                "aadhar_number": "aadhaar_no",
                "aadhaar_number": "aadhaar_no",

                "pan_number": "pan_no",
                "employee_pan": "pan_no",

                "email": "email_id",
                "email_address": "email_id",
                "employee_email": "email_id"
            }

            for source_column, target_column in column_aliases.items():
                if (
                    source_column in df.columns
                    and target_column not in df.columns
                ):
                    df[target_column] = df[source_column]

            required_columns = [
                "emp_code",
                "employee_name",
                "role",
                "monthly_salary"
            ]

            missing_columns = [
                col for col in required_columns
                if col not in df.columns
            ]

            if missing_columns:
                session["error_report"] = create_error_report(
                    [
                        f"Missing required column: {col}"
                        for col in missing_columns
                    ],
                    "employee_upload_errors.xlsx"
                )

                flash(
                    "Upload failed. Required columns are missing. "
                    "Please download the error report.",
                    "danger"
                )
                return redirect(url_for("upload_employees"))

            row_errors = []

            cleaned_emp_codes = df["emp_code"].apply(
                lambda x: clean_text(x)
            )

            duplicate_emp_codes = cleaned_emp_codes[
                cleaned_emp_codes.duplicated()
                & (cleaned_emp_codes != "")
            ].unique()

            for emp_code in duplicate_emp_codes:
                row_errors.append(
                    f"Duplicate emp_code found in file: {emp_code}"
                )

            valid_genders = ["male", "female", "other"]
            valid_tax_regimes = ["old", "new"]

            for index, row in df.iterrows():
                row_no = index + 2

                emp_code = clean_text(row.get("emp_code"))
                employee_name = clean_text(
                    row.get("employee_name")
                )
                role = clean_text(row.get("role"))
                gender = clean_text(
                    row.get("gender"),
                    "male"
                ).lower()
                tax_regime = clean_text(
                    row.get("tax_regime"),
                    "new"
                ).lower()

                monthly_salary = clean_float(
                    row.get("monthly_salary"),
                    0
                )
                special_allowance = clean_float(
                    row.get("special_allowance"),
                    0
                )
                other_annual_deductions = clean_float(
                    row.get("other_annual_deductions"),
                    0
                )

                mobile_no = normalize_employee_mobile(
                    row.get("mobile_no")
                )
                aadhaar_no = clean_employee_digits(
                    row.get("aadhaar_no")
                )
                pan_no = clean_employee_text_value(
                    row.get("pan_no")
                ).upper()
                employee_address = clean_employee_text_value(
                    row.get("address")
                )
                email_id = clean_employee_text_value(
                    row.get("email_id")
                ).lower()

                if emp_code == "":
                    row_errors.append(
                        f"Row {row_no}: emp_code missing"
                    )

                if employee_name == "":
                    row_errors.append(
                        f"Row {row_no}: employee_name missing"
                    )

                if role == "":
                    row_errors.append(
                        f"Row {row_no}: role missing"
                    )

                if monthly_salary <= 0:
                    row_errors.append(
                        f"Row {row_no}: monthly_salary "
                        "must be greater than 0"
                    )

                if gender and gender not in valid_genders:
                    row_errors.append(
                        f"Row {row_no}: gender must be "
                        "male, female, or other"
                    )

                if (
                    tax_regime
                    and tax_regime not in valid_tax_regimes
                ):
                    row_errors.append(
                        f"Row {row_no}: tax_regime must be "
                        "old or new"
                    )

                if special_allowance < 0:
                    row_errors.append(
                        f"Row {row_no}: special_allowance "
                        "cannot be negative"
                    )

                if other_annual_deductions < 0:
                    row_errors.append(
                        f"Row {row_no}: "
                        "other_annual_deductions "
                        "cannot be negative"
                    )

                # New optional field validations
                if mobile_no:
                    if (
                        len(mobile_no) != 10
                        or mobile_no[0] not in "6789"
                    ):
                        row_errors.append(
                            f"Row {row_no}: mobile_no must be "
                            "a valid 10-digit Indian mobile number"
                        )

                if aadhaar_no and len(aadhaar_no) != 12:
                    row_errors.append(
                        f"Row {row_no}: aadhaar_no must "
                        "contain 12 digits"
                    )

                if pan_no and not re.fullmatch(
                    r"[A-Z]{5}[0-9]{4}[A-Z]",
                    pan_no
                ):
                    row_errors.append(
                        f"Row {row_no}: pan_no must be "
                        "in format ABCDE1234F"
                    )

                if len(employee_address) > 500:
                    row_errors.append(
                        f"Row {row_no}: address is too long "
                        "(maximum 500 characters)"
                    )

                if email_id and not valid_employee_email(email_id):
                    row_errors.append(
                        f"Row {row_no}: email_id is invalid"
                    )

            if row_errors:
                session["error_report"] = create_error_report(
                    row_errors,
                    "employee_upload_errors.xlsx"
                )

                flash(
                    "Upload failed. Please download the error "
                    "report and fix the file.",
                    "danger"
                )
                return redirect(url_for("upload_employees"))

            conn = get_db()
            cur = conn.cursor()

            # Free plan upload limit check:
            # only NEW employees count honge.
            if not is_admin_user():
                active_plan = get_active_plan()

                if not active_plan.get("is_pro"):
                    cur.execute("""
                        SELECT emp_code
                        FROM employees
                        WHERE company_id = ?
                    """, (company_id,))

                    existing_emp_codes = {
                        clean_text(row["emp_code"]).lower()
                        for row in cur.fetchall()
                    }

                    upload_emp_codes = {
                        clean_text(emp_code).lower()
                        for emp_code in df["emp_code"].tolist()
                        if clean_text(emp_code) != ""
                    }

                    new_emp_codes = (
                        upload_emp_codes
                        - existing_emp_codes
                    )

                    existing_count = len(existing_emp_codes)
                    new_upload_count = len(new_emp_codes)

                    if (
                        existing_count + new_upload_count
                        > FREE_EMPLOYEE_LIMIT
                    ):
                        flash(
                            f"Free plan allows up to "
                            f"{FREE_EMPLOYEE_LIMIT} employees only. "
                            f"You already have {existing_count} "
                            f"employee(s), and this file will add "
                            f"{new_upload_count} new employee(s). "
                            "Please upgrade to PRO for unlimited "
                            "employees.",
                            "warning"
                        )
                        return redirect(url_for("pricing"))

            # Get leave policy for default balances
            cur.execute("""
                SELECT
                    casual_leave_limit,
                    sick_leave_limit,
                    paid_leave_limit
                FROM leave_policy
                WHERE company_id = ?
                LIMIT 1
            """, (company_id,))

            leave_policy = cur.fetchone()

            if leave_policy:
                default_casual_leave = float(
                    leave_policy["casual_leave_limit"] or 6
                )
                default_sick_leave = float(
                    leave_policy["sick_leave_limit"] or 6
                )
                default_paid_leave = float(
                    leave_policy["paid_leave_limit"] or 12
                )
            else:
                default_casual_leave = 6
                default_sick_leave = 6
                default_paid_leave = 12

            added_count = 0
            updated_count = 0

            for _, row in df.iterrows():
                emp_code = clean_text(row.get("emp_code"))
                employee_name = clean_text(
                    row.get("employee_name")
                )
                role = clean_text(row.get("role"))
                department = clean_text(
                    row.get("department"),
                    "General"
                )
                gender = clean_text(
                    row.get("gender"),
                    "male"
                ).lower()
                monthly_salary = clean_float(
                    row.get("monthly_salary"),
                    0
                )
                tax_regime = clean_text(
                    row.get("tax_regime"),
                    "new"
                ).lower()
                other_annual_deductions = clean_float(
                    row.get("other_annual_deductions"),
                    0
                )
                special_allowance = clean_float(
                    row.get("special_allowance"),
                    0
                )

                uan_no = clean_employee_text_value(
                    row.get("uan_no")
                )
                esic_no = clean_employee_text_value(
                    row.get("esic_no")
                )
                bank_name = clean_text(row.get("bank_name"))
                account_no = clean_employee_text_value(
                    row.get("account_no")
                )
                ifsc_code = clean_text(
                    row.get("ifsc_code")
                ).upper()

                mobile_no = normalize_employee_mobile(
                    row.get("mobile_no")
                )
                aadhaar_no = clean_employee_digits(
                    row.get("aadhaar_no")
                )
                pan_no = clean_employee_text_value(
                    row.get("pan_no")
                ).upper()
                employee_address = clean_employee_text_value(
                    row.get("address")
                )
                email_id = clean_employee_text_value(
                    row.get("email_id")
                ).lower()

                if gender not in valid_genders:
                    gender = "male"

                if tax_regime not in valid_tax_regimes:
                    tax_regime = "new"

                if department == "":
                    department = "General"

                cur.execute("""
                    SELECT id
                    FROM employees
                    WHERE company_id = ?
                      AND emp_code = ?
                """, (company_id, emp_code))

                existing_employee = cur.fetchone()

                if existing_employee:
                    updated_count += 1

                    cur.execute("""
                        UPDATE employees
                        SET
                            employee_name = ?,
                            role = ?,
                            department = ?,
                            gender = ?,
                            monthly_salary = ?,
                            tax_regime = ?,
                            other_annual_deductions = ?,
                            special_allowance = ?,
                            uan_no = ?,
                            esic_no = ?,
                            bank_name = ?,
                            account_no = ?,
                            ifsc_code = ?,
                            mobile_no = ?,
                            aadhaar_no = ?,
                            pan_no = ?,
                            address = ?,
                            email_id = ?
                        WHERE company_id = ?
                          AND emp_code = ?
                    """, (
                        employee_name,
                        role,
                        department,
                        gender,
                        monthly_salary,
                        tax_regime,
                        other_annual_deductions,
                        special_allowance,
                        uan_no,
                        esic_no,
                        bank_name,
                        account_no,
                        ifsc_code,
                        mobile_no,
                        aadhaar_no,
                        pan_no,
                        employee_address,
                        email_id,
                        company_id,
                        emp_code
                    ))

                else:
                    added_count += 1

                    cur.execute("""
                        INSERT INTO employees
                        (
                            company_id,
                            emp_code,
                            employee_name,
                            role,
                            department,
                            gender,
                            monthly_salary,
                            tax_regime,
                            other_annual_deductions,
                            special_allowance,
                            uan_no,
                            esic_no,
                            bank_name,
                            account_no,
                            ifsc_code,
                            mobile_no,
                            aadhaar_no,
                            pan_no,
                            address,
                            email_id
                        )
                        VALUES (
                            ?, ?, ?, ?, ?, ?,
                            ?, ?, ?, ?, ?, ?,
                            ?, ?, ?, ?, ?, ?,
                            ?, ?
                        )
                    """, (
                        company_id,
                        emp_code,
                        employee_name,
                        role,
                        department,
                        gender,
                        monthly_salary,
                        tax_regime,
                        other_annual_deductions,
                        special_allowance,
                        uan_no,
                        esic_no,
                        bank_name,
                        account_no,
                        ifsc_code,
                        mobile_no,
                        aadhaar_no,
                        pan_no,
                        employee_address,
                        email_id
                    ))

                # Create leave balance only for new employees /
                # missing balance
                cur.execute("""
                    SELECT id
                    FROM leave_balances
                    WHERE company_id = ?
                      AND emp_code = ?
                """, (company_id, emp_code))

                existing_balance = cur.fetchone()

                if not existing_balance:
                    cur.execute("""
                        INSERT INTO leave_balances
                        (
                            company_id,
                            emp_code,
                            casual_leave,
                            sick_leave,
                            paid_leave,
                            used_leave
                        )
                        VALUES (?, ?, ?, ?, ?, 0)
                    """, (
                        company_id,
                        emp_code,
                        default_casual_leave,
                        default_sick_leave,
                        default_paid_leave
                    ))

            conn.commit()
            session.pop("error_report", None)

            flash(
                "Employee master uploaded successfully. "
                f"Added: {added_count}, "
                f"Updated: {updated_count}. "
                "Leave balances checked/created.",
                "success"
            )

            return redirect(url_for("employees_list"))

        except Exception as e:
            if conn:
                conn.rollback()

            print("Employee upload error:", e)

            flash(
                f"Upload failed: {str(e)}",
                "danger"
            )
            return redirect(url_for("upload_employees"))

        finally:
            if conn:
                conn.close()

    return render_template("upload_employees.html")


@app.route("/download-employee-sample")
@login_required
def download_employee_sample():
    """
    Updated Payroll Pro employee-master sample:
    mobile_no, aadhaar_no, pan_no, address, email_id included.
    """
    ensure_employee_personal_columns()

    data = {
        "emp_code": [
            "EMP001",
            "EMP002"
        ],
        "employee_name": [
            "Rahul Sharma",
            "Priya Verma"
        ],
        "role": [
            "Payroll Executive",
            "HR Executive"
        ],
        "department": [
            "Accounts",
            "HR"
        ],
        "gender": [
            "male",
            "female"
        ],
        "mobile_no": [
            "9876543210",
            "9123456789"
        ],
        "email_id": [
            "rahul.sharma@example.com",
            "priya.verma@example.com"
        ],
        "aadhaar_no": [
            "123412341234",
            "567856785678"
        ],
        "pan_no": [
            "ABCDE1234F",
            "FGHIJ5678K"
        ],
        "address": [
            "Nagpur, Maharashtra",
            "Pune, Maharashtra"
        ],
        "monthly_salary": [
            25000,
            32000
        ],
        "tax_regime": [
            "new",
            "new"
        ],
        "other_annual_deductions": [
            0,
            0
        ],
        "special_allowance": [
            2000,
            3000
        ],
        "uan_no": [
            "123456789012",
            "222233334444"
        ],
        "esic_no": [
            "9876543210",
            "111122223333"
        ],
        "bank_name": [
            "HDFC Bank",
            "ICICI Bank"
        ],
        "account_no": [
            "50100234567890",
            "123456789012"
        ],
        "ifsc_code": [
            "HDFC0001234",
            "ICIC0005678"
        ]
    }

    df = pd.DataFrame(data)

    file_path = os.path.join(
        UPLOAD_FOLDER,
        "employee_master_sample.xlsx"
    )

    instructions = pd.DataFrame({
        "Field": [
            "emp_code",
            "employee_name",
            "role",
            "department",
            "gender",
            "mobile_no",
            "email_id",
            "aadhaar_no",
            "pan_no",
            "address",
            "monthly_salary",
            "tax_regime",
            "other_annual_deductions",
            "special_allowance",
            "uan_no",
            "esic_no",
            "bank_name",
            "account_no",
            "ifsc_code"
        ],
        "Required": [
            "Yes",
            "Yes",
            "Yes",
            "Optional",
            "Optional",
            "Optional",
            "Optional",
            "Optional",
            "Optional",
            "Optional",
            "Yes",
            "Optional",
            "Optional",
            "Optional",
            "Optional",
            "Optional",
            "Optional",
            "Optional",
            "Optional"
        ],
        "Example": [
            "EMP001",
            "Rahul Sharma",
            "Payroll Executive",
            "Accounts",
            "male / female / other",
            "9876543210",
            "employee@example.com",
            "123412341234",
            "ABCDE1234F",
            "Nagpur, Maharashtra",
            "25000",
            "old / new",
            "0",
            "2000",
            "123456789012",
            "9876543210",
            "HDFC Bank",
            "50100234567890",
            "HDFC0001234"
        ],
        "Notes": [
            "Must be unique for each employee. Existing emp_code updates employee details.",
            "Employee full name.",
            "Designation or job role.",
            "Blank department is treated as General.",
            "Use male, female, or other. Blank defaults to male.",
            "Optional on upload. Required only for WhatsApp payslip sending. Enter 10 digits without +91.",
            "Optional employee email address.",
            "Optional. Enter exactly 12 digits and keep the column as text.",
            "Optional. Use uppercase format ABCDE1234F.",
            "Optional employee residential address, maximum 500 characters.",
            "Monthly salary must be greater than 0.",
            "Use old or new. Blank defaults to new.",
            "Use 0 if not applicable.",
            "Use 0 if not applicable.",
            "Keep as text to avoid number formatting issues.",
            "Keep as text to avoid number formatting issues.",
            "Employee bank name.",
            "Keep as text to avoid number formatting issues.",
            "Use uppercase IFSC code."
        ]
    })

    with pd.ExcelWriter(
        file_path,
        engine="openpyxl"
    ) as writer:
        df.to_excel(
            writer,
            index=False,
            sheet_name="Employee Master"
        )
        instructions.to_excel(
            writer,
            index=False,
            sheet_name="Instructions"
        )

        workbook = writer.book

        from openpyxl.styles import (
            Font,
            PatternFill,
            Border,
            Side,
            Alignment
        )
        from openpyxl.utils import get_column_letter

        header_fill = PatternFill(
            start_color="2563EB",
            end_color="2563EB",
            fill_type="solid"
        )
        header_font = Font(
            color="FFFFFF",
            bold=True
        )
        thin_border = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB")
        )

        text_columns = {
            "emp_code",
            "mobile_no",
            "aadhaar_no",
            "pan_no",
            "uan_no",
            "esic_no",
            "account_no",
            "ifsc_code"
        }

        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )
                cell.border = thin_border

            for row_cells in ws.iter_rows():
                for cell in row_cells:
                    cell.border = thin_border
                    cell.alignment = Alignment(
                        vertical="top",
                        wrap_text=True
                    )

            for column_index, column_cells in enumerate(
                ws.columns,
                start=1
            ):
                max_length = max(
                    len(str(cell.value or ""))
                    for cell in column_cells
                )

                width = min(max(max_length + 2, 12), 34)
                ws.column_dimensions[
                    get_column_letter(column_index)
                ].width = width

        master_ws = workbook["Employee Master"]

        for column_index, header_cell in enumerate(
            master_ws[1],
            start=1
        ):
            if str(header_cell.value) in text_columns:
                for row_number in range(
                    2,
                    master_ws.max_row + 1
                ):
                    master_ws.cell(
                        row=row_number,
                        column=column_index
                    ).number_format = "@"

    return send_file(
        file_path,
        as_attachment=True,
        download_name="employee_master_sample.xlsx"
    )


@app.route("/employees")
@login_required
def employees_list():
    # New employee personal fields database me available rahenge
    ensure_employee_personal_columns()

    department = request.args.get("department", "").strip()
    search = request.args.get("search", "").strip()

    company_id = current_company_id()

    if not company_id:
        flash("Company not found. Please login again.", "danger")
        return redirect(url_for("login"))

    conn = get_db()
    cur = conn.cursor()

    query = """
        SELECT *
        FROM employees
        WHERE company_id = ?
    """

    params = [company_id]

    if department:
        query += " AND department = ?"
        params.append(department)

    if search:
        query += """
            AND (
                emp_code LIKE ?
                OR employee_name LIKE ?
                OR role LIKE ?
                OR department LIKE ?
                OR COALESCE(mobile_no, '') LIKE ?
                OR COALESCE(email_id, '') LIKE ?
                OR COALESCE(pan_no, '') LIKE ?
            )
        """

        search_value = f"%{search}%"

        params.extend([
            search_value,
            search_value,
            search_value,
            search_value,
            search_value,
            search_value,
            search_value
        ])

    query += " ORDER BY id DESC"

    cur.execute(query, tuple(params))
    employees = cur.fetchall()

    cur.execute("""
        SELECT DISTINCT department
        FROM employees
        WHERE company_id = ?
          AND department IS NOT NULL
          AND department != ''
        ORDER BY department
    """, (company_id,))

    departments = cur.fetchall()

    cur.execute("""
        SELECT COUNT(*) AS total_employees
        FROM employees
        WHERE company_id = ?
    """, (company_id,))

    total_employees = (
        cur.fetchone()["total_employees"] or 0
    )

    cur.execute("""
        SELECT COUNT(DISTINCT department) AS total_departments
        FROM employees
        WHERE company_id = ?
          AND department IS NOT NULL
          AND department != ''
    """, (company_id,))

    total_departments = (
        cur.fetchone()["total_departments"] or 0
    )

    cur.execute("""
        SELECT COALESCE(SUM(monthly_salary), 0) AS total_salary
        FROM employees
        WHERE company_id = ?
    """, (company_id,))

    total_salary = round(
        float(cur.fetchone()["total_salary"] or 0)
    )

    cur.execute("""
        SELECT COUNT(*) AS male_count
        FROM employees
        WHERE company_id = ?
          AND LOWER(COALESCE(gender, '')) = 'male'
    """, (company_id,))

    male_count = cur.fetchone()["male_count"] or 0

    cur.execute("""
        SELECT COUNT(*) AS female_count
        FROM employees
        WHERE company_id = ?
          AND LOWER(COALESCE(gender, '')) = 'female'
    """, (company_id,))

    female_count = cur.fetchone()["female_count"] or 0

    # WhatsApp payslip readiness summary
    cur.execute("""
        SELECT COUNT(*) AS whatsapp_ready_count
        FROM employees
        WHERE company_id = ?
          AND LENGTH(
                REPLACE(
                    REPLACE(
                        REPLACE(
                            COALESCE(mobile_no, ''),
                            ' ',
                            ''
                        ),
                        '-',
                        ''
                    ),
                    '+91',
                    ''
                )
              ) = 10
    """, (company_id,))

    whatsapp_ready_count = (
        cur.fetchone()["whatsapp_ready_count"] or 0
    )

    mobile_missing_count = max(
        total_employees - whatsapp_ready_count,
        0
    )

    conn.close()

    return render_template(
        "employees.html",
        employees=employees,
        departments=departments,

        selected_department=department,
        search=search,

        total_employees=total_employees,
        total_departments=total_departments,
        total_salary=total_salary,
        male_count=male_count,
        female_count=female_count,

        whatsapp_ready_count=whatsapp_ready_count,
        mobile_missing_count=mobile_missing_count
    )


# ============================================================
# PAYROLL PRO - EDIT EMPLOYEE ROUTE
# Paste this block in app.py after employees_list() route.
# Required import at top: import re
# ============================================================

@app.route("/employees/<path:emp_code>/edit", methods=["GET", "POST"])
@login_required
def edit_employee(emp_code):
    ensure_employee_personal_columns()

    company_id = current_company_id()

    if not company_id:
        flash("Company not found. Please login again.", "danger")
        return redirect(url_for("login"))

    def local_clean_text(value, default=""):
        if value is None:
            return default
        text = str(value).strip()
        if not text or text.lower() in ["nan", "none", "null"]:
            return default
        if text.endswith(".0") and text[:-2].replace("-", "").isdigit():
            text = text[:-2]
        return text

    def local_clean_float(value, default=0):
        try:
            text = local_clean_text(value)
            return float(text) if text else float(default)
        except Exception:
            return float(default)

    def local_digits(value):
        return "".join(ch for ch in local_clean_text(value) if ch.isdigit())

    def local_mobile(value):
        digits = local_digits(value)
        if not digits:
            return ""
        if len(digits) == 12 and digits.startswith("91"):
            digits = digits[2:]
        if len(digits) == 11 and digits.startswith("0"):
            digits = digits[1:]
        return digits

    def local_valid_email(value):
        email = local_clean_text(value).lower()
        if not email:
            return True
        return "@" in email and "." in email.split("@")[-1] and " " not in email

    conn = get_db()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT *
            FROM employees
            WHERE company_id = ?
              AND emp_code = ?
            LIMIT 1
        """, (company_id, emp_code))

        employee = cur.fetchone()

        if not employee:
            flash("Employee record not found.", "danger")
            return redirect(url_for("employees_list"))

        if request.method == "POST":
            employee_name = local_clean_text(request.form.get("employee_name"))
            role = local_clean_text(request.form.get("role"))
            department = local_clean_text(request.form.get("department"), "General")
            gender = local_clean_text(request.form.get("gender"), "male").lower()

            monthly_salary = local_clean_float(request.form.get("monthly_salary"), 0)
            tax_regime = local_clean_text(request.form.get("tax_regime"), "new").lower()
            other_annual_deductions = local_clean_float(request.form.get("other_annual_deductions"), 0)
            special_allowance = local_clean_float(request.form.get("special_allowance"), 0)

            uan_no = local_clean_text(request.form.get("uan_no"))
            esic_no = local_clean_text(request.form.get("esic_no"))
            bank_name = local_clean_text(request.form.get("bank_name"))
            account_no = local_clean_text(request.form.get("account_no"))
            ifsc_code = local_clean_text(request.form.get("ifsc_code")).upper()

            mobile_no = local_mobile(request.form.get("mobile_no"))
            email_id = local_clean_text(request.form.get("email_id")).lower()
            aadhaar_no = local_digits(request.form.get("aadhaar_no"))
            pan_no = local_clean_text(request.form.get("pan_no")).upper()
            employee_address = local_clean_text(request.form.get("address"))

            errors = []

            if not employee_name:
                errors.append("Employee name is required.")
            if len(employee_name) > 150:
                errors.append("Employee name is too long.")
            if not role:
                errors.append("Role / designation is required.")
            if len(role) > 150:
                errors.append("Role / designation is too long.")
            if len(department) > 150:
                errors.append("Department name is too long.")
            if gender not in ["male", "female", "other"]:
                errors.append("Gender must be male, female, or other.")
            if monthly_salary <= 0:
                errors.append("Monthly salary must be greater than 0.")
            if tax_regime not in ["old", "new"]:
                errors.append("Tax regime must be old or new.")
            if other_annual_deductions < 0:
                errors.append("Other annual deductions cannot be negative.")
            if special_allowance < 0:
                errors.append("Special allowance cannot be negative.")
            if mobile_no and (len(mobile_no) != 10 or mobile_no[0] not in "6789"):
                errors.append("Mobile number must be a valid 10-digit Indian mobile number.")
            if aadhaar_no and len(aadhaar_no) != 12:
                errors.append("Aadhaar number must contain exactly 12 digits.")
            if pan_no and not re.fullmatch(r"[A-Z]{5}[0-9]{4}[A-Z]", pan_no):
                errors.append("PAN number must be in format ABCDE1234F.")
            if email_id and not local_valid_email(email_id):
                errors.append("Employee email address is invalid.")
            if len(employee_address) > 500:
                errors.append("Address is too long. Maximum 500 characters allowed.")

            if errors:
                for error in errors:
                    flash(error, "danger")

                form_data = {
                    "emp_code": emp_code,
                    "employee_name": employee_name,
                    "role": role,
                    "department": department,
                    "gender": gender,
                    "monthly_salary": monthly_salary,
                    "tax_regime": tax_regime,
                    "other_annual_deductions": other_annual_deductions,
                    "special_allowance": special_allowance,
                    "uan_no": uan_no,
                    "esic_no": esic_no,
                    "bank_name": bank_name,
                    "account_no": account_no,
                    "ifsc_code": ifsc_code,
                    "mobile_no": mobile_no,
                    "email_id": email_id,
                    "aadhaar_no": aadhaar_no,
                    "pan_no": pan_no,
                    "address": employee_address
                }

                return render_template("edit_employee.html", employee=form_data)

            cur.execute("""
                UPDATE employees
                SET
                    employee_name = ?,
                    role = ?,
                    department = ?,
                    gender = ?,
                    monthly_salary = ?,
                    tax_regime = ?,
                    other_annual_deductions = ?,
                    special_allowance = ?,
                    uan_no = ?,
                    esic_no = ?,
                    bank_name = ?,
                    account_no = ?,
                    ifsc_code = ?,
                    mobile_no = ?,
                    email_id = ?,
                    aadhaar_no = ?,
                    pan_no = ?,
                    address = ?
                WHERE company_id = ?
                  AND emp_code = ?
            """, (
                employee_name,
                role,
                department,
                gender,
                monthly_salary,
                tax_regime,
                other_annual_deductions,
                special_allowance,
                uan_no,
                esic_no,
                bank_name,
                account_no,
                ifsc_code,
                mobile_no,
                email_id,
                aadhaar_no,
                pan_no,
                employee_address,
                company_id,
                emp_code
            ))

            conn.commit()
            flash(f"Employee {employee_name} updated successfully.", "success")
            return redirect(url_for("employees_list", search=emp_code))

        return render_template("edit_employee.html", employee=employee)

    except Exception as e:
        conn.rollback()
        print("Edit employee error:", e)
        flash(f"Employee update failed: {str(e)}", "danger")
        return redirect(url_for("employees_list"))

    finally:
        conn.close()


# ---------------------------
# ATTENDANCE
# ---------------------------
@app.route("/upload-attendance", methods=["GET", "POST"])
@login_required
def upload_attendance():
    if request.method == "POST":
        company_id = current_company_id()

        if not company_id:
            flash("Company not found. Please login again.", "danger")
            return redirect(url_for("login"))

        if "file" not in request.files:
            flash("Please select a file.", "warning")
            return redirect(url_for("upload_attendance"))

        file = request.files["file"]

        if not file or file.filename.strip() == "":
            flash("Please select a file.", "warning")
            return redirect(url_for("upload_attendance"))

        filename = file.filename.lower()

        if not (filename.endswith(".csv") or filename.endswith(".xlsx")):
            flash("Only CSV or Excel (.xlsx) file allowed.", "danger")
            return redirect(url_for("upload_attendance"))

        path = os.path.join(UPLOAD_FOLDER, secure_filename(file.filename))
        file.save(path)

        conn = None

        try:
            if filename.endswith(".csv"):
                df = pd.read_csv(path)
            else:
                df = pd.read_excel(path, engine="openpyxl")

            # Clean column names
            df.columns = [str(col).strip().lower() for col in df.columns]

            required_columns = [
                "emp_code",
                "month",
                "working_days",
                "present_days",
                "weekly_off",
                "paid_leave",
                "holiday",
                "lop_days",
                "overtime_hours"
            ]

            # Required column validation - helper function dependency removed
            missing_columns = [
                col for col in required_columns
                if col not in df.columns
            ]

            if missing_columns:
                session["error_report"] = create_error_report(
                    [f"Missing required column: {col}" for col in missing_columns],
                    "attendance_upload_errors.xlsx"
                )
                flash("Upload failed. Required columns are missing. Please download the error report.", "danger")
                return redirect(url_for("upload_attendance"))

            row_errors = []

            df["emp_code_clean"] = df["emp_code"].apply(lambda x: clean_text(x))
            df["month_clean"] = df["month"].apply(lambda x: clean_text(x))

            duplicate_rows = df[
                df.duplicated(subset=["emp_code_clean", "month_clean"], keep=False)
                & (df["emp_code_clean"] != "")
                & (df["month_clean"] != "")
            ]

            if not duplicate_rows.empty:
                duplicate_pairs = duplicate_rows[["emp_code_clean", "month_clean"]].drop_duplicates()

                for _, dup in duplicate_pairs.iterrows():
                    row_errors.append(
                        f"Duplicate attendance found for emp_code {dup['emp_code_clean']} in month {dup['month_clean']}"
                    )

            for index, row in df.iterrows():
                row_no = index + 2

                emp_code = clean_text(row.get("emp_code"))
                month = clean_text(row.get("month"))

                working_days = clean_float(row.get("working_days"), -1)
                present_days = clean_float(row.get("present_days"), -1)
                weekly_off = clean_float(row.get("weekly_off"), -1)
                paid_leave = clean_float(row.get("paid_leave"), -1)
                holiday = clean_float(row.get("holiday"), -1)
                lop_days = clean_float(row.get("lop_days"), -1)
                paid_days = clean_float(row.get("paid_days"), 0)
                overtime_hours = clean_float(row.get("overtime_hours"), -1)
                bonus = clean_float(row.get("bonus"), 0)
                manual_deduction = clean_float(row.get("manual_deduction"), 0)

                if emp_code == "":
                    row_errors.append(f"Row {row_no}: emp_code missing")

                if month == "":
                    row_errors.append(f"Row {row_no}: month missing")
                else:
                    try:
                        datetime.datetime.strptime(month, "%Y-%m")
                    except Exception:
                        row_errors.append(f"Row {row_no}: month must be in YYYY-MM format, example 2026-12")

                if working_days <= 0:
                    row_errors.append(f"Row {row_no}: working_days must be greater than 0")

                if present_days < 0:
                    row_errors.append(f"Row {row_no}: present_days cannot be negative")

                if weekly_off < 0:
                    row_errors.append(f"Row {row_no}: weekly_off cannot be negative")

                if paid_leave < 0:
                    row_errors.append(f"Row {row_no}: paid_leave cannot be negative")

                if holiday < 0:
                    row_errors.append(f"Row {row_no}: holiday cannot be negative")

                if lop_days < 0:
                    row_errors.append(f"Row {row_no}: lop_days cannot be negative")

                if paid_days < 0:
                    row_errors.append(f"Row {row_no}: paid_days cannot be negative")

                if overtime_hours < 0:
                    row_errors.append(f"Row {row_no}: overtime_hours cannot be negative")

                if bonus < 0:
                    row_errors.append(f"Row {row_no}: bonus cannot be negative")

                if manual_deduction < 0:
                    row_errors.append(f"Row {row_no}: manual_deduction cannot be negative")

                if working_days > 31:
                    row_errors.append(f"Row {row_no}: working_days cannot be greater than 31")

                if present_days > 31:
                    row_errors.append(f"Row {row_no}: present_days cannot be greater than 31")

                if weekly_off > 31:
                    row_errors.append(f"Row {row_no}: weekly_off cannot be greater than 31")

                if paid_leave > 31:
                    row_errors.append(f"Row {row_no}: paid_leave cannot be greater than 31")

                if holiday > 31:
                    row_errors.append(f"Row {row_no}: holiday cannot be greater than 31")

                if lop_days > 31:
                    row_errors.append(f"Row {row_no}: lop_days cannot be greater than 31")

                calculated_paid_days = present_days + weekly_off + paid_leave + holiday - lop_days

                if calculated_paid_days < 0:
                    row_errors.append(f"Row {row_no}: calculated paid_days cannot be negative")

                if calculated_paid_days > 31:
                    row_errors.append(f"Row {row_no}: calculated paid_days cannot be greater than 31")

                if paid_days > 0 and abs(paid_days - calculated_paid_days) > 0.01:
                    row_errors.append(
                        f"Row {row_no}: paid_days mismatch. Expected {calculated_paid_days}, found {paid_days}"
                    )

            if row_errors:
                session["error_report"] = create_error_report(
                    row_errors,
                    "attendance_upload_errors.xlsx"
                )
                flash("Upload failed. Please download the error report and fix the file.", "danger")
                return redirect(url_for("upload_attendance"))

            conn = get_db()
            cur = conn.cursor()

            missing_employee_errors = []

            for index, row in df.iterrows():
                row_no = index + 2
                emp_code = clean_text(row.get("emp_code"))

                cur.execute("""
                    SELECT id
                    FROM employees
                    WHERE company_id = ?
                      AND emp_code = ?
                """, (company_id, emp_code))

                if not cur.fetchone():
                    missing_employee_errors.append(
                        f"Row {row_no}: Employee code not found in Employee Master: {emp_code}"
                    )

            if missing_employee_errors:
                session["error_report"] = create_error_report(
                    missing_employee_errors,
                    "attendance_upload_errors.xlsx"
                )
                flash("Upload failed. Some employee codes were not found. Please download the error report.", "danger")
                return redirect(url_for("upload_attendance"))

            uploaded_months = df["month_clean"].dropna().unique().tolist()

            for uploaded_month in uploaded_months:
                cur.execute("""
                    DELETE FROM attendance
                    WHERE company_id = ?
                      AND month = ?
                """, (company_id, uploaded_month))

            success_count = 0

            for _, row in df.iterrows():
                emp_code = clean_text(row.get("emp_code"))
                month = clean_text(row.get("month"))

                working_days = clean_float(row.get("working_days"), 0)
                present_days = clean_float(row.get("present_days"), 0)
                weekly_off = clean_float(row.get("weekly_off"), 0)
                paid_leave = clean_float(row.get("paid_leave"), 0)
                holiday = clean_float(row.get("holiday"), 0)
                lop_days = clean_float(row.get("lop_days"), 0)
                paid_days = clean_float(row.get("paid_days"), 0)

                if paid_days <= 0:
                    paid_days = present_days + weekly_off + paid_leave + holiday - lop_days

                overtime_hours = clean_float(row.get("overtime_hours"), 0)
                bonus = clean_float(row.get("bonus"), 0)
                manual_deduction = clean_float(row.get("manual_deduction"), 0)

                cur.execute("""
                    INSERT INTO attendance
                    (
                        company_id,
                        emp_code,
                        month,
                        working_days,
                        present_days,
                        weekly_off,
                        paid_leave,
                        holiday,
                        lop_days,
                        paid_days,
                        overtime_hours,
                        bonus,
                        manual_deduction
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    company_id,
                    emp_code,
                    month,
                    working_days,
                    present_days,
                    weekly_off,
                    paid_leave,
                    holiday,
                    lop_days,
                    paid_days,
                    overtime_hours,
                    bonus,
                    manual_deduction
                ))

                success_count += 1

            conn.commit()
            session.pop("error_report", None)

            flash(
                f"Attendance uploaded successfully. Added: {success_count}. Existing attendance for uploaded month(s) was replaced.",
                "success"
            )

            return redirect(url_for("run_payroll"))

        except Exception as e:
            if conn:
                conn.rollback()

            flash(f"Upload failed: {str(e)}", "danger")
            return redirect(url_for("upload_attendance"))

        finally:
            if conn:
                conn.close()

    return render_template("upload_attendance.html")


@app.route("/attendance")
@login_required
def attendance_list():
    month = request.args.get("month", "").strip()
    department = request.args.get("department", "").strip()
    search = request.args.get("search", "").strip()

    company_id = current_company_id()

    if not company_id:
        flash("Company not found. Please login again.", "danger")
        return redirect(url_for("login"))

    conn = get_db()
    cur = conn.cursor()

    query = """
        SELECT
            a.*,
            COALESCE(e.employee_name, '') AS employee_name,
            COALESCE(e.role, '') AS role,
            COALESCE(e.department, '') AS department,
            COALESCE(e.gender, '') AS gender
        FROM attendance a
        LEFT JOIN employees e
          ON a.company_id = e.company_id
         AND a.emp_code = e.emp_code
        WHERE a.company_id = ?
    """

    params = [company_id]

    if month:
        query += " AND a.month = ?"
        params.append(month)

    if department:
        query += " AND e.department = ?"
        params.append(department)

    if search:
        query += """
            AND (
                a.emp_code LIKE ?
                OR e.employee_name LIKE ?
                OR e.role LIKE ?
                OR e.department LIKE ?
            )
        """
        search_value = f"%{search}%"
        params.extend([search_value, search_value, search_value, search_value])

    query += " ORDER BY a.month DESC, a.id DESC"

    cur.execute(query, tuple(params))
    rows = cur.fetchall()

    cur.execute("""
        SELECT DISTINCT department
        FROM employees
        WHERE company_id = ?
          AND department IS NOT NULL
          AND department != ''
        ORDER BY department
    """, (company_id,))
    departments = cur.fetchall()

    cur.execute("""
        SELECT DISTINCT month
        FROM attendance
        WHERE company_id = ?
          AND month IS NOT NULL
          AND month != ''
        ORDER BY month DESC
    """, (company_id,))
    months = cur.fetchall()

    total_records = len(rows)

    total_working_days = round(sum(float(r["working_days"] or 0) for r in rows), 2)
    total_present_days = round(sum(float(r["present_days"] or 0) for r in rows), 2)

    total_weekly_off = round(sum(float(r["weekly_off"] or 0) for r in rows), 2)
    total_paid_leave = round(sum(float(r["paid_leave"] or 0) for r in rows), 2)
    total_holiday = round(sum(float(r["holiday"] or 0) for r in rows), 2)
    total_lop_days = round(sum(float(r["lop_days"] or 0) for r in rows), 2)
    total_paid_days = round(sum(float(r["paid_days"] or 0) for r in rows), 2)

    total_overtime_hours = round(sum(float(r["overtime_hours"] or 0) for r in rows), 2)
    total_bonus = round(sum(float(r["bonus"] or 0) for r in rows))
    total_manual_deduction = round(sum(float(r["manual_deduction"] or 0) for r in rows))

    total_absent_days = round(
        total_working_days
        - total_present_days
        - total_weekly_off
        - total_paid_leave
        - total_holiday,
        2
    )

    if total_absent_days < 0:
        total_absent_days = 0

    if total_working_days > 0:
        attendance_percentage = round((total_present_days / total_working_days) * 100, 2)
        paid_days_percentage = round((total_paid_days / total_working_days) * 100, 2)
    else:
        attendance_percentage = 0
        paid_days_percentage = 0

    conn.close()

    return render_template(
        "attendance.html",
        rows=rows,
        departments=departments,
        months=months,

        selected_month=month,
        selected_department=department,
        search=search,

        total_records=total_records,
        total_working_days=total_working_days,
        total_present_days=total_present_days,
        total_weekly_off=total_weekly_off,
        total_paid_leave=total_paid_leave,
        total_holiday=total_holiday,
        total_lop_days=total_lop_days,
        total_paid_days=total_paid_days,
        total_absent_days=total_absent_days,
        total_overtime_hours=total_overtime_hours,
        total_bonus=total_bonus,
        total_manual_deduction=total_manual_deduction,
        attendance_percentage=attendance_percentage,
        paid_days_percentage=paid_days_percentage
    )


@app.route("/leave-management", methods=["GET", "POST"])
@login_required
def leave_management():
    company_id = current_company_id()

    if not company_id:
        flash("Company not found. Please login again.", "danger")
        return redirect(url_for("login"))

    conn = get_db()
    cur = conn.cursor()

    try:
        # Ensure default leave policy exists
        cur.execute("""
            INSERT OR IGNORE INTO leave_policy_settings
            (
                company_id,
                casual_leave_limit,
                sick_leave_limit,
                paid_leave_limit
            )
            VALUES (?, 6, 6, 12)
        """, (company_id,))

        conn.commit()

        cur.execute("""
            SELECT
                casual_leave_limit,
                sick_leave_limit,
                paid_leave_limit
            FROM leave_policy_settings
            WHERE company_id = ?
        """, (company_id,))

        leave_policy = cur.fetchone()

        if request.method == "POST":
            emp_code = request.form.get("emp_code", "").strip()
            leave_type = request.form.get("leave_type", "").strip()
            start_date = request.form.get("start_date", "").strip()
            end_date = request.form.get("end_date", "").strip()
            total_days_input = request.form.get("total_days", "").strip()
            reason = request.form.get("reason", "").strip()

            errors = []

            if not emp_code:
                errors.append("Employee is required.")

            if not leave_type:
                errors.append("Leave type is required.")

            allowed_leave_types = [
                "Casual Leave",
                "Sick Leave",
                "Paid Leave",
                "Leave Without Pay",
                "Unpaid Leave",
                "LWP"
            ]

            if leave_type and leave_type not in allowed_leave_types:
                errors.append("Invalid leave type selected.")

            start_date_obj = None
            end_date_obj = None

            if not start_date:
                errors.append("Start date is required.")
            else:
                try:
                    start_date_obj = datetime.datetime.strptime(start_date, "%Y-%m-%d")
                except Exception:
                    errors.append("Start date must be a valid date.")

            if not end_date:
                errors.append("End date is required.")
            else:
                try:
                    end_date_obj = datetime.datetime.strptime(end_date, "%Y-%m-%d")
                except Exception:
                    errors.append("End date must be a valid date.")

            if start_date_obj and end_date_obj and end_date_obj < start_date_obj:
                errors.append("End date cannot be earlier than start date.")

            # Auto calculate total days if blank or 0
            total_days = 0

            try:
                total_days = float(total_days_input or 0)
            except Exception:
                total_days = 0

            if total_days <= 0 and start_date_obj and end_date_obj:
                total_days = (end_date_obj - start_date_obj).days + 1

            if total_days <= 0:
                errors.append("Total leave days must be greater than 0.")

            if total_days > 365:
                errors.append("Total leave days cannot be greater than 365.")

            # Employee exists check
            if emp_code:
                cur.execute("""
                    SELECT emp_code
                    FROM employees
                    WHERE company_id = ?
                      AND emp_code = ?
                """, (company_id, emp_code))

                emp = cur.fetchone()

                if not emp:
                    errors.append("Employee code not found in Employee Master.")

            if errors:
                flash(" ".join(errors), "danger")
                return redirect(url_for("leave_management"))

            # Duplicate / overlapping leave request check
            # Same employee ke liye same date range me Pending ya Approved leave dobara create nahi hoga.
            cur.execute("""
                SELECT id, leave_type, start_date, end_date, status
                FROM leave_requests
                WHERE company_id = ?
                  AND emp_code = ?
                  AND status IN ('Pending', 'Approved')
                  AND (
                    date(start_date) <= date(?)
                  AND date(end_date) >= date(?)
                  )
                LIMIT 1
            """, (
    company_id,
    emp_code,
    end_date,
    start_date
))

            existing_leave = cur.fetchone()

            if existing_leave:
                flash(
                    f"Leave request already exists for this employee between "
                    f"{existing_leave['start_date']} and {existing_leave['end_date']} "
                    f"with status {existing_leave['status']}. Duplicate leave not allowed.",
                    "warning"
                )
                return redirect(url_for("leave_management"))

            cur.execute("""
                INSERT INTO leave_requests
                (
                    company_id,
                    emp_code,
                    leave_type,
                    start_date,
                    end_date,
                    total_days,
                    reason,
                    status
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, 'Pending')
            """, (
                company_id,
                emp_code,
                leave_type,
                start_date,
                end_date,
                total_days,
                reason
            ))

            conn.commit()

            flash("Leave request added successfully.", "success")
            return redirect(url_for("leave_management"))

        # Employees list
        cur.execute("""
            SELECT emp_code, employee_name, department
            FROM employees
            WHERE company_id = ?
            ORDER BY employee_name
        """, (company_id,))

        employees = cur.fetchall()

        # Create missing leave balances only
        for emp in employees:
            cur.execute("""
                INSERT OR IGNORE INTO leave_balances
                (
                    company_id,
                    emp_code,
                    casual_leave,
                    sick_leave,
                    paid_leave,
                    used_leave
                )
                VALUES (?, ?, ?, ?, ?, 0)
            """, (
                company_id,
                emp["emp_code"],
                float(leave_policy["casual_leave_limit"] or 6),
                float(leave_policy["sick_leave_limit"] or 6),
                float(leave_policy["paid_leave_limit"] or 12)
            ))

        conn.commit()

        cur.execute("""
            SELECT
                lr.*,
                e.employee_name,
                e.department
            FROM leave_requests lr
            LEFT JOIN employees e
              ON lr.company_id = e.company_id
             AND lr.emp_code = e.emp_code
            WHERE lr.company_id = ?
            ORDER BY lr.id DESC
        """, (company_id,))

        leave_requests = cur.fetchall()

        cur.execute("""
            SELECT
                lb.*,
                e.employee_name,
                e.department
            FROM leave_balances lb
            LEFT JOIN employees e
              ON lb.company_id = e.company_id
             AND lb.emp_code = e.emp_code
            WHERE lb.company_id = ?
            ORDER BY e.employee_name
        """, (company_id,))

        leave_balances = cur.fetchall()

        return render_template(
            "leave_management.html",
            employees=employees,
            leave_requests=leave_requests,
            leave_balances=leave_balances,
            leave_policy=leave_policy
        )

    except Exception as e:
        conn.rollback()
        flash(f"Error in Leave Management: {str(e)}", "danger")
        return redirect(url_for("dashboard"))

    finally:
        conn.close()


@app.route("/update-leave-policy", methods=["POST"])
@login_required
def update_leave_policy():
    company_id = current_company_id()

    if not company_id:
        flash("Company not found. Please login again.", "danger")
        return redirect(url_for("login"))

    def to_float(value, default=0):
        try:
            if value is None or str(value).strip() == "":
                return default
            return float(value)
        except Exception:
            return default

    casual_leave_limit = to_float(request.form.get("casual_leave_limit"), 0)
    sick_leave_limit = to_float(request.form.get("sick_leave_limit"), 0)
    paid_leave_limit = to_float(request.form.get("paid_leave_limit"), 0)

    errors = []

    if casual_leave_limit < 0:
        errors.append("Casual Leave limit cannot be negative.")

    if sick_leave_limit < 0:
        errors.append("Sick Leave limit cannot be negative.")

    if paid_leave_limit < 0:
        errors.append("Paid Leave limit cannot be negative.")

    if casual_leave_limit > 365:
        errors.append("Casual Leave limit cannot be greater than 365.")

    if sick_leave_limit > 365:
        errors.append("Sick Leave limit cannot be greater than 365.")

    if paid_leave_limit > 365:
        errors.append("Paid Leave limit cannot be greater than 365.")

    if errors:
        flash(" ".join(errors), "danger")
        return redirect(url_for("leave_management"))

    conn = get_db()
    cur = conn.cursor()

    try:
        cur.execute("""
            INSERT INTO leave_policy_settings
            (
                company_id,
                casual_leave_limit,
                sick_leave_limit,
                paid_leave_limit,
                updated_at
            )
            VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(company_id) DO UPDATE SET
                casual_leave_limit = excluded.casual_leave_limit,
                sick_leave_limit = excluded.sick_leave_limit,
                paid_leave_limit = excluded.paid_leave_limit,
                updated_at = CURRENT_TIMESTAMP
        """, (
            company_id,
            casual_leave_limit,
            sick_leave_limit,
            paid_leave_limit
        ))

        # Create leave balances for employees who do not have balance records yet.
        # Existing balances are not overwritten.
        cur.execute("""
            SELECT emp_code
            FROM employees
            WHERE company_id = ?
        """, (company_id,))

        employees = cur.fetchall()

        for emp in employees:
            cur.execute("""
                INSERT OR IGNORE INTO leave_balances
                (
                    company_id,
                    emp_code,
                    casual_leave,
                    sick_leave,
                    paid_leave,
                    used_leave
                )
                VALUES (?, ?, ?, ?, ?, 0)
            """, (
                company_id,
                emp["emp_code"],
                casual_leave_limit,
                sick_leave_limit,
                paid_leave_limit
            ))

        conn.commit()

        flash("Leave policy updated successfully. Existing leave balances were not overwritten.", "success")

    except Exception as e:
        conn.rollback()
        flash(f"Error while updating leave policy: {str(e)}", "danger")

    finally:
        conn.close()

    return redirect(url_for("leave_management"))


@app.route("/approve-leave/<int:leave_id>", methods=["POST"])
@login_required
def approve_leave(leave_id):
    company_id = current_company_id()

    if not company_id:
        flash("Company not found. Please login again.", "danger")
        return redirect(url_for("login"))

    conn = get_db()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT *
            FROM leave_requests
            WHERE id = ?
              AND company_id = ?
        """, (leave_id, company_id))

        leave = cur.fetchone()

        if not leave:
            flash("Leave request not found.", "warning")
            return redirect(url_for("leave_management"))

        current_status = str(leave["status"] or "").strip()

        if current_status == "Approved":
            flash("This leave is already approved.", "warning")
            return redirect(url_for("leave_management"))

        if current_status == "Rejected":
            flash("Rejected leave cannot be approved directly. Please create a new leave request.", "warning")
            return redirect(url_for("leave_management"))

        emp_code = str(leave["emp_code"] or "").strip()
        leave_type = str(leave["leave_type"] or "").strip()
        total_days = float(leave["total_days"] or 0)

        if not emp_code:
            flash("Employee code missing in leave request.", "danger")
            return redirect(url_for("leave_management"))

        if total_days <= 0:
            flash("Leave days must be greater than 0.", "danger")
            return redirect(url_for("leave_management"))

        # Paid leave types reduce balance.
        leave_column_map = {
            "Casual Leave": "casual_leave",
            "Sick Leave": "sick_leave",
            "Paid Leave": "paid_leave"
        }

        # LWP / unpaid leave does not reduce leave balance.
        unpaid_leave_types = [
            "Leave Without Pay",
            "Unpaid Leave",
            "LWP"
        ]

        if leave_type in unpaid_leave_types:
            cur.execute("""
                UPDATE leave_requests
                SET status = 'Approved'
                WHERE id = ?
                  AND company_id = ?
            """, (leave_id, company_id))

            conn.commit()

            flash("Leave Without Pay approved successfully. Leave balance was not changed.", "success")
            return redirect(url_for("leave_management"))

        if leave_type not in leave_column_map:
            flash(f"Invalid leave type: {leave_type}", "danger")
            return redirect(url_for("leave_management"))

        balance_column = leave_column_map[leave_type]

        cur.execute("""
            SELECT *
            FROM leave_balances
            WHERE company_id = ?
              AND emp_code = ?
        """, (company_id, emp_code))

        balance_row = cur.fetchone()

        if not balance_row:
            flash("Leave balance not found for this employee. Please check leave balance setup.", "danger")
            return redirect(url_for("leave_management"))

        current_balance = float(balance_row[balance_column] or 0)
        used_leave = float(balance_row["used_leave"] or 0)

        if current_balance < total_days:
            flash(
                f"Insufficient {leave_type} balance. Available: {current_balance}, Required: {total_days}.",
                "danger"
            )
            return redirect(url_for("leave_management"))

        new_balance = current_balance - total_days
        new_used_leave = used_leave + total_days

        cur.execute(f"""
            UPDATE leave_balances
            SET {balance_column} = ?,
                used_leave = ?
            WHERE company_id = ?
              AND emp_code = ?
        """, (new_balance, new_used_leave, company_id, emp_code))

        cur.execute("""
            UPDATE leave_requests
            SET status = 'Approved'
            WHERE id = ?
              AND company_id = ?
        """, (leave_id, company_id))

        conn.commit()

        flash(
            f"Leave approved successfully. {leave_type} balance updated: {current_balance} → {new_balance}.",
            "success"
        )

    except Exception as e:
        conn.rollback()
        flash(f"Error while approving leave: {str(e)}", "danger")

    finally:
        conn.close()

    return redirect(url_for("leave_management"))


@app.route("/reject-leave/<int:leave_id>", methods=["POST"])
@login_required
def reject_leave(leave_id):
    company_id = current_company_id()

    if not company_id:
        flash("Company not found. Please login again.", "danger")
        return redirect(url_for("login"))

    conn = get_db()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT *
            FROM leave_requests
            WHERE id = ?
              AND company_id = ?
        """, (leave_id, company_id))

        leave = cur.fetchone()

        if not leave:
            flash("Leave request not found.", "warning")
            return redirect(url_for("leave_management"))

        current_status = str(leave["status"] or "").strip()

        if current_status == "Rejected":
            flash("This leave is already rejected.", "warning")
            return redirect(url_for("leave_management"))

        if current_status == "Approved":
            flash("Approved leave cannot be rejected directly because leave balance is already updated. Delete/cancel reversal logic is required.", "warning")
            return redirect(url_for("leave_management"))

        cur.execute("""
            UPDATE leave_requests
            SET status = 'Rejected'
            WHERE id = ?
              AND company_id = ?
        """, (leave_id, company_id))

        conn.commit()

        flash("Leave rejected successfully. Leave balance was not changed.", "success")

    except Exception as e:
        conn.rollback()
        flash(f"Error while rejecting leave: {str(e)}", "danger")

    finally:
        conn.close()

    return redirect(url_for("leave_management"))


@app.route("/cancel-leave/<int:leave_id>", methods=["POST"])
@login_required
def cancel_leave(leave_id):
    company_id = current_company_id()

    if not company_id:
        flash("Company not found. Please login again.", "danger")
        return redirect(url_for("login"))

    conn = get_db()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT *
            FROM leave_requests
            WHERE id = ?
              AND company_id = ?
        """, (leave_id, company_id))

        leave = cur.fetchone()

        if not leave:
            flash("Leave request not found.", "warning")
            return redirect(url_for("leave_management"))

        emp_code = str(leave["emp_code"] or "").strip()
        leave_type = str(leave["leave_type"] or "").strip()
        status = str(leave["status"] or "").strip()
        total_days = float(leave["total_days"] or 0)

        leave_column_map = {
            "Casual Leave": "casual_leave",
            "Sick Leave": "sick_leave",
            "Paid Leave": "paid_leave"
        }

        unpaid_leave_types = [
            "Leave Without Pay",
            "Unpaid Leave",
            "LWP"
        ]

        # Approved paid leave cancel/reverse karega
        if status == "Approved" and leave_type in leave_column_map:
            balance_column = leave_column_map[leave_type]

            cur.execute("""
                SELECT *
                FROM leave_balances
                WHERE company_id = ?
                  AND emp_code = ?
            """, (company_id, emp_code))

            balance_row = cur.fetchone()

            if not balance_row:
                flash("Leave balance not found. Cannot reverse approved leave.", "danger")
                return redirect(url_for("leave_management"))

            current_balance = float(balance_row[balance_column] or 0)
            used_leave = float(balance_row["used_leave"] or 0)

            new_balance = current_balance + total_days
            new_used_leave = used_leave - total_days

            if new_used_leave < 0:
                new_used_leave = 0

            cur.execute(f"""
                UPDATE leave_balances
                SET {balance_column} = ?,
                    used_leave = ?
                WHERE company_id = ?
                  AND emp_code = ?
            """, (
                new_balance,
                new_used_leave,
                company_id,
                emp_code
            ))

        # Approved LWP / Pending / Rejected delete only, balance unchanged
        elif status == "Approved" and leave_type in unpaid_leave_types:
            pass

        elif status in ["Pending", "Rejected"]:
            pass

        else:
            flash("Invalid leave status. Cannot cancel this leave request.", "warning")
            return redirect(url_for("leave_management"))

        cur.execute("""
            DELETE FROM leave_requests
            WHERE id = ?
              AND company_id = ?
        """, (leave_id, company_id))

        conn.commit()

        flash("Leave request cancelled/deleted successfully.", "success")

    except Exception as e:
        conn.rollback()
        flash(f"Error while cancelling leave: {str(e)}", "danger")

    finally:
        conn.close()

    return redirect(url_for("leave_management"))


@app.route("/download-attendance-sample")
@login_required
def download_attendance_sample():
    company_id = current_company_id()

    if not company_id:
        flash("Company not found. Please login again.", "danger")
        return redirect(url_for("login"))

    data = {
        "emp_code": ["EMP001", "EMP002"],
        "month": ["2026-12", "2026-12"],
        "working_days": [30, 30],
        "present_days": [26, 24],
        "weekly_off": [4, 4],
        "paid_leave": [0, 1],
        "holiday": [0, 0],
        "lop_days": [0, 1],
        "paid_days": [30, 28],
        "overtime_hours": [2, 2],
        "bonus": [0, 0],
        "manual_deduction": [0, 0]
    }

    df = pd.DataFrame(data)

    file_path = os.path.join(UPLOAD_FOLDER, "attendance_sample.xlsx")

    instructions = pd.DataFrame({
        "Field": [
            "emp_code",
            "month",
            "working_days",
            "present_days",
            "weekly_off",
            "paid_leave",
            "holiday",
            "lop_days",
            "paid_days",
            "overtime_hours",
            "bonus",
            "manual_deduction"
        ],
        "Required": [
            "Yes",
            "Yes",
            "Yes",
            "Yes",
            "Yes",
            "Yes",
            "Yes",
            "Yes",
            "Optional",
            "Yes",
            "Optional",
            "Optional"
        ],
        "Example": [
            "EMP001",
            "2026-12",
            "30",
            "26",
            "4",
            "0",
            "0",
            "0",
            "30",
            "2",
            "0",
            "0"
        ],
        "Notes": [
            "Employee code must exist in Employee Master.",
            "Use YYYY-MM format only, example 2026-12.",
            "Month working days / salary days as per company policy.",
            "Actual present days.",
            "Weekly off days. Example: 4 Sundays.",
            "Paid leave days included in salary payable days.",
            "Paid holiday days included in salary payable days.",
            "Loss of Pay / unpaid leave days.",
            "Optional. If blank or 0, system calculates: present_days + weekly_off + paid_leave + holiday - lop_days.",
            "Use 0 if no overtime.",
            "Optional attendance bonus. Use 0 if not applicable.",
            "Optional manual deduction. Use 0 if not applicable."
        ]
    })

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Attendance")
        instructions.to_excel(writer, index=False, sheet_name="Instructions")

        workbook = writer.book

        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        required_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        optional_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

        thin_border = Border(
            left=Side(style="thin", color="E2E8F0"),
            right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin", color="E2E8F0"),
            bottom=Side(style="thin", color="E2E8F0")
        )

        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            for row_cells in ws.iter_rows(min_row=2):
                for cell in row_cells:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center")

            for column_cells in ws.columns:
                max_length = 0
                column_letter = column_cells[0].column_letter

                for cell in column_cells:
                    cell_value = str(cell.value) if cell.value is not None else ""
                    max_length = max(max_length, len(cell_value))

                ws.column_dimensions[column_letter].width = max_length + 4

        # Keep emp_code and month as text
        attendance_ws = workbook["Attendance"]

        for col in ["A", "B"]:
            for cell in attendance_ws[col]:
                cell.number_format = "@"

        # Format instruction required/optional rows
        instruction_ws = workbook["Instructions"]

        for row_idx in range(2, instruction_ws.max_row + 1):
            required_value = str(instruction_ws.cell(row=row_idx, column=2).value or "").strip().lower()

            if required_value == "yes":
                fill = required_fill
            else:
                fill = optional_fill

            for col_idx in range(1, instruction_ws.max_column + 1):
                instruction_ws.cell(row=row_idx, column=col_idx).fill = fill

    return send_file(
        file_path,
        as_attachment=True,
        download_name="attendance_sample.xlsx"
    )


@app.route("/download-error-report")
@login_required
def download_error_report():
    file_path = session.get("error_report")
    if not file_path or not os.path.exists(file_path):
        flash("No error report available.")
        return redirect(url_for("dashboard"))
    return send_file(file_path, as_attachment=True)


# ---------------------------
# PAYROLL
# ---------------------------
@app.route("/run-payroll", methods=["GET", "POST"])
@login_required
def run_payroll():
    if request.method == "GET":
        return render_template("run_payroll.html")

    company_id = current_company_id()
    month = request.form.get("month")

    if not month:
        flash("Please select payroll month.")
        return redirect(url_for("run_payroll"))

    conn = get_db()
    cur = conn.cursor()
    run_id = datetime.datetime.now().strftime("%Y%m%d%H%M%S")

    # Company overtime policy
    cur.execute("""
        SELECT COALESCE(overtime_multiplier, 1) AS overtime_multiplier
        FROM companies
        WHERE id = ?
    """, (company_id,))
    company = cur.fetchone()

    overtime_multiplier = 1
    if company:
        try:
            overtime_multiplier = float(company["overtime_multiplier"] or 1)
        except Exception:
            overtime_multiplier = 1

    if overtime_multiplier not in [1, 2]:
        overtime_multiplier = 1

    # Compliance settings
    settings = get_compliance_settings(company_id)

    pf_employee_rate = float(settings["pf_employee_rate"] or 12) / 100
    pf_employer_rate = float(settings["pf_employer_rate"] or 12) / 100
    pf_wage_ceiling = float(settings["pf_wage_ceiling"] or 15000)
    pf_max_deduction = float(settings["pf_max_deduction"] or 1800)

    esic_employee_rate = float(settings["esic_employee_rate"] or 0.75) / 100
    esic_employer_rate = float(settings["esic_employer_rate"] or 3.25) / 100
    esic_wage_limit = float(settings["esic_wage_limit"] or 21000)

    gratuity_rate = float(settings["gratuity_rate"] or 4.81) / 100
    bonus_rate = float(settings["bonus_rate"] or 8.33) / 100
    tds_enabled = int(settings["tds_enabled"] or 0)

    salary_days_policy = settings["salary_days_policy"] or "attendance"
    custom_salary_days = float(settings["custom_salary_days"] or 30)

    count_weekly_off_paid = int(settings["count_weekly_off_paid"] or 0)
    count_paid_leave_paid = int(settings["count_paid_leave_paid"] or 0)
    count_holiday_paid = int(settings["count_holiday_paid"] or 0)
    deduct_lop = int(settings["deduct_lop"] or 0)

    festival_bonus_enabled = int(settings["festival_bonus_enabled"] or 0)
    festival_bonus_month = int(settings["festival_bonus_month"] or 10)

    bonus_min_service_days = int(settings["bonus_min_service_days"] or 30)
    bonus_prorata_enabled = int(settings["bonus_prorata_enabled"] or 1)

    # Mark previous payroll of same month as old
    cur.execute("""
        UPDATE payroll_history
        SET is_current = 0
        WHERE company_id = ?
          AND month = ?
    """, (company_id, month))

    # Fetch employees + attendance
    cur.execute("""
        SELECT
            e.*,
            a.working_days,
            a.present_days,
            COALESCE(a.weekly_off, 0) AS weekly_off,
            COALESCE(a.paid_leave, 0) AS attendance_paid_leave,
            COALESCE(a.holiday, 0) AS holiday,
            COALESCE(a.lop_days, 0) AS attendance_lop_days,
            COALESCE(a.paid_days, 0) AS attendance_paid_days,
            a.overtime_hours,
            a.bonus,
            a.manual_deduction
        FROM employees e
        JOIN attendance a
          ON e.emp_code = a.emp_code
         AND e.company_id = a.company_id
        WHERE e.company_id = ?
          AND a.month = ?
    """, (company_id, month))

    rows = cur.fetchall()

    if not rows:
        conn.close()
        flash("No attendance found for selected month.")
        return redirect(url_for("run_payroll"))

    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    year, month_num = map(int, month.split("-"))
    calendar_days = calendar.monthrange(year, month_num)[1]

    # Financial year range
    financial_year_start = f"{year}-04"
    financial_year_end = f"{year + 1}-03"

    if month_num < 4:
        financial_year_start = f"{year - 1}-04"
        financial_year_end = f"{year}-03"

    for row in rows:
        emp_code = row["emp_code"]
        monthly_salary = float(row["monthly_salary"] or 0)

        attendance_working_days = float(row["working_days"] or 30)

        # Salary days policy
        if salary_days_policy == "fixed_26":
            working_days = 26
        elif salary_days_policy == "fixed_30":
            working_days = 30
        elif salary_days_policy == "calendar":
            working_days = calendar_days
        elif salary_days_policy == "custom":
            working_days = custom_salary_days
        else:
            working_days = attendance_working_days

        if working_days <= 0:
            working_days = 30

        present_days = float(row["present_days"] or 0)
        weekly_off = float(row["weekly_off"] or 0)
        attendance_paid_leave = float(row["attendance_paid_leave"] or 0)
        holiday = float(row["holiday"] or 0)
        attendance_lop_days = float(row["attendance_lop_days"] or 0)

        overtime_hours = float(row["overtime_hours"] or 0)
        manual_deduction = float(row["manual_deduction"] or 0)
        attendance_bonus = float(row["bonus"] or 0)

        gender = str(row["gender"] or "male").strip().lower()

        # Approved leave data
        cur.execute("""
            SELECT
                COALESCE(SUM(
                    CASE
                        WHEN leave_type IN ('Casual Leave', 'Sick Leave', 'Paid Leave')
                        THEN total_days
                        ELSE 0
                    END
                ), 0) AS approved_paid_leave_days,

                COALESCE(SUM(
                    CASE
                        WHEN leave_type = 'Unpaid Leave'
                        THEN total_days
                        ELSE 0
                    END
                ), 0) AS approved_lwp_days
            FROM leave_requests
            WHERE company_id = ?
              AND emp_code = ?
              AND status = 'Approved'
              AND (
                    substr(start_date, 1, 7) = ?
                    OR substr(end_date, 1, 7) = ?
              )
        """, (company_id, emp_code, month, month))

        leave_data = cur.fetchone()

        approved_paid_leave_days = float(leave_data["approved_paid_leave_days"] or 0)
        approved_lwp_days = float(leave_data["approved_lwp_days"] or 0)

        paid_leave_days = attendance_paid_leave if attendance_paid_leave > 0 else approved_paid_leave_days
        lwp_days = attendance_lop_days if attendance_lop_days > 0 else approved_lwp_days

        # Payable days calculation
        payable_days = present_days

        if count_weekly_off_paid == 1:
            payable_days += weekly_off

        if count_paid_leave_paid == 1:
            payable_days += paid_leave_days

        if count_holiday_paid == 1:
            payable_days += holiday

        if deduct_lop == 1:
            payable_days -= lwp_days

        if payable_days < 0:
            payable_days = 0

        if payable_days > working_days:
            payable_days = working_days

        per_day_salary = monthly_salary / working_days
        lwp_deduction = rupee(per_day_salary * lwp_days)

        earned_salary = per_day_salary * payable_days

        basic = rupee(earned_salary * 0.40)
        da = rupee(earned_salary * 0.10)
        hra = rupee(earned_salary * 0.20)

        special_allowance = float(row["special_allowance"] or 0)
        special_allowance = rupee((special_allowance / working_days) * payable_days)

        other_allowance = earned_salary - basic - da - hra - special_allowance
        if other_allowance < 0:
            other_allowance = 0

        other_allowance = rupee(other_allowance)

        gross = rupee(basic + da + hra + special_allowance + other_allowance)

        # Overtime calculation
        if overtime_hours > 0:
            hourly_rate = monthly_salary / 30 / 8
            overtime_amount = rupee(hourly_rate * overtime_hours * overtime_multiplier)
        else:
            overtime_amount = 0

        # PF calculation with wage ceiling
        pf_base = basic + da

        if pf_wage_ceiling > 0:
            pf_base_for_calculation = min(pf_base, pf_wage_ceiling)
        else:
            pf_base_for_calculation = pf_base

        pf_employee = min(rupee(pf_base_for_calculation * pf_employee_rate), pf_max_deduction)
        pf_employer = min(rupee(pf_base_for_calculation * pf_employer_rate), pf_max_deduction)

        # ESIC calculation
        if gross <= esic_wage_limit:
            esi_employee = round(gross * esic_employee_rate)
            esi_employer = round(gross * esic_employer_rate)
        else:
            esi_employee = 0
            esi_employer = 0

        # Professional tax
        professional_tax = rupee(
            calculate_professional_tax_maharashtra(gross, gender, month)
        )

        # LWF
        lwf = calculate_lwf_maharashtra(month)
        lwf_employee = rupee(lwf["employee"])
        lwf_employer = rupee(lwf["employer"])

        # TDS placeholder
        # Future me actual income tax regime logic add kar sakte hain
        if tds_enabled == 1:
            tds = 0
        else:
            tds = 0

        # Bonus accrual logic
        monthly_bonus_accrual = 0

        if payable_days >= bonus_min_service_days:
            if bonus_prorata_enabled == 1:
                monthly_bonus_accrual = rupee(basic * bonus_rate)
            else:
                full_month_basic = monthly_salary * 0.40
                monthly_bonus_accrual = rupee(full_month_basic * bonus_rate)

        bonus_ctc = monthly_bonus_accrual

        # Festival bonus payout logic
        festival_bonus = 0

        if festival_bonus_enabled == 1 and month_num == festival_bonus_month:
            cur.execute("""
                SELECT COALESCE(SUM(bonus_ctc), 0) AS accumulated_bonus
                FROM payroll_history
                WHERE company_id = ?
                  AND emp_code = ?
                  AND is_current = 1
                  AND month >= ?
                  AND month <= ?
            """, (
                company_id,
                emp_code,
                financial_year_start,
                month
            ))

            bonus_data = cur.fetchone()
            accumulated_bonus = float(bonus_data["accumulated_bonus"] or 0)

            festival_bonus = rupee(accumulated_bonus + monthly_bonus_accrual)

        festival_bonus = rupee(festival_bonus + attendance_bonus)

        total_deductions = rupee(
            esi_employee
            + professional_tax
            + pf_employee
            + lwf_employee
            + tds
            + manual_deduction
        )

        gratuity = rupee(basic * gratuity_rate)

        total_contributions = rupee(
            esi_employer
            + pf_employer
            + gratuity
            + lwf_employer
        )

        net_pay = rupee(
            gross
            + overtime_amount
            + festival_bonus
            - total_deductions
        )

        monthly_ctc = rupee(
            gross
            + overtime_amount
            + total_contributions
            + bonus_ctc
        )

        annual_ctc = rupee(monthly_ctc * 12)

        cur.execute("""
            INSERT INTO payroll_history (
                company_id,
                emp_code,
                employee_name,
                role,
                department,
                gender,
                month,
                monthly_salary,

                paid_leave_days,
                lwp_days,
                lwp_deduction,
                payable_days,

                basic,
                da,
                hra,
                special_allowance,
                other_allowance,
                gross,

                esi_employee,
                professional_tax,
                pf_employee,
                lwf_employee,
                tds,
                manual_deduction,
                total_deductions,

                esi_employer,
                pf_employer,
                gratuity,
                bonus_ctc,
                festival_bonus,
                lwf_employer,
                total_contributions,

                net_pay,
                monthly_ctc,
                annual_ctc,

                overtime_hours,
                overtime_amount,
                created_at,
                run_id,
                is_current
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            company_id,
            emp_code,
            row["employee_name"],
            row["role"],
            row["department"],
            gender,

            month,
            monthly_salary,

            paid_leave_days,
            lwp_days,
            lwp_deduction,
            payable_days,

            basic,
            da,
            hra,
            special_allowance,
            other_allowance,
            gross,

            esi_employee,
            professional_tax,
            pf_employee,
            lwf_employee,
            tds,
            manual_deduction,
            total_deductions,

            esi_employer,
            pf_employer,
            gratuity,
            bonus_ctc,
            festival_bonus,
            lwf_employer,
            total_contributions,

            net_pay,
            monthly_ctc,
            annual_ctc,

            overtime_hours,
            overtime_amount,
            now,
            run_id,
            1
        ))

    conn.commit()
    conn.close()

    flash("Payroll run completed successfully.")
    return redirect(url_for("payroll_history"))


@app.route("/payroll-history")
@login_required
def payroll_history():
    ensure_employee_personal_columns()
    ensure_whatsapp_payslip_log_table()

    month = request.args.get("month", "").strip()
    department = request.args.get("department", "").strip()
    search = request.args.get("search", "").strip()

    company_id = current_company_id()

    if not company_id:
        flash("Company not found. Please login again.", "danger")
        return redirect(url_for("login"))

    conn = get_db()
    cur = conn.cursor()

    query = """
        SELECT
            p.*,

            COALESCE(p.payable_days, 0) AS payable_days,
            COALESCE(p.paid_leave_days, 0) AS paid_leave_days,
            COALESCE(p.lwp_days, 0) AS lwp_days,
            COALESCE(p.lwp_deduction, 0) AS lwp_deduction,

            COALESCE(p.overtime_amount, 0) AS overtime_amount,
            COALESCE(p.festival_bonus, 0) AS festival_bonus,
            COALESCE(p.bonus_ctc, 0) AS bonus_ctc,

            COALESCE(p.pf_employee, 0) AS pf_employee,
            COALESCE(p.esi_employee, 0) AS esi_employee,
            COALESCE(p.professional_tax, 0) AS professional_tax,
            COALESCE(p.lwf_employee, 0) AS lwf_employee,
            COALESCE(p.tds, 0) AS tds,
            COALESCE(p.manual_deduction, 0) AS manual_deduction,

            COALESCE(p.pf_employer, 0) AS pf_employer,
            COALESCE(p.esi_employer, 0) AS esi_employer,
            COALESCE(p.gratuity, 0) AS gratuity,
            COALESCE(p.lwf_employer, 0) AS lwf_employer,

            COALESCE(p.total_contributions, 0) AS total_contributions,
            COALESCE(p.monthly_ctc, 0) AS monthly_ctc,
            COALESCE(p.annual_ctc, 0) AS annual_ctc,

            COALESCE(e.mobile_no, '') AS employee_mobile_no,
            COALESCE(e.email_id, '') AS employee_email_id,

            COALESCE(w.status, '') AS whatsapp_send_status,
            COALESCE(w.error_message, '') AS whatsapp_error_message,
            COALESCE(w.sent_at, '') AS whatsapp_sent_at

        FROM payroll_history AS p

        LEFT JOIN employees AS e
          ON e.company_id = p.company_id
         AND e.emp_code = p.emp_code

        LEFT JOIN whatsapp_payslip_logs AS w
          ON w.id = (
                SELECT wl.id
                FROM whatsapp_payslip_logs AS wl
                WHERE wl.company_id = p.company_id
                  AND wl.payroll_id = p.id
                ORDER BY wl.id DESC
                LIMIT 1
          )

        WHERE p.company_id = ?
          AND p.is_current = 1
    """

    params = [company_id]

    if month:
        query += " AND p.month = ?"
        params.append(month)

    if department:
        query += " AND p.department = ?"
        params.append(department)

    if search:
        query += """
            AND (
                p.emp_code LIKE ?
                OR p.employee_name LIKE ?
                OR p.role LIKE ?
                OR p.department LIKE ?
                OR COALESCE(e.mobile_no, '') LIKE ?
                OR COALESCE(e.email_id, '') LIKE ?
            )
        """

        search_value = f"%{search}%"

        params.extend([
            search_value,
            search_value,
            search_value,
            search_value,
            search_value,
            search_value
        ])

    query += " ORDER BY p.month DESC, p.id DESC"

    cur.execute(query, tuple(params))
    records = cur.fetchall()

    cur.execute("""
        SELECT DISTINCT department
        FROM employees
        WHERE company_id = ?
          AND department IS NOT NULL
          AND department != ''
        ORDER BY department
    """, (company_id,))
    departments = cur.fetchall()

    total_employees = len(records)

    total_gross = round(
        sum(float(row["gross"] or 0) for row in records)
    )
    total_overtime_amount = round(
        sum(float(row["overtime_amount"] or 0) for row in records)
    )
    total_festival_bonus = round(
        sum(float(row["festival_bonus"] or 0) for row in records)
    )
    total_bonus_ctc = round(
        sum(float(row["bonus_ctc"] or 0) for row in records)
    )
    total_net_pay = round(
        sum(float(row["net_pay"] or 0) for row in records)
    )
    total_deductions = round(
        sum(float(row["total_deductions"] or 0) for row in records)
    )

    total_pf_employee = round(
        sum(float(row["pf_employee"] or 0) for row in records)
    )
    total_esi_employee = round(
        sum(float(row["esi_employee"] or 0) for row in records)
    )
    total_professional_tax = round(
        sum(float(row["professional_tax"] or 0) for row in records)
    )
    total_lwf_employee = round(
        sum(float(row["lwf_employee"] or 0) for row in records)
    )
    total_tds = round(
        sum(float(row["tds"] or 0) for row in records)
    )
    total_manual_deduction = round(
        sum(float(row["manual_deduction"] or 0) for row in records)
    )

    total_pf_employer = round(
        sum(float(row["pf_employer"] or 0) for row in records)
    )
    total_esi_employer = round(
        sum(float(row["esi_employer"] or 0) for row in records)
    )
    total_gratuity = round(
        sum(float(row["gratuity"] or 0) for row in records)
    )
    total_lwf_employer = round(
        sum(float(row["lwf_employer"] or 0) for row in records)
    )

    total_employer_cost = round(
        total_pf_employer
        + total_esi_employer
        + total_gratuity
        + total_lwf_employer
    )

    total_contributions = round(
        sum(float(row["total_contributions"] or 0) for row in records)
    )
    total_monthly_ctc = round(
        sum(float(row["monthly_ctc"] or 0) for row in records)
    )
    total_annual_ctc = round(
        sum(float(row["annual_ctc"] or 0) for row in records)
    )

    total_paid_leave_days = round(
        sum(float(row["paid_leave_days"] or 0) for row in records),
        2
    )
    total_lwp_days = round(
        sum(float(row["lwp_days"] or 0) for row in records),
        2
    )
    total_lwp_deduction = round(
        sum(float(row["lwp_deduction"] or 0) for row in records)
    )

    whatsapp_ready_count = 0
    mobile_missing_count = 0

    for row in records:
        if clean_whatsapp_number_for_send(
            row["employee_mobile_no"]
        ):
            whatsapp_ready_count += 1
        else:
            mobile_missing_count += 1

    conn.close()

    return render_template(
        "payroll_history.html",
        records=records,
        departments=departments,

        selected_month=month,
        selected_department=department,
        search=search,

        total_employees=total_employees,
        total_gross=total_gross,
        total_overtime_amount=total_overtime_amount,
        total_festival_bonus=total_festival_bonus,
        total_bonus_ctc=total_bonus_ctc,
        total_net_pay=total_net_pay,
        total_deductions=total_deductions,

        total_pf_employee=total_pf_employee,
        total_esi_employee=total_esi_employee,
        total_professional_tax=total_professional_tax,
        total_lwf_employee=total_lwf_employee,
        total_tds=total_tds,
        total_manual_deduction=total_manual_deduction,

        total_pf_employer=total_pf_employer,
        total_esi_employer=total_esi_employer,
        total_gratuity=total_gratuity,
        total_lwf_employer=total_lwf_employer,
        total_employer_cost=total_employer_cost,

        total_contributions=total_contributions,
        total_monthly_ctc=total_monthly_ctc,
        total_annual_ctc=total_annual_ctc,

        total_paid_leave_days=total_paid_leave_days,
        total_lwp_days=total_lwp_days,
        total_lwp_deduction=total_lwp_deduction,

        whatsapp_ready_count=whatsapp_ready_count,
        mobile_missing_count=mobile_missing_count,

        now=datetime.datetime.now()
    )


# ------------------------------------------------------------
# PAYROLL PRO - PROFESSIONAL TAX REPORT
# ------------------------------------------------------------
# Paste this complete block in app.py after the payroll_history() route.
#
# Required imports (add only if not already present):
# from reportlab.lib.pagesizes import A4
# from reportlab.pdfgen import canvas
# from flask import send_file
# import os
# import datetime
# ------------------------------------------------------------


def _pt_clean(value, default="-"):
    value = str(value or "").strip()

    if not value or value.lower() in ["nan", "none", "null"]:
        return default

    return value


def _pt_float(value, default=0):
    try:
        if value is None or str(value).strip() == "":
            return float(default)

        return float(value)

    except Exception:
        return float(default)


def _pt_money(value):
    try:
        return round(float(value or 0), 2)

    except Exception:
        return 0.0


def _pt_month_label(month):
    try:
        return datetime.datetime.strptime(month, "%Y-%m").strftime("%B %Y")

    except Exception:
        return month or "-"


def _pt_safe_filename(value, default="file"):
    value = _pt_clean(value, default)

    for ch in [" ", "/", "\\", ":", "*", "?", '"', "<", ">", "|"]:
        value = value.replace(ch, "_")

    return value


def _get_pt_company_details(cur, company_id):
    """
    companies table se available company fields safely read karta hai.
    Missing optional columns ki wajah se PT report crash nahi hogi.
    """
    company = {
        "company_name": "",
        "company_address": "",
        "company_email": "",
        "company_phone": "",
        "pt_registration_no": ""
    }

    cur.execute("""
        SELECT *
        FROM companies
        WHERE id = ?
        LIMIT 1
    """, (company_id,))

    row = cur.fetchone()

    if not row:
        return company

    keys = set(row.keys())

    if "company_name" in keys:
        company["company_name"] = _pt_clean(row["company_name"], "")

    if "address" in keys:
        company["company_address"] = _pt_clean(row["address"], "")

    if "email" in keys:
        company["company_email"] = _pt_clean(row["email"], "")

    if "phone" in keys:
        company["company_phone"] = _pt_clean(row["phone"], "")

    if "pt_registration_no" in keys:
        company["pt_registration_no"] = _pt_clean(
            row["pt_registration_no"],
            ""
        )

    return company


def _get_pt_settings(cur, company_id, month):
    """
    Compliance settings me saved Maharashtra PT slabs read karta hai.
    Missing settings par current default Maharashtra slabs use hote hain.
    """
    settings = {
        "pt_enabled": 1,
        "pt_state": "Maharashtra",
        "pt_male_slab_1_limit": 7500,
        "pt_male_slab_1_amount": 0,
        "pt_male_slab_2_limit": 10000,
        "pt_male_slab_2_amount": 175,
        "pt_male_above_amount": 200,
        "pt_february_amount": 300,
        "pt_female_exemption_limit": 25000,
        "pt_exempt_designations": "DIRECTOR,CEO"
    }

    try:
        cur.execute("""
            SELECT *
            FROM compliance_settings
            WHERE company_id = ?
            LIMIT 1
        """, (company_id,))

        row = cur.fetchone()

        if row:
            keys = set(row.keys())

            for key in settings:
                if (
                    key in keys
                    and row[key] is not None
                    and str(row[key]).strip() != ""
                ):
                    if key in ["pt_state", "pt_exempt_designations"]:
                        settings[key] = _pt_clean(row[key], settings[key])
                    else:
                        settings[key] = _pt_float(row[key], settings[key])

    except Exception as e:
        print("PT settings read failed:", e)

    try:
        month_no = int(str(month).split("-")[1])

    except Exception:
        month_no = 0

    settings["pt_above_slab_amount_for_month"] = (
        settings["pt_february_amount"]
        if month_no == 2
        else settings["pt_male_above_amount"]
    )

    return settings


def get_professional_tax_report_data(company_id, month=""):
    """
    Payroll Pro payroll_history ke saved gross/professional_tax values se
    month-wise PT report data banata hai.
    """
    conn = get_db()
    cur = conn.cursor()

    data = {
        "company": {},
        "month": month,
        "month_label": _pt_month_label(month),
        "settings": {},
        "summary_rows": [],
        "details": [],
        "total_amount": 0,
        "total_employees": 0,
        "pt_deducted_employees": 0,
        "months": []
    }

    try:
        # Available payroll months
        cur.execute("""
            SELECT DISTINCT month
            FROM payroll_history
            WHERE company_id = ?
              AND COALESCE(is_current, 1) = 1
              AND month IS NOT NULL
              AND TRIM(month) != ''
            ORDER BY month DESC
        """, (company_id,))

        data["months"] = [
            row["month"]
            for row in cur.fetchall()
            if row["month"]
        ]

        if not month and data["months"]:
            month = data["months"][0]
            data["month"] = month
            data["month_label"] = _pt_month_label(month)

        data["company"] = _get_pt_company_details(cur, company_id)
        data["settings"] = _get_pt_settings(cur, company_id, month)

        if not month:
            return data

        # Saved payroll values only; PT is not recalculated here.
        cur.execute("""
            SELECT
                COALESCE(emp_code, '') AS emp_code,
                COALESCE(employee_name, '') AS employee_name,
                COALESCE(gender, '') AS gender,
                COALESCE(role, '') AS designation,
                COALESCE(department, '') AS department,
                COALESCE(gross, 0) AS gross_earning,
                COALESCE(professional_tax, 0) AS pt_amount
            FROM payroll_history
            WHERE company_id = ?
              AND month = ?
              AND COALESCE(is_current, 1) = 1
            ORDER BY employee_name COLLATE NOCASE, emp_code
        """, (company_id, month))

        details = []

        for row in cur.fetchall():
            details.append({
                "emp_code": _pt_clean(row["emp_code"], ""),
                "employee_name": _pt_clean(row["employee_name"], "-"),
                "gender": _pt_clean(row["gender"], ""),
                "designation": _pt_clean(row["designation"], ""),
                "department": _pt_clean(row["department"], ""),
                "gross_earning": _pt_money(row["gross_earning"]),
                "pt_amount": _pt_money(row["pt_amount"])
            })

        data["details"] = details
        data["total_employees"] = len(details)
        data["pt_deducted_employees"] = sum(
            1 for item in details if item["pt_amount"] > 0
        )
        data["total_amount"] = _pt_money(
            sum(item["pt_amount"] for item in details)
        )

        settings = data["settings"]

        slab1_to = _pt_float(
            settings.get("pt_male_slab_1_limit"),
            7500
        )
        slab2_to = _pt_float(
            settings.get("pt_male_slab_2_limit"),
            10000
        )
        slab1_rate = _pt_float(
            settings.get("pt_male_slab_1_amount"),
            0
        )
        slab2_rate = _pt_float(
            settings.get("pt_male_slab_2_amount"),
            175
        )
        above_rate = _pt_float(
            settings.get("pt_above_slab_amount_for_month"),
            200
        )

        slabs = [
            {
                "from": 0,
                "to": slab1_to,
                "rate": slab1_rate
            },
            {
                "from": slab1_to + 1,
                "to": slab2_to,
                "rate": slab2_rate
            },
            {
                "from": slab2_to + 1,
                "to": 999999999,
                "rate": above_rate
            }
        ]

        summary_rows = []

        for slab in slabs:
            slab_items = [
                item
                for item in details
                if slab["from"] <= item["gross_earning"] <= slab["to"]
            ]

            summary_rows.append({
                "from": slab["from"],
                "to": slab["to"],
                "rate": slab["rate"],
                "employees": len(slab_items),
                # Actual saved payroll PT amount is used.
                "amount": _pt_money(
                    sum(item["pt_amount"] for item in slab_items)
                )
            })

        data["summary_rows"] = summary_rows

        return data

    finally:
        conn.close()


@app.route("/professional-tax-report")
@login_required
def professional_tax_report():
    if not require_pro_feature(
        "Upgrade to PRO to view Professional Tax Report."
    ):
        return redirect(url_for("pricing"))

    company_id = current_company_id()

    if not company_id:
        flash(
            "Company not found. Please login again.",
            "danger"
        )
        return redirect(url_for("login"))

    month = request.args.get("month", "").strip()

    try:
        data = get_professional_tax_report_data(
            company_id,
            month
        )

    except Exception as e:
        print("Professional Tax report error:", e)
        flash(
            f"Error loading Professional Tax report: {str(e)}",
            "danger"
        )
        return redirect(url_for("dashboard"))

    return render_template(
        "professional_tax_report.html",
        data=data,
        selected_month=data.get("month", month)
    )


@app.route("/download-professional-tax-pdf")
@login_required
def download_professional_tax_pdf():
    if not require_pro_feature(
        "Upgrade to PRO to download Professional Tax PDF."
    ):
        return redirect(url_for("pricing"))

    company_id = current_company_id()

    if not company_id:
        flash(
            "Company not found. Please login again.",
            "danger"
        )
        return redirect(url_for("login"))

    month = request.args.get("month", "").strip()

    try:
        data = get_professional_tax_report_data(
            company_id,
            month
        )

    except Exception as e:
        print("Professional Tax PDF error:", e)
        flash(
            f"Error creating Professional Tax PDF: {str(e)}",
            "danger"
        )
        return redirect(url_for("professional_tax_report"))

    if not data.get("month"):
        flash(
            "Please select a payroll month.",
            "warning"
        )
        return redirect(url_for("professional_tax_report"))

    if not data.get("details"):
        flash(
            "No payroll data found for selected month.",
            "warning"
        )
        return redirect(
            url_for(
                "professional_tax_report",
                month=data["month"]
            )
        )

    folder = globals().get(
        "PAYSLIP_FOLDER",
        globals().get("UPLOAD_FOLDER", "uploads")
    )
    os.makedirs(folder, exist_ok=True)

    month_safe = _pt_safe_filename(
        data["month"],
        "month"
    )

    file_name = (
        f"professional_tax_report_{month_safe}.pdf"
    )
    file_path = os.path.join(folder, file_name)

    company = data.get("company", {})
    settings = data.get("settings", {})

    company_name = _pt_clean(
        company.get("company_name"),
        "Company Name"
    )
    company_address = _pt_clean(
        company.get("company_address"),
        "-"
    )
    pt_reg_no = _pt_clean(
        company.get("pt_registration_no"),
        "-"
    )
    state = _pt_clean(
        settings.get("pt_state"),
        "Maharashtra"
    )

    c = canvas.Canvas(file_path, pagesize=A4)
    width, height = A4

    left = 38
    right = width - 38
    top = height - 42
    line_height = 13

    font_name = "Courier"
    font_bold = "Courier-Bold"

    def draw_separator(y):
        c.setFont(font_name, 8)
        c.drawString(left, y, "=" * 94)
        return y - line_height

    def draw_detail_header(y, continued=False):
        if continued:
            c.setFont(font_bold, 9)
            c.drawCentredString(
                width / 2,
                y,
                (
                    "PROFESSIONAL TAX REPORT DETAILS "
                    f"CONTINUED - {data['month_label'].upper()}"
                )
            )
            y -= line_height
            y = draw_separator(y)

        c.setFont(font_bold, 8.5)
        c.drawString(left, y, "Sr.")
        c.drawString(left + 30, y, "Employee Name")
        c.drawString(left + 275, y, "Employee ID")
        c.drawRightString(right - 92, y, "Gross Earning")
        c.drawRightString(right, y, "PT Amount")

        y -= line_height
        return draw_separator(y)

    y = top

    # Company header
    c.setFont(font_bold, 9)
    c.drawString(left, y, "Company Name    :")
    c.drawString(
        left + 145,
        y,
        company_name.upper()[:58]
    )
    y -= line_height

    c.drawString(left, y, "Company Address :")
    c.setFont(font_name, 9)
    c.drawString(
        left + 145,
        y,
        company_address[:63]
    )
    y -= line_height

    c.setFont(font_bold, 9)
    c.drawString(left, y, "State           :")
    c.setFont(font_name, 9)
    c.drawString(left + 145, y, state)
    y -= line_height

    c.setFont(font_bold, 9)
    c.drawString(left, y, "PT Reg. No.     :")
    c.setFont(font_name, 9)
    c.drawString(left + 145, y, pt_reg_no)
    y -= line_height * 2

    c.setFont(font_bold, 10)
    c.drawCentredString(
        width / 2,
        y,
        (
            "PROFESSIONAL TAX REPORT FOR "
            f"{data['month_label'].upper()}"
        )
    )
    y -= line_height
    y = draw_separator(y)

    # Slab summary
    c.setFont(font_bold, 8.5)
    c.drawString(left, y, "From")
    c.drawString(left + 85, y, "To")
    c.drawString(left + 170, y, "Rate")
    c.drawString(left + 245, y, "No. of Employees")
    c.drawRightString(right, y, "Amount")
    y -= line_height
    y = draw_separator(y)

    c.setFont(font_name, 8.5)

    for item in data["summary_rows"]:
        c.drawString(left, y, f"{int(item['from'])}")
        c.drawString(left + 85, y, f"{int(item['to'])}")
        c.drawString(left + 170, y, f"{int(item['rate'])}")
        c.drawString(left + 245, y, f"{int(item['employees'])}")
        c.drawRightString(
            right,
            y,
            f"{item['amount']:.2f}"
        )
        y -= line_height

    y = draw_separator(y)

    c.setFont(font_bold, 9)
    c.drawRightString(
        right,
        y,
        f"TOTAL PT AMOUNT: {data['total_amount']:.2f}"
    )
    y -= line_height
    y = draw_separator(y)
    y -= line_height

    # Employee details
    c.setFont(font_bold, 9)
    c.drawCentredString(
        width / 2,
        y,
        "PT REPORT DETAILS"
    )
    y -= line_height
    y = draw_separator(y)
    y = draw_detail_header(y)

    c.setFont(font_name, 8.2)

    for index, item in enumerate(data["details"], start=1):
        if y < 65:
            c.showPage()
            y = top
            y = draw_detail_header(y, continued=True)
            c.setFont(font_name, 8.2)

        c.drawString(left, y, str(index))
        c.drawString(
            left + 30,
            y,
            item["employee_name"][:38]
        )
        c.drawString(
            left + 275,
            y,
            item["emp_code"][:16]
        )
        c.drawRightString(
            right - 92,
            y,
            f"{item['gross_earning']:.2f}"
        )
        c.drawRightString(
            right,
            y,
            f"{item['pt_amount']:.2f}"
        )
        y -= line_height

    if y < 80:
        c.showPage()
        y = top

    y = draw_separator(y)

    c.setFont(font_bold, 9)
    c.drawRightString(
        right,
        y,
        f"TOTAL PT AMOUNT: {data['total_amount']:.2f}"
    )
    y -= line_height * 2

    c.setFont(font_name, 7.5)
    c.drawString(
        left,
        y,
        (
            "This report uses the Professional Tax amount "
            "saved in Payroll History."
        )
    )

    c.save()

    return send_file(
        file_path,
        as_attachment=True,
        download_name=file_name
    )


@app.route("/export-excel")
@login_required
def export_excel():
    if not require_pro_feature("Upgrade to PRO to use Excel export."):
        return redirect(url_for("pricing"))

    month = request.args.get("month", "").strip()
    department = request.args.get("department", "").strip()

    if not month:
        flash("Please select month to export payroll", "warning")
        return redirect(url_for("payroll_history"))

    company_id = current_company_id()

    if not company_id:
        flash("Company not found. Please login again.", "danger")
        return redirect(url_for("login"))

    conn = get_db()

    query = """
        SELECT
            p.emp_code,
            p.employee_name,
            p.role,
            p.department,
            p.gender,

            COALESCE(e.uan_no, '') AS uan_no,
            COALESCE(e.esic_no, '') AS esic_no,
            COALESCE(e.bank_name, '') AS bank_name,
            COALESCE(e.account_no, '') AS account_no,
            COALESCE(e.ifsc_code, '') AS ifsc_code,

            p.month,
            p.monthly_salary,

            COALESCE(a.working_days, 0) AS attendance_working_days,
            COALESCE(a.present_days, 0) AS attendance_present_days,
            COALESCE(a.weekly_off, 0) AS attendance_weekly_off,
            COALESCE(a.paid_leave, 0) AS attendance_paid_leave,
            COALESCE(a.holiday, 0) AS attendance_holiday,
            COALESCE(a.lop_days, 0) AS attendance_lop_days,
            COALESCE(a.paid_days, 0) AS attendance_paid_days,

            COALESCE(p.payable_days, 0) AS payable_days,
            COALESCE(p.paid_leave_days, 0) AS paid_leave_days,
            COALESCE(p.lwp_days, 0) AS lwp_days,
            COALESCE(p.lwp_deduction, 0) AS lwp_deduction,

            COALESCE(p.basic, 0) AS basic,
            COALESCE(p.da, 0) AS da,
            COALESCE(p.hra, 0) AS hra,
            COALESCE(p.special_allowance, 0) AS special_allowance,
            COALESCE(p.other_allowance, 0) AS other_allowance,
            COALESCE(p.gross, 0) AS gross,

            COALESCE(p.overtime_hours, 0) AS overtime_hours,
            COALESCE(p.overtime_amount, 0) AS overtime_amount,

            COALESCE(p.esi_employee, 0) AS esi_employee,
            COALESCE(p.professional_tax, 0) AS professional_tax,
            COALESCE(p.pf_employee, 0) AS pf_employee,
            COALESCE(p.lwf_employee, 0) AS lwf_employee,
            COALESCE(p.tds, 0) AS tds,
            COALESCE(p.manual_deduction, 0) AS manual_deduction,
            COALESCE(p.total_deductions, 0) AS total_deductions,

            COALESCE(p.esi_employer, 0) AS esi_employer,
            COALESCE(p.pf_employer, 0) AS pf_employer,
            COALESCE(p.gratuity, 0) AS gratuity,
            COALESCE(p.lwf_employer, 0) AS lwf_employer,

            COALESCE(p.bonus_ctc, 0) AS bonus_ctc,
            COALESCE(p.festival_bonus, 0) AS festival_bonus,
            COALESCE(p.total_contributions, 0) AS total_contributions,

            COALESCE(p.net_pay, 0) AS net_pay,
            COALESCE(p.monthly_ctc, 0) AS monthly_ctc,
            COALESCE(p.annual_ctc, 0) AS annual_ctc,

            p.created_at

        FROM payroll_history p

        LEFT JOIN employees e
          ON p.company_id = e.company_id
         AND p.emp_code = e.emp_code

        LEFT JOIN attendance a
          ON p.company_id = a.company_id
         AND p.emp_code = a.emp_code
         AND p.month = a.month

        WHERE p.company_id = ?
          AND p.month = ?
          AND p.is_current = 1
    """

    params = [company_id, month]

    if department:
        query += " AND p.department = ?"
        params.append(department)

    query += " ORDER BY p.emp_code"

    raw_df = pd.read_sql_query(query, conn, params=tuple(params))
    conn.close()

    if raw_df.empty:
        flash("No payroll data found for selected month", "warning")
        return redirect(url_for("payroll_history", month=month))

    raw_df["employer_total"] = (
        raw_df["pf_employer"].fillna(0)
        + raw_df["esi_employer"].fillna(0)
        + raw_df["gratuity"].fillna(0)
        + raw_df["lwf_employer"].fillna(0)
    )

    df = raw_df.copy()

    numeric_columns = [
        "monthly_salary",
        "attendance_working_days",
        "attendance_present_days",
        "attendance_weekly_off",
        "attendance_paid_leave",
        "attendance_holiday",
        "attendance_lop_days",
        "attendance_paid_days",
        "payable_days",
        "paid_leave_days",
        "lwp_days",
        "lwp_deduction",
        "basic",
        "da",
        "hra",
        "special_allowance",
        "other_allowance",
        "gross",
        "overtime_hours",
        "overtime_amount",
        "esi_employee",
        "professional_tax",
        "pf_employee",
        "lwf_employee",
        "tds",
        "manual_deduction",
        "total_deductions",
        "esi_employer",
        "pf_employer",
        "gratuity",
        "lwf_employer",
        "employer_total",
        "bonus_ctc",
        "festival_bonus",
        "total_contributions",
        "net_pay",
        "monthly_ctc",
        "annual_ctc"
    ]

    for col in numeric_columns:
        if col in df.columns:
            df[col] = df[col].fillna(0).round().astype(int)

    text_columns = [
        "emp_code",
        "employee_name",
        "role",
        "department",
        "gender",
        "uan_no",
        "esic_no",
        "bank_name",
        "account_no",
        "ifsc_code",
        "month",
        "created_at"
    ]

    for col in text_columns:
        if col in df.columns:
            df[col] = df[col].fillna("").astype(str)

    df = df.rename(columns={
        "emp_code": "Employee Code",
        "employee_name": "Employee Name",
        "role": "Designation",
        "department": "Department",
        "gender": "Gender",

        "uan_no": "UAN No",
        "esic_no": "ESIC No",
        "bank_name": "Bank Name",
        "account_no": "Account No",
        "ifsc_code": "IFSC Code",

        "month": "Month",
        "monthly_salary": "Monthly Salary",

        "attendance_working_days": "Working Days",
        "attendance_present_days": "Present Days",
        "attendance_weekly_off": "Weekly Off",
        "attendance_paid_leave": "Attendance Paid Leave",
        "attendance_holiday": "Holiday",
        "attendance_lop_days": "Attendance LOP Days",
        "attendance_paid_days": "Attendance Paid Days",

        "payable_days": "Final Payable Days",
        "paid_leave_days": "Final Paid Leave Days",
        "lwp_days": "Final LWP Days",
        "lwp_deduction": "LWP Deduction",

        "basic": "Basic",
        "da": "DA",
        "hra": "HRA",
        "special_allowance": "Special Allowance",
        "other_allowance": "Other Allowance",
        "gross": "Gross Salary",

        "overtime_hours": "Overtime Hours",
        "overtime_amount": "Overtime Amount",

        "esi_employee": "ESIC Employee",
        "professional_tax": "Professional Tax",
        "pf_employee": "PF Employee",
        "lwf_employee": "LWF Employee",
        "tds": "TDS",
        "manual_deduction": "Manual Deduction",
        "total_deductions": "Total Deductions",

        "esi_employer": "ESIC Employer",
        "pf_employer": "PF Employer",
        "gratuity": "Gratuity",
        "lwf_employer": "LWF Employer",
        "employer_total": "Employer Total",

        "bonus_ctc": "Bonus CTC",
        "festival_bonus": "Festival Bonus",
        "total_contributions": "Total Contributions",

        "net_pay": "Net Pay",
        "monthly_ctc": "Monthly CTC",
        "annual_ctc": "Annual CTC",

        "created_at": "Created At"
    })

    final_columns = [
        "Employee Code",
        "Employee Name",
        "Designation",
        "Department",
        "Gender",

        "UAN No",
        "ESIC No",
        "Bank Name",
        "Account No",
        "IFSC Code",

        "Month",
        "Monthly Salary",

        "Working Days",
        "Present Days",
        "Weekly Off",
        "Attendance Paid Leave",
        "Holiday",
        "Attendance LOP Days",
        "Attendance Paid Days",

        "Final Payable Days",
        "Final Paid Leave Days",
        "Final LWP Days",
        "LWP Deduction",

        "Basic",
        "DA",
        "HRA",
        "Special Allowance",
        "Other Allowance",
        "Gross Salary",

        "Overtime Hours",
        "Overtime Amount",

        "PF Employee",
        "ESIC Employee",
        "Professional Tax",
        "LWF Employee",
        "TDS",
        "Manual Deduction",
        "Total Deductions",

        "Festival Bonus",
        "Net Pay",

        "PF Employer",
        "ESIC Employer",
        "Gratuity",
        "LWF Employer",
        "Employer Total",

        "Bonus CTC",
        "Total Contributions",
        "Monthly CTC",
        "Annual CTC",

        "Created At"
    ]

    final_columns = [col for col in final_columns if col in df.columns]
    payroll_register_df = df[final_columns]

    # Bank Payment Sheet
    bank_payment_df = raw_df[[
        "emp_code",
        "employee_name",
        "department",
        "bank_name",
        "account_no",
        "ifsc_code",
        "net_pay"
    ]].copy()

    bank_payment_df.insert(0, "Sr No", range(1, len(bank_payment_df) + 1))
    bank_payment_df.insert(1, "Month", month)

    bank_payment_df.rename(columns={
        "emp_code": "Emp Code",
        "employee_name": "Employee Name",
        "department": "Department",
        "bank_name": "Bank Name",
        "account_no": "Account No",
        "ifsc_code": "IFSC Code",
        "net_pay": "Net Pay"
    }, inplace=True)

    bank_payment_df["Net Pay"] = bank_payment_df["Net Pay"].fillna(0).round().astype(int)
    bank_payment_df["Payment Mode"] = "Bank Transfer"
    bank_payment_df["Payment Status"] = "Pending"
    bank_payment_df["Remarks"] = ""

    bank_payment_df = bank_payment_df[[
        "Sr No",
        "Month",
        "Emp Code",
        "Employee Name",
        "Department",
        "Bank Name",
        "Account No",
        "IFSC Code",
        "Net Pay",
        "Payment Mode",
        "Payment Status",
        "Remarks"
    ]]

    total_payment_amount = round(float(bank_payment_df["Net Pay"].fillna(0).sum()))

    total_row = pd.DataFrame([{
        "Sr No": "",
        "Month": "",
        "Emp Code": "",
        "Employee Name": "TOTAL PAYMENT AMOUNT",
        "Department": "",
        "Bank Name": "",
        "Account No": "",
        "IFSC Code": "",
        "Net Pay": total_payment_amount,
        "Payment Mode": "",
        "Payment Status": "",
        "Remarks": ""
    }])

    bank_payment_df = pd.concat([bank_payment_df, total_row], ignore_index=True)

    # Deduction Summary
    deduction_items = [
        ("PF Employee", raw_df["pf_employee"].fillna(0).sum()),
        ("ESIC Employee", raw_df["esi_employee"].fillna(0).sum()),
        ("Professional Tax", raw_df["professional_tax"].fillna(0).sum()),
        ("LWF Employee", raw_df["lwf_employee"].fillna(0).sum()),
        ("TDS", raw_df["tds"].fillna(0).sum()),
        ("Manual Deduction", raw_df["manual_deduction"].fillna(0).sum()),
        ("LWP Deduction", raw_df["lwp_deduction"].fillna(0).sum()),
        ("Total Deductions", raw_df["total_deductions"].fillna(0).sum())
    ]

    deduction_summary_df = pd.DataFrame({
        "Month": [month] * len(deduction_items),
        "Deduction Head": [item[0] for item in deduction_items],
        "Amount": [round(float(item[1])) for item in deduction_items]
    })

    # Employer Contribution Summary
    employer_items = [
        ("PF Employer", raw_df["pf_employer"].fillna(0).sum()),
        ("ESIC Employer", raw_df["esi_employer"].fillna(0).sum()),
        ("Gratuity", raw_df["gratuity"].fillna(0).sum()),
        ("LWF Employer", raw_df["lwf_employer"].fillna(0).sum()),
        ("Employer Total", raw_df["employer_total"].fillna(0).sum()),
        ("Bonus CTC", raw_df["bonus_ctc"].fillna(0).sum()),
        ("Total Contributions", raw_df["total_contributions"].fillna(0).sum()),
        ("Total Monthly CTC", raw_df["monthly_ctc"].fillna(0).sum()),
        ("Total Annual CTC", raw_df["annual_ctc"].fillna(0).sum())
    ]

    employer_summary_df = pd.DataFrame({
        "Month": [month] * len(employer_items),
        "Employer Cost Head": [item[0] for item in employer_items],
        "Amount": [round(float(item[1])) for item in employer_items]
    })

    file_name = f"payroll_{month}.xlsx"
    file_path = os.path.join(UPLOAD_FOLDER, file_name)

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        payroll_register_df.to_excel(writer, index=False, sheet_name="Payroll Register")
        bank_payment_df.to_excel(writer, index=False, sheet_name="Bank Payment Sheet")
        deduction_summary_df.to_excel(writer, index=False, sheet_name="Deduction Summary")
        employer_summary_df.to_excel(writer, index=False, sheet_name="Employer Contribution")

        workbook = writer.book

        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        total_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        warning_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for column_cells in ws.columns:
                max_length = 0
                column_letter = column_cells[0].column_letter

                for cell in column_cells:
                    cell_value = str(cell.value) if cell.value is not None else ""
                    max_length = max(max_length, len(cell_value))

                ws.column_dimensions[column_letter].width = max_length + 3

            # Text formatting
            text_format_headers = [
                "UAN No",
                "ESIC No",
                "Account No",
                "IFSC Code"
            ]

            for col_idx, header_cell in enumerate(ws[1], start=1):
                if header_cell.value in text_format_headers:
                    for row_cells in ws.iter_rows(
                        min_row=2,
                        max_row=ws.max_row,
                        min_col=col_idx,
                        max_col=col_idx
                    ):
                        for cell in row_cells:
                            cell.number_format = "@"

            # Amount formatting by header names
            amount_headers = [
                "Monthly Salary",
                "LWP Deduction",
                "Basic",
                "DA",
                "HRA",
                "Special Allowance",
                "Other Allowance",
                "Gross Salary",
                "Overtime Amount",
                "PF Employee",
                "ESIC Employee",
                "Professional Tax",
                "LWF Employee",
                "TDS",
                "Manual Deduction",
                "Total Deductions",
                "Festival Bonus",
                "Net Pay",
                "PF Employer",
                "ESIC Employer",
                "Gratuity",
                "LWF Employer",
                "Employer Total",
                "Bonus CTC",
                "Total Contributions",
                "Monthly CTC",
                "Annual CTC",
                "Amount"
            ]

            for col_idx, header_cell in enumerate(ws[1], start=1):
                if header_cell.value in amount_headers:
                    for row_cells in ws.iter_rows(
                        min_row=2,
                        max_row=ws.max_row,
                        min_col=col_idx,
                        max_col=col_idx
                    ):
                        for cell in row_cells:
                            cell.number_format = '₹#,##0'

            # Bank Payment Sheet special formatting
            if sheet_name == "Bank Payment Sheet":
                for row_idx in range(2, ws.max_row):
                    bank_name = str(ws.cell(row=row_idx, column=6).value or "").strip()
                    account_no = str(ws.cell(row=row_idx, column=7).value or "").strip()
                    ifsc_code = str(ws.cell(row=row_idx, column=8).value or "").strip()

                    if bank_name == "" or account_no == "" or ifsc_code == "":
                        for col_idx in range(1, ws.max_column + 1):
                            ws.cell(row=row_idx, column=col_idx).fill = warning_fill

                        ws.cell(row=row_idx, column=12).value = "Bank details missing"

                total_row_idx = ws.max_row

                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=total_row_idx, column=col_idx)
                    cell.font = Font(bold=True, color="166534")
                    cell.fill = total_fill

                ws.cell(row=total_row_idx, column=4).value = "TOTAL PAYMENT AMOUNT"
                ws.cell(row=total_row_idx, column=9).number_format = '₹#,##0'

            # Summary sheets total row highlight
            if sheet_name in ["Deduction Summary", "Employer Contribution"]:
                last_row = ws.max_row

                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=last_row, column=col_idx)
                    cell.font = Font(bold=True)

    return send_file(
        file_path,
        as_attachment=True,
        download_name=file_name
    )


@app.route("/hr-audit-report")
@login_required
def hr_audit_report():
    if not require_pro_feature("Upgrade to PRO to download HR Audit Report."):
        return redirect(url_for("pricing"))

    month = request.args.get("month", "").strip()

    if not month:
        flash("Please select month for HR Audit Report.")
        return redirect(url_for("payroll_history"))

    company_id = current_company_id()
    conn = get_db()

    employees_df = pd.read_sql_query("""
        SELECT
            emp_code,
            employee_name,
            role,
            department,
            gender,
            monthly_salary,
            tax_regime,
            other_annual_deductions,
            special_allowance,
            uan_no,
            esic_no,
            bank_name,
            account_no,
            ifsc_code
        FROM employees
        WHERE company_id = ?
        ORDER BY emp_code
    """, conn, params=(company_id,))

    attendance_df = pd.read_sql_query("""
        SELECT
            a.emp_code,
            e.employee_name,
            e.role,
            e.department,
            a.month,
            a.working_days,
            a.present_days,
            COALESCE(a.weekly_off, 0) AS weekly_off,
            COALESCE(a.paid_leave, 0) AS paid_leave,
            COALESCE(a.holiday, 0) AS holiday,
            COALESCE(a.lop_days, 0) AS lop_days,
            COALESCE(a.paid_days, 0) AS paid_days,
            COALESCE(a.overtime_hours, 0) AS overtime_hours,
            COALESCE(a.bonus, 0) AS attendance_bonus,
            COALESCE(a.manual_deduction, 0) AS manual_deduction
        FROM attendance a
        LEFT JOIN employees e
          ON a.company_id = e.company_id
         AND a.emp_code = e.emp_code
        WHERE a.company_id = ?
          AND a.month = ?
        ORDER BY a.emp_code
    """, conn, params=(company_id, month))

    payroll_df = pd.read_sql_query("""
        SELECT
            emp_code,
            employee_name,
            role,
            department,
            gender,
            month,

            COALESCE(payable_days, 0) AS payable_days,
            COALESCE(paid_leave_days, 0) AS paid_leave_days,
            COALESCE(lwp_days, 0) AS lwp_days,
            COALESCE(lwp_deduction, 0) AS lwp_deduction,

            COALESCE(basic, 0) AS basic,
            COALESCE(da, 0) AS da,
            COALESCE(hra, 0) AS hra,
            COALESCE(special_allowance, 0) AS special_allowance,
            COALESCE(other_allowance, 0) AS other_allowance,
            COALESCE(gross, 0) AS gross,

            COALESCE(overtime_hours, 0) AS overtime_hours,
            COALESCE(overtime_amount, 0) AS overtime_amount,

            COALESCE(pf_employee, 0) AS pf_employee,
            COALESCE(esi_employee, 0) AS esi_employee,
            COALESCE(professional_tax, 0) AS professional_tax,
            COALESCE(lwf_employee, 0) AS lwf_employee,
            COALESCE(tds, 0) AS tds,
            COALESCE(manual_deduction, 0) AS manual_deduction,
            COALESCE(total_deductions, 0) AS total_deductions,

            COALESCE(pf_employer, 0) AS pf_employer,
            COALESCE(esi_employer, 0) AS esi_employer,
            COALESCE(gratuity, 0) AS gratuity,
            COALESCE(lwf_employer, 0) AS lwf_employer,

            COALESCE(bonus_ctc, 0) AS bonus_ctc,
            COALESCE(festival_bonus, 0) AS festival_bonus,
            COALESCE(total_contributions, 0) AS total_contributions,

            COALESCE(net_pay, 0) AS net_pay,
            COALESCE(monthly_ctc, 0) AS monthly_ctc,
            COALESCE(annual_ctc, 0) AS annual_ctc,

            created_at
        FROM payroll_history
        WHERE company_id = ?
          AND month = ?
          AND is_current = 1
        ORDER BY emp_code
    """, conn, params=(company_id, month))

    leave_df = pd.read_sql_query("""
        SELECT
            lr.emp_code,
            e.employee_name,
            e.department,
            lr.leave_type,
            lr.start_date,
            lr.end_date,
            lr.total_days,
            lr.status,
            lr.reason
        FROM leave_requests lr
        LEFT JOIN employees e
          ON lr.company_id = e.company_id
         AND lr.emp_code = e.emp_code
        WHERE lr.company_id = ?
          AND (
                substr(lr.start_date, 1, 7) = ?
                OR substr(lr.end_date, 1, 7) = ?
          )
        ORDER BY lr.emp_code, lr.start_date
    """, conn, params=(company_id, month, month))

    conn.close()

    if employees_df.empty:
        flash("No employee data found for HR audit.")
        return redirect(url_for("employees_list"))

    def is_missing(value):
        if pd.isna(value):
            return True
        value = str(value).strip()
        return value == "" or value.lower() in ["nan", "none", "null", "-"]

    employees_df["uan_status"] = employees_df["uan_no"].apply(lambda x: "Missing" if is_missing(x) else "OK")
    employees_df["esic_status"] = employees_df["esic_no"].apply(lambda x: "Missing" if is_missing(x) else "OK")
    employees_df["bank_status"] = employees_df["bank_name"].apply(lambda x: "Missing" if is_missing(x) else "OK")
    employees_df["account_status"] = employees_df["account_no"].apply(lambda x: "Missing" if is_missing(x) else "OK")
    employees_df["ifsc_status"] = employees_df["ifsc_code"].apply(lambda x: "Missing" if is_missing(x) else "OK")
    employees_df["department_status"] = employees_df["department"].apply(lambda x: "Missing" if is_missing(x) else "OK")
    employees_df["gender_status"] = employees_df["gender"].apply(lambda x: "Missing" if is_missing(x) else "OK")

    employees_df["salary_status"] = employees_df["monthly_salary"].apply(
        lambda x: "Invalid" if pd.isna(x) or float(x or 0) <= 0 else "OK"
    )

    attendance_emp_codes = set(attendance_df["emp_code"].astype(str)) if not attendance_df.empty else set()
    payroll_emp_codes = set(payroll_df["emp_code"].astype(str)) if not payroll_df.empty else set()

    employees_df["attendance_status"] = employees_df["emp_code"].astype(str).apply(
        lambda x: "Attendance Missing" if x not in attendance_emp_codes else "OK"
    )

    employees_df["payroll_status"] = employees_df["emp_code"].astype(str).apply(
        lambda x: "Payroll Missing" if x not in payroll_emp_codes else "OK"
    )

    employees_df["audit_month"] = month

    audit_columns = [
        "audit_month",
        "emp_code",
        "employee_name",
        "role",
        "department",
        "gender",
        "monthly_salary",
        "uan_no",
        "esic_no",
        "bank_name",
        "account_no",
        "ifsc_code",
        "uan_status",
        "esic_status",
        "bank_status",
        "account_status",
        "ifsc_status",
        "department_status",
        "gender_status",
        "salary_status",
        "attendance_status",
        "payroll_status"
    ]

    employee_audit_df = employees_df[audit_columns]

    if attendance_df.empty:
        total_working_days = 0
        total_present_days = 0
        total_weekly_off = 0
        total_attendance_paid_leave = 0
        total_holiday = 0
        total_attendance_lop_days = 0
        total_attendance_paid_days = 0
        total_overtime_hours = 0
        total_attendance_bonus = 0
        total_attendance_manual_deduction = 0
    else:
        total_working_days = round(float(attendance_df["working_days"].fillna(0).sum()), 2)
        total_present_days = round(float(attendance_df["present_days"].fillna(0).sum()), 2)
        total_weekly_off = round(float(attendance_df["weekly_off"].fillna(0).sum()), 2)
        total_attendance_paid_leave = round(float(attendance_df["paid_leave"].fillna(0).sum()), 2)
        total_holiday = round(float(attendance_df["holiday"].fillna(0).sum()), 2)
        total_attendance_lop_days = round(float(attendance_df["lop_days"].fillna(0).sum()), 2)
        total_attendance_paid_days = round(float(attendance_df["paid_days"].fillna(0).sum()), 2)
        total_overtime_hours = round(float(attendance_df["overtime_hours"].fillna(0).sum()), 2)
        total_attendance_bonus = round(float(attendance_df["attendance_bonus"].fillna(0).sum()))
        total_attendance_manual_deduction = round(float(attendance_df["manual_deduction"].fillna(0).sum()))

    if payroll_df.empty:
        total_payable_days = 0
        total_paid_leave_days = 0
        total_lwp_days = 0
        total_lwp_deduction = 0
        employees_with_lwp = 0

        total_gross = 0
        total_overtime_amount = 0
        total_festival_bonus = 0
        total_bonus_ctc = 0
        total_pf_employee = 0
        total_esi_employee = 0
        total_professional_tax = 0
        total_lwf_employee = 0
        total_tds = 0
        total_manual_deduction = 0
        total_deductions = 0

        total_pf_employer = 0
        total_esi_employer = 0
        total_gratuity = 0
        total_lwf_employer = 0
        total_employer_contribution = 0
        total_contributions = 0

        total_net_pay = 0
        total_monthly_ctc = 0
        total_annual_ctc = 0
    else:
        payroll_df["employer_total"] = (
            payroll_df["pf_employer"].fillna(0)
            + payroll_df["esi_employer"].fillna(0)
            + payroll_df["gratuity"].fillna(0)
            + payroll_df["lwf_employer"].fillna(0)
        )

        total_payable_days = round(float(payroll_df["payable_days"].fillna(0).sum()), 2)
        total_paid_leave_days = round(float(payroll_df["paid_leave_days"].fillna(0).sum()), 2)
        total_lwp_days = round(float(payroll_df["lwp_days"].fillna(0).sum()), 2)
        total_lwp_deduction = round(float(payroll_df["lwp_deduction"].fillna(0).sum()))
        employees_with_lwp = int((payroll_df["lwp_days"].fillna(0) > 0).sum())

        total_gross = round(float(payroll_df["gross"].fillna(0).sum()))
        total_overtime_amount = round(float(payroll_df["overtime_amount"].fillna(0).sum()))
        total_festival_bonus = round(float(payroll_df["festival_bonus"].fillna(0).sum()))
        total_bonus_ctc = round(float(payroll_df["bonus_ctc"].fillna(0).sum()))

        total_pf_employee = round(float(payroll_df["pf_employee"].fillna(0).sum()))
        total_esi_employee = round(float(payroll_df["esi_employee"].fillna(0).sum()))
        total_professional_tax = round(float(payroll_df["professional_tax"].fillna(0).sum()))
        total_lwf_employee = round(float(payroll_df["lwf_employee"].fillna(0).sum()))
        total_tds = round(float(payroll_df["tds"].fillna(0).sum()))
        total_manual_deduction = round(float(payroll_df["manual_deduction"].fillna(0).sum()))
        total_deductions = round(float(payroll_df["total_deductions"].fillna(0).sum()))

        total_pf_employer = round(float(payroll_df["pf_employer"].fillna(0).sum()))
        total_esi_employer = round(float(payroll_df["esi_employer"].fillna(0).sum()))
        total_gratuity = round(float(payroll_df["gratuity"].fillna(0).sum()))
        total_lwf_employer = round(float(payroll_df["lwf_employer"].fillna(0).sum()))
        total_employer_contribution = round(float(payroll_df["employer_total"].fillna(0).sum()))
        total_contributions = round(float(payroll_df["total_contributions"].fillna(0).sum()))

        total_net_pay = round(float(payroll_df["net_pay"].fillna(0).sum()))
        total_monthly_ctc = round(float(payroll_df["monthly_ctc"].fillna(0).sum()))
        total_annual_ctc = round(float(payroll_df["annual_ctc"].fillna(0).sum()))

    approved_leave_count = 0
    rejected_leave_count = 0
    pending_leave_count = 0

    if not leave_df.empty:
        approved_leave_count = int((leave_df["status"] == "Approved").sum())
        rejected_leave_count = int((leave_df["status"] == "Rejected").sum())
        pending_leave_count = int((leave_df["status"] == "Pending").sum())

    summary_items = [
        ("Total Employees", len(employees_df)),
        ("Attendance Uploaded", len(attendance_emp_codes)),
        ("Payroll Processed", len(payroll_emp_codes)),

        ("Total Working Days", total_working_days),
        ("Total Present Days", total_present_days),
        ("Total Weekly Off", total_weekly_off),
        ("Total Attendance Paid Leave", total_attendance_paid_leave),
        ("Total Holiday", total_holiday),
        ("Total Attendance LOP Days", total_attendance_lop_days),
        ("Total Attendance Paid Days", total_attendance_paid_days),
        ("Total Overtime Hours", total_overtime_hours),
        ("Total Attendance Bonus", total_attendance_bonus),
        ("Total Attendance Manual Deduction", total_attendance_manual_deduction),

        ("Total Final Payable Days", total_payable_days),
        ("Total Final Paid Leave Days", total_paid_leave_days),
        ("Total Final LWP Days", total_lwp_days),
        ("Total LWP Deduction", total_lwp_deduction),
        ("Employees With LWP", employees_with_lwp),

        ("Total Gross Salary", total_gross),
        ("Total Overtime Amount", total_overtime_amount),
        ("Total Festival Bonus Paid", total_festival_bonus),
        ("Total Bonus CTC", total_bonus_ctc),

        ("Total PF Employee", total_pf_employee),
        ("Total ESIC Employee", total_esi_employee),
        ("Total Professional Tax", total_professional_tax),
        ("Total LWF Employee", total_lwf_employee),
        ("Total TDS", total_tds),
        ("Total Manual Deduction", total_manual_deduction),
        ("Total Deductions", total_deductions),

        ("Total PF Employer", total_pf_employer),
        ("Total ESIC Employer", total_esi_employer),
        ("Total Gratuity", total_gratuity),
        ("Total LWF Employer", total_lwf_employer),
        ("Total Employer Contribution", total_employer_contribution),
        ("Total Contributions", total_contributions),

        ("Total Net Pay", total_net_pay),
        ("Total Monthly CTC", total_monthly_ctc),
        ("Total Annual CTC", total_annual_ctc),

        ("Approved Leave Requests", approved_leave_count),
        ("Rejected Leave Requests", rejected_leave_count),
        ("Pending Leave Requests", pending_leave_count),

        ("Missing UAN", int((employees_df["uan_status"] == "Missing").sum())),
        ("Missing ESIC No.", int((employees_df["esic_status"] == "Missing").sum())),
        ("Missing Bank Name", int((employees_df["bank_status"] == "Missing").sum())),
        ("Missing Account No.", int((employees_df["account_status"] == "Missing").sum())),
        ("Missing IFSC", int((employees_df["ifsc_status"] == "Missing").sum())),
        ("Missing Department", int((employees_df["department_status"] == "Missing").sum())),
        ("Missing Gender", int((employees_df["gender_status"] == "Missing").sum())),
        ("Invalid Salary", int((employees_df["salary_status"] == "Invalid").sum())),
        ("Attendance Missing", int((employees_df["attendance_status"] == "Attendance Missing").sum())),
        ("Payroll Missing", int((employees_df["payroll_status"] == "Payroll Missing").sum()))
    ]

    summary_df = pd.DataFrame({
        "Audit Month": [month] * len(summary_items),
        "Audit Item": [item[0] for item in summary_items],
        "Count / Amount": [item[1] for item in summary_items]
    })

    payroll_missing_df = employee_audit_df[
        employee_audit_df["payroll_status"] == "Payroll Missing"
    ]

    attendance_missing_df = employee_audit_df[
        employee_audit_df["attendance_status"] == "Attendance Missing"
    ]

    # Bank Payment Sheet
    # Payroll processed employees ke net pay ko bank payment format me export karega.
    if payroll_df.empty:
        bank_payment_df = pd.DataFrame(columns=[
            "Sr No",
            "Audit Month",
            "Emp Code",
            "Employee Name",
            "Department",
            "Bank Name",
            "Account No",
            "IFSC Code",
            "Net Pay",
            "Payment Mode",
            "Payment Status",
            "Remarks"
        ])
        bank_payment_total = 0

    else:
        bank_master_df = employees_df[[
            "emp_code",
            "bank_name",
            "account_no",
            "ifsc_code"
        ]].copy()

        bank_payment_df = payroll_df[[
            "emp_code",
            "employee_name",
            "department",
            "net_pay"
        ]].copy()

        bank_payment_df = bank_payment_df.merge(
            bank_master_df,
            on="emp_code",
            how="left"
        )

        bank_payment_df.insert(0, "Sr No", range(1, len(bank_payment_df) + 1))
        bank_payment_df.insert(1, "Audit Month", month)

        bank_payment_df.rename(columns={
            "emp_code": "Emp Code",
            "employee_name": "Employee Name",
            "department": "Department",
            "bank_name": "Bank Name",
            "account_no": "Account No",
            "ifsc_code": "IFSC Code",
            "net_pay": "Net Pay"
        }, inplace=True)

        bank_payment_df["Payment Mode"] = "Bank Transfer"
        bank_payment_df["Payment Status"] = "Pending"
        bank_payment_df["Remarks"] = ""

        bank_payment_df = bank_payment_df[[
            "Sr No",
            "Audit Month",
            "Emp Code",
            "Employee Name",
            "Department",
            "Bank Name",
            "Account No",
            "IFSC Code",
            "Net Pay",
            "Payment Mode",
            "Payment Status",
            "Remarks"
        ]]

        bank_payment_total = round(float(bank_payment_df["Net Pay"].fillna(0).sum()))

        # Total payment row at bottom
        total_row = pd.DataFrame([{
            "Sr No": "",
            "Audit Month": "",
            "Emp Code": "",
            "Employee Name": "TOTAL PAYMENT AMOUNT",
            "Department": "",
            "Bank Name": "",
            "Account No": "",
            "IFSC Code": "",
            "Net Pay": bank_payment_total,
            "Payment Mode": "",
            "Payment Status": "",
            "Remarks": ""
        }])

        bank_payment_df = pd.concat(
            [bank_payment_df, total_row],
            ignore_index=True
        )

    # Add bank payment total in audit summary also
    summary_df = pd.concat([
        summary_df,
        pd.DataFrame({
            "Audit Month": [month],
            "Audit Item": ["Bank Payment Total Amount"],
            "Count / Amount": [bank_payment_total]
        })
    ], ignore_index=True)

    file_name = f"hr_audit_report_{month}.xlsx"
    file_path = os.path.join(UPLOAD_FOLDER, file_name)

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, index=False, sheet_name="Audit Summary")
        employee_audit_df.to_excel(writer, index=False, sheet_name="Employee Master Audit")
        attendance_df.to_excel(writer, index=False, sheet_name="Attendance Audit")
        payroll_df.to_excel(writer, index=False, sheet_name="Payroll Audit")
        leave_df.to_excel(writer, index=False, sheet_name="Leave Audit")
        bank_payment_df.to_excel(writer, index=False, sheet_name="Bank Payment Sheet")
        attendance_missing_df.to_excel(writer, index=False, sheet_name="Attendance Missing")
        payroll_missing_df.to_excel(writer, index=False, sheet_name="Payroll Missing")

        workbook = writer.book

        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        missing_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        ok_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        total_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")

        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for column_cells in ws.columns:
                max_length = 0
                column_letter = column_cells[0].column_letter

                for cell in column_cells:
                    cell_value = str(cell.value) if cell.value is not None else ""
                    max_length = max(max_length, len(cell_value))

                    if cell.row != 1:
                        if cell_value in ["Missing", "Invalid", "Attendance Missing", "Payroll Missing", "Pending"]:
                            cell.fill = missing_fill
                        elif cell_value == "OK":
                            cell.fill = ok_fill

                ws.column_dimensions[column_letter].width = max_length + 3

            # Bank Payment Sheet special formatting
            if sheet_name == "Bank Payment Sheet":
                # Net Pay amount formatting
                for row_idx in range(2, ws.max_row + 1):
                    ws.cell(row=row_idx, column=9).number_format = '₹#,##0'

                # Keep bank account number and IFSC as text
                for col in ["G", "H"]:
                    for cell in ws[col]:
                        cell.number_format = "@"

                # Highlight total row
                total_row_idx = ws.max_row

                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=total_row_idx, column=col_idx)
                    cell.font = Font(bold=True, color="166534")
                    cell.fill = total_fill

                ws.cell(row=total_row_idx, column=4).value = "TOTAL PAYMENT AMOUNT"
                ws.cell(row=total_row_idx, column=9).number_format = '₹#,##0'

            # Audit Summary amount formatting
            if sheet_name == "Audit Summary":
                for row_idx in range(2, ws.max_row + 1):
                    item_name = str(ws.cell(row=row_idx, column=2).value or "")

                    if "Amount" in item_name or "Salary" in item_name or "Pay" in item_name or "CTC" in item_name or "Deduction" in item_name:
                        ws.cell(row=row_idx, column=3).number_format = '₹#,##0'

    return send_file(
        file_path,
        as_attachment=True,
        download_name=file_name
    )


# ============================================================
# PAYROLL PRO - ANNUAL HR AUDIT REPORT
#
# New route:
# /annual-hr-audit-report?start_month=2026-04&end_month=2027-03
#
# This route does not replace the existing monthly HR Audit route.
#
# Required imports in app.py (add only when missing):
# import os
# import re
# import datetime
# import pandas as pd
# from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# ============================================================


# ============================================================
# PAYROLL PRO - CUSTOM PERIOD HR AUDIT REPORT
#
# New route:
# /annual-hr-audit-report?start_month_number=1&start_year=2026&end_month_number=6&end_year=2026
#
# This route does not replace the existing monthly HR Audit route.
#
# Required imports in app.py (add only when missing):
# import os
# import re
# import datetime
# import pandas as pd
# from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# ============================================================


# ============================================================
# PAYROLL PRO - CUSTOM PERIOD HR AUDIT REPORT
#
# New route:
# /annual-hr-audit-report?start_month_number=1&start_year=2026&end_month_number=6&end_year=2026
#
# This route does not replace the existing monthly HR Audit route.
#
# Required imports in app.py (add only when missing):
# import os
# import re
# import datetime
# import pandas as pd
# from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# ============================================================


@app.route("/annual-hr-audit-report")
@login_required
def annual_hr_audit_report():
    if not require_pro_feature(
        "Upgrade to PRO to download HR Audit Report."
    ):
        return redirect(url_for("pricing"))

    ensure_employee_personal_columns()

    start_month_number = request.args.get(
        "start_month_number",
        ""
    ).strip()

    start_year = request.args.get(
        "start_year",
        ""
    ).strip()

    end_month_number = request.args.get(
        "end_month_number",
        ""
    ).strip()

    end_year = request.args.get(
        "end_year",
        ""
    ).strip()

    if not all([
        start_month_number,
        start_year,
        end_month_number,
        end_year
    ]):
        flash(
            "Please select From Month, From Year, "
            "To Month and To Year.",
            "warning"
        )
        return redirect(url_for("payroll_history"))

    try:
        start_month_int = int(start_month_number)
        start_year_int = int(start_year)
        end_month_int = int(end_month_number)
        end_year_int = int(end_year)

        if start_month_int not in range(1, 13):
            raise ValueError

        if end_month_int not in range(1, 13):
            raise ValueError

        if not 2000 <= start_year_int <= 2100:
            raise ValueError

        if not 2000 <= end_year_int <= 2100:
            raise ValueError

        start_month = (
            f"{start_year_int}-{start_month_int:02d}"
        )

        end_month = (
            f"{end_year_int}-{end_month_int:02d}"
        )

        start_dt = datetime.datetime.strptime(
            start_month,
            "%Y-%m"
        )

        end_dt = datetime.datetime.strptime(
            end_month,
            "%Y-%m"
        )

    except (TypeError, ValueError):
        flash(
            "Invalid month or year selected.",
            "danger"
        )
        return redirect(url_for("payroll_history"))

    if start_dt > end_dt:
        flash(
            "From Month cannot be after To Month.",
            "danger"
        )
        return redirect(url_for("payroll_history"))

    def build_month_range(start_value, end_value):
        months = []
        current = start_value

        while current <= end_value:
            months.append(current.strftime("%Y-%m"))

            if current.month == 12:
                current = current.replace(
                    year=current.year + 1,
                    month=1
                )
            else:
                current = current.replace(
                    month=current.month + 1
                )

        return months

    def month_end_text(month_value):
        month_dt = datetime.datetime.strptime(
            month_value,
            "%Y-%m"
        )

        if month_dt.month == 12:
            next_month = month_dt.replace(
                year=month_dt.year + 1,
                month=1
            )
        else:
            next_month = month_dt.replace(
                month=month_dt.month + 1
            )

        return (
            next_month - datetime.timedelta(days=1)
        ).strftime("%Y-%m-%d")

    months = build_month_range(start_dt, end_dt)

    company_id = current_company_id()

    if not company_id:
        flash(
            "Company not found. Please login again.",
            "danger"
        )
        return redirect(url_for("login"))

    period_start_date = f"{start_month}-01"
    period_end_date = month_end_text(end_month)

    conn = get_db()

    company_df = pd.read_sql_query("""
        SELECT
            COALESCE(company_name, '') AS company_name,
            COALESCE(address, '') AS company_address,
            COALESCE(email, '') AS company_email,
            COALESCE(phone, '') AS company_phone
        FROM companies
        WHERE id = ?
        LIMIT 1
    """, conn, params=(company_id,))

    employees_df = pd.read_sql_query("""
        SELECT
            emp_code,
            employee_name,
            role,
            department,
            gender,
            monthly_salary,
            tax_regime,
            other_annual_deductions,
            special_allowance,

            COALESCE(mobile_no, '') AS mobile_no,
            COALESCE(email_id, '') AS email_id,
            COALESCE(aadhaar_no, '') AS aadhaar_no,
            COALESCE(pan_no, '') AS pan_no,
            COALESCE(address, '') AS address,

            COALESCE(uan_no, '') AS uan_no,
            COALESCE(esic_no, '') AS esic_no,

            COALESCE(bank_name, '') AS bank_name,
            COALESCE(account_no, '') AS account_no,
            COALESCE(ifsc_code, '') AS ifsc_code

        FROM employees
        WHERE company_id = ?
        ORDER BY emp_code
    """, conn, params=(company_id,))

    attendance_df = pd.read_sql_query("""
        SELECT
            a.emp_code,
            e.employee_name,
            e.role,
            e.department,
            a.month,

            COALESCE(a.working_days, 0) AS working_days,
            COALESCE(a.present_days, 0) AS present_days,
            COALESCE(a.weekly_off, 0) AS weekly_off,
            COALESCE(a.paid_leave, 0) AS paid_leave,
            COALESCE(a.holiday, 0) AS holiday,
            COALESCE(a.lop_days, 0) AS lop_days,
            COALESCE(a.paid_days, 0) AS paid_days,
            COALESCE(a.overtime_hours, 0) AS overtime_hours,
            COALESCE(a.bonus, 0) AS attendance_bonus,
            COALESCE(a.manual_deduction, 0) AS manual_deduction

        FROM attendance AS a

        LEFT JOIN employees AS e
            ON a.company_id = e.company_id
           AND a.emp_code = e.emp_code

        WHERE a.company_id = ?
          AND a.month >= ?
          AND a.month <= ?

        ORDER BY a.month, a.emp_code
    """, conn, params=(
        company_id,
        start_month,
        end_month
    ))

    payroll_df = pd.read_sql_query("""
        SELECT
            emp_code,
            employee_name,
            role,
            department,
            gender,
            month,

            COALESCE(payable_days, 0) AS payable_days,
            COALESCE(paid_leave_days, 0) AS paid_leave_days,
            COALESCE(lwp_days, 0) AS lwp_days,
            COALESCE(lwp_deduction, 0) AS lwp_deduction,

            COALESCE(basic, 0) AS basic,
            COALESCE(da, 0) AS da,
            COALESCE(hra, 0) AS hra,
            COALESCE(special_allowance, 0) AS special_allowance,
            COALESCE(other_allowance, 0) AS other_allowance,
            COALESCE(gross, 0) AS gross,

            COALESCE(overtime_hours, 0) AS overtime_hours,
            COALESCE(overtime_amount, 0) AS overtime_amount,
            COALESCE(festival_bonus, 0) AS festival_bonus,

            COALESCE(pf_employee, 0) AS pf_employee,
            COALESCE(esi_employee, 0) AS esi_employee,
            COALESCE(professional_tax, 0) AS professional_tax,
            COALESCE(lwf_employee, 0) AS lwf_employee,
            COALESCE(tds, 0) AS tds,
            COALESCE(manual_deduction, 0) AS manual_deduction,
            COALESCE(total_deductions, 0) AS total_deductions,
            COALESCE(net_pay, 0) AS net_pay,

            COALESCE(pf_employer, 0) AS pf_employer,
            COALESCE(esi_employer, 0) AS esi_employer,
            COALESCE(gratuity, 0) AS gratuity,
            COALESCE(lwf_employer, 0) AS lwf_employer,
            COALESCE(total_contributions, 0) AS total_contributions,

            COALESCE(bonus_ctc, 0) AS bonus_ctc,
            COALESCE(monthly_ctc, 0) AS monthly_ctc,
            COALESCE(annual_ctc, 0) AS annual_ctc,

            created_at

        FROM payroll_history

        WHERE company_id = ?
          AND month >= ?
          AND month <= ?
          AND is_current = 1

        ORDER BY month, emp_code
    """, conn, params=(
        company_id,
        start_month,
        end_month
    ))

    leave_df = pd.read_sql_query("""
        SELECT
            lr.emp_code,
            e.employee_name,
            e.department,
            lr.leave_type,
            lr.start_date,
            lr.end_date,
            lr.total_days,
            lr.status,
            lr.reason

        FROM leave_requests AS lr

        LEFT JOIN employees AS e
            ON lr.company_id = e.company_id
           AND lr.emp_code = e.emp_code

        WHERE lr.company_id = ?
          AND date(lr.start_date) <= date(?)
          AND date(lr.end_date) >= date(?)

        ORDER BY lr.start_date, lr.emp_code
    """, conn, params=(
        company_id,
        period_end_date,
        period_start_date
    ))

    conn.close()

    if employees_df.empty:
        flash(
            "No employee data found for the selected audit period.",
            "warning"
        )
        return redirect(url_for("employees_list"))

    def clean_text(value):
        if pd.isna(value):
            return ""

        text = str(value).strip()

        if text.lower() in ["nan", "none", "null", "-"]:
            return ""

        if text.endswith(".0") and text[:-2].isdigit():
            text = text[:-2]

        return text

    def only_digits(value):
        return "".join(
            ch for ch in clean_text(value)
            if ch.isdigit()
        )

    def valid_mobile(value):
        digits = only_digits(value)

        if len(digits) == 12 and digits.startswith("91"):
            digits = digits[2:]

        if len(digits) == 11 and digits.startswith("0"):
            digits = digits[1:]

        return (
            len(digits) == 10
            and digits[0] in "6789"
        )

    def valid_email(value):
        email = clean_text(value).lower()

        return (
            email != ""
            and "@" in email
            and "." in email.split("@")[-1]
            and " " not in email
        )

    def valid_aadhaar(value):
        return len(only_digits(value)) == 12

    def valid_pan(value):
        pan = clean_text(value).upper()

        return bool(
            re.fullmatch(
                r"[A-Z]{5}[0-9]{4}[A-Z]",
                pan
            )
        )

    def status_present(value):
        return "OK" if clean_text(value) else "Missing"

    employees_df["mobile_status"] = employees_df[
        "mobile_no"
    ].apply(
        lambda value: "OK"
        if valid_mobile(value)
        else "Missing / Invalid"
    )

    employees_df["email_status"] = employees_df[
        "email_id"
    ].apply(
        lambda value: "OK"
        if valid_email(value)
        else "Missing / Invalid"
    )

    employees_df["aadhaar_status"] = employees_df[
        "aadhaar_no"
    ].apply(
        lambda value: "OK"
        if valid_aadhaar(value)
        else "Missing / Invalid"
    )

    employees_df["pan_status"] = employees_df[
        "pan_no"
    ].apply(
        lambda value: "OK"
        if valid_pan(value)
        else "Missing / Invalid"
    )

    employees_df["address_status"] = employees_df[
        "address"
    ].apply(status_present)

    employees_df["uan_status"] = employees_df[
        "uan_no"
    ].apply(status_present)

    employees_df["esic_status"] = employees_df[
        "esic_no"
    ].apply(status_present)

    employees_df["bank_status"] = employees_df[
        "bank_name"
    ].apply(status_present)

    employees_df["account_status"] = employees_df[
        "account_no"
    ].apply(status_present)

    employees_df["ifsc_status"] = employees_df[
        "ifsc_code"
    ].apply(status_present)

    employees_df["department_status"] = employees_df[
        "department"
    ].apply(status_present)

    employees_df["gender_status"] = employees_df[
        "gender"
    ].apply(status_present)

    employees_df["salary_status"] = employees_df[
        "monthly_salary"
    ].apply(
        lambda value: "Invalid"
        if pd.isna(value) or float(value or 0) <= 0
        else "OK"
    )

    employees_df["whatsapp_status"] = employees_df[
        "mobile_no"
    ].apply(
        lambda value: "Ready"
        if valid_mobile(value)
        else "Mobile Missing"
    )

    # Sensitive values are masked in the exported audit workbook.
    employees_df["aadhaar_no"] = employees_df[
        "aadhaar_no"
    ].apply(
        lambda value: (
            "XXXXXXXX" + only_digits(value)[-4:]
            if len(only_digits(value)) >= 4
            else ""
        )
    )

    employees_df["pan_no"] = employees_df[
        "pan_no"
    ].apply(
        lambda value: (
            clean_text(value).upper()[:5]
            + "****"
            + clean_text(value).upper()[-1:]
            if len(clean_text(value)) == 10
            else ""
        )
    )

    total_employees = len(employees_df)
    employee_codes = employees_df[
        "emp_code"
    ].astype(str).tolist()

    month_labels = {
        month_value: datetime.datetime.strptime(
            month_value,
            "%Y-%m"
        ).strftime("%b-%Y")
        for month_value in months
    }

    month_summary_rows = []
    attendance_missing_rows = []
    payroll_missing_rows = []
    employee_period_rows = []

    attendance_month_sets = {}
    payroll_month_sets = {}

    for month_value in months:
        month_attendance = attendance_df[
            attendance_df["month"].astype(str) == month_value
        ].copy()

        month_payroll = payroll_df[
            payroll_df["month"].astype(str) == month_value
        ].copy()

        attendance_codes = set(
            month_attendance["emp_code"].astype(str)
        ) if not month_attendance.empty else set()

        payroll_codes = set(
            month_payroll["emp_code"].astype(str)
        ) if not month_payroll.empty else set()

        attendance_month_sets[month_value] = attendance_codes
        payroll_month_sets[month_value] = payroll_codes

        month_start_date = datetime.datetime.strptime(
            month_value,
            "%Y-%m"
        )
        month_start_text = month_start_date.strftime(
            "%Y-%m-%d"
        )
        month_end_value = month_end_text(month_value)

        if leave_df.empty:
            month_leave = leave_df.copy()
        else:
            leave_start_dates = pd.to_datetime(
                leave_df["start_date"],
                errors="coerce"
            )
            leave_end_dates = pd.to_datetime(
                leave_df["end_date"],
                errors="coerce"
            )

            month_leave = leave_df[
                (
                    leave_start_dates
                    <= pd.to_datetime(month_end_value)
                )
                &
                (
                    leave_end_dates
                    >= pd.to_datetime(month_start_text)
                )
            ].copy()

        approved_leave_count = 0
        rejected_leave_count = 0
        pending_leave_count = 0

        if not month_leave.empty:
            status_series = (
                month_leave["status"]
                .fillna("")
                .astype(str)
                .str.strip()
                .str.lower()
            )

            approved_leave_count = int(
                (status_series == "approved").sum()
            )
            rejected_leave_count = int(
                (status_series == "rejected").sum()
            )
            pending_leave_count = int(
                (status_series == "pending").sum()
            )

        def column_total(dataframe, column_name, decimals=2):
            if (
                dataframe.empty
                or column_name not in dataframe.columns
            ):
                return 0

            total = pd.to_numeric(
                dataframe[column_name],
                errors="coerce"
            ).fillna(0).sum()

            return round(float(total), decimals)

        employees_with_lwp = 0

        if not month_payroll.empty:
            employees_with_lwp = int(
                (
                    pd.to_numeric(
                        month_payroll["lwp_days"],
                        errors="coerce"
                    ).fillna(0) > 0
                ).sum()
            )

        month_summary_rows.append({
            "Month": month_value,
            "Month Name": month_labels[month_value],

            "Total Employees": total_employees,
            "Attendance Uploaded": len(attendance_codes),
            "Attendance Missing": (
                total_employees - len(attendance_codes)
            ),
            "Payroll Processed": len(payroll_codes),
            "Payroll Missing": (
                total_employees - len(payroll_codes)
            ),

            "Working Days": column_total(
                month_attendance,
                "working_days"
            ),
            "Present Days": column_total(
                month_attendance,
                "present_days"
            ),
            "Weekly Off": column_total(
                month_attendance,
                "weekly_off"
            ),
            "Attendance Paid Leave": column_total(
                month_attendance,
                "paid_leave"
            ),
            "Holiday": column_total(
                month_attendance,
                "holiday"
            ),
            "Attendance LOP Days": column_total(
                month_attendance,
                "lop_days"
            ),
            "Attendance Paid Days": column_total(
                month_attendance,
                "paid_days"
            ),
            "Attendance OT Hours": column_total(
                month_attendance,
                "overtime_hours"
            ),
            "Attendance Bonus": column_total(
                month_attendance,
                "attendance_bonus",
                0
            ),
            "Attendance Manual Deduction": column_total(
                month_attendance,
                "manual_deduction",
                0
            ),

            "Payroll Payable Days": column_total(
                month_payroll,
                "payable_days"
            ),
            "Payroll Paid Leave Days": column_total(
                month_payroll,
                "paid_leave_days"
            ),
            "Payroll LWP Days": column_total(
                month_payroll,
                "lwp_days"
            ),
            "LWP Deduction": column_total(
                month_payroll,
                "lwp_deduction",
                0
            ),
            "Employees With LWP": employees_with_lwp,

            "Gross": column_total(
                month_payroll,
                "gross",
                0
            ),
            "Overtime Amount": column_total(
                month_payroll,
                "overtime_amount",
                0
            ),
            "Festival Bonus": column_total(
                month_payroll,
                "festival_bonus",
                0
            ),
            "Bonus CTC": column_total(
                month_payroll,
                "bonus_ctc",
                0
            ),

            "PF Employee": column_total(
                month_payroll,
                "pf_employee",
                0
            ),
            "ESIC Employee": column_total(
                month_payroll,
                "esi_employee",
                0
            ),
            "Professional Tax": column_total(
                month_payroll,
                "professional_tax",
                0
            ),
            "LWF Employee": column_total(
                month_payroll,
                "lwf_employee",
                0
            ),
            "TDS": column_total(
                month_payroll,
                "tds",
                0
            ),
            "Payroll Manual Deduction": column_total(
                month_payroll,
                "manual_deduction",
                0
            ),
            "Total Deductions": column_total(
                month_payroll,
                "total_deductions",
                0
            ),

            "PF Employer": column_total(
                month_payroll,
                "pf_employer",
                0
            ),
            "ESIC Employer": column_total(
                month_payroll,
                "esi_employer",
                0
            ),
            "Gratuity": column_total(
                month_payroll,
                "gratuity",
                0
            ),
            "LWF Employer": column_total(
                month_payroll,
                "lwf_employer",
                0
            ),
            "Employer Contribution": round(
                column_total(
                    month_payroll,
                    "pf_employer",
                    0
                )
                + column_total(
                    month_payroll,
                    "esi_employer",
                    0
                )
                + column_total(
                    month_payroll,
                    "gratuity",
                    0
                )
                + column_total(
                    month_payroll,
                    "lwf_employer",
                    0
                )
            ),
            "Total Contributions": column_total(
                month_payroll,
                "total_contributions",
                0
            ),

            "Net Pay": column_total(
                month_payroll,
                "net_pay",
                0
            ),
            "Monthly CTC": column_total(
                month_payroll,
                "monthly_ctc",
                0
            ),
            "Bank Payment Total": column_total(
                month_payroll,
                "net_pay",
                0
            ),

            "Approved Leave Requests": approved_leave_count,
            "Rejected Leave Requests": rejected_leave_count,
            "Pending Leave Requests": pending_leave_count
        })

        missing_attendance_codes = [
            code for code in employee_codes
            if code not in attendance_codes
        ]

        missing_payroll_codes = [
            code for code in employee_codes
            if code not in payroll_codes
        ]

        for employee_code in missing_attendance_codes:
            employee_row = employees_df[
                employees_df["emp_code"].astype(str)
                == employee_code
            ].iloc[0]

            attendance_missing_rows.append({
                "Month": month_value,
                "Employee Code": employee_code,
                "Employee Name": employee_row[
                    "employee_name"
                ],
                "Department": employee_row["department"],
                "Status": "Attendance Missing"
            })

        for employee_code in missing_payroll_codes:
            employee_row = employees_df[
                employees_df["emp_code"].astype(str)
                == employee_code
            ].iloc[0]

            payroll_missing_rows.append({
                "Month": month_value,
                "Employee Code": employee_code,
                "Employee Name": employee_row[
                    "employee_name"
                ],
                "Department": employee_row["department"],
                "Status": "Payroll Missing"
            })

    month_summary_df = pd.DataFrame(month_summary_rows)

    for _, employee_row in employees_df.iterrows():
        employee_code = str(employee_row["emp_code"])

        period_row = {
            "Employee Code": employee_code,
            "Employee Name": employee_row["employee_name"],
            "Role": employee_row["role"],
            "Department": employee_row["department"],

            "Attendance Months Uploaded": sum(
                employee_code in attendance_month_sets[
                    month_value
                ]
                for month_value in months
            ),
            "Attendance Months Missing": sum(
                employee_code not in attendance_month_sets[
                    month_value
                ]
                for month_value in months
            ),
            "Payroll Months Processed": sum(
                employee_code in payroll_month_sets[
                    month_value
                ]
                for month_value in months
            ),
            "Payroll Months Missing": sum(
                employee_code not in payroll_month_sets[
                    month_value
                ]
                for month_value in months
            )
        }

        for month_value in months:
            label = month_labels[month_value]

            period_row[
                f"{label} Attendance"
            ] = (
                "OK"
                if employee_code in attendance_month_sets[
                    month_value
                ]
                else "Attendance Missing"
            )

            period_row[
                f"{label} Payroll"
            ] = (
                "OK"
                if employee_code in payroll_month_sets[
                    month_value
                ]
                else "Payroll Missing"
            )

        employee_period_rows.append(period_row)

    employee_period_df = pd.DataFrame(
        employee_period_rows
    )

    attendance_missing_df = pd.DataFrame(
        attendance_missing_rows,
        columns=[
            "Month",
            "Employee Code",
            "Employee Name",
            "Department",
            "Status"
        ]
    )

    payroll_missing_df = pd.DataFrame(
        payroll_missing_rows,
        columns=[
            "Month",
            "Employee Code",
            "Employee Name",
            "Department",
            "Status"
        ]
    )

    # --------------------------------------------------------
    # PERIOD BANK PAYMENT SHEET
    # One row per employee per processed payroll month.
    # --------------------------------------------------------
    bank_payment_columns = [
        "Sr No",
        "Audit Month",
        "Emp Code",
        "Employee Name",
        "Department",
        "Bank Name",
        "Account No",
        "IFSC Code",
        "Net Pay",
        "Payment Mode",
        "Payment Status",
        "Remarks"
    ]

    if payroll_df.empty:
        bank_payment_df = pd.DataFrame(
            columns=bank_payment_columns
        )
        bank_payment_total = 0

    else:
        bank_master_df = employees_df[[
            "emp_code",
            "bank_name",
            "account_no",
            "ifsc_code"
        ]].copy()

        bank_payment_df = payroll_df[[
            "month",
            "emp_code",
            "employee_name",
            "department",
            "net_pay"
        ]].copy()

        bank_payment_df = bank_payment_df.merge(
            bank_master_df,
            on="emp_code",
            how="left"
        )

        bank_payment_df.sort_values(
            by=["month", "emp_code"],
            inplace=True
        )

        bank_payment_df.insert(
            0,
            "Sr No",
            range(1, len(bank_payment_df) + 1)
        )

        bank_payment_df.rename(
            columns={
                "month": "Audit Month",
                "emp_code": "Emp Code",
                "employee_name": "Employee Name",
                "department": "Department",
                "bank_name": "Bank Name",
                "account_no": "Account No",
                "ifsc_code": "IFSC Code",
                "net_pay": "Net Pay"
            },
            inplace=True
        )

        bank_payment_df["Payment Mode"] = "Bank Transfer"
        bank_payment_df["Payment Status"] = "Pending"
        bank_payment_df["Remarks"] = ""

        bank_payment_df = bank_payment_df[
            bank_payment_columns
        ]

        bank_payment_total = round(
            float(
                pd.to_numeric(
                    bank_payment_df["Net Pay"],
                    errors="coerce"
                ).fillna(0).sum()
            )
        )

        total_row = pd.DataFrame([{
            "Sr No": "",
            "Audit Month": "",
            "Emp Code": "",
            "Employee Name": "TOTAL PAYMENT AMOUNT",
            "Department": "",
            "Bank Name": "",
            "Account No": "",
            "IFSC Code": "",
            "Net Pay": bank_payment_total,
            "Payment Mode": "",
            "Payment Status": "",
            "Remarks": ""
        }])

        bank_payment_df = pd.concat(
            [bank_payment_df, total_row],
            ignore_index=True
        )

    company_name = ""

    if not company_df.empty:
        company_name = clean_text(
            company_df.iloc[0]["company_name"]
        )

    def period_total(column_name, decimals=0):
        if (
            month_summary_df.empty
            or column_name not in month_summary_df.columns
        ):
            return 0

        value = pd.to_numeric(
            month_summary_df[column_name],
            errors="coerce"
        ).fillna(0).sum()

        return round(float(value), decimals)

    period_summary_items = [
        (
            "Company",
            company_name or "SmartHire Payroll"
        ),
        (
            "Audit Period",
            (
                f"{start_dt.strftime('%B %Y')} "
                f"to {end_dt.strftime('%B %Y')}"
            )
        ),
        ("Total Months", len(months)),
        ("Current Employee Master Count", total_employees),

        (
            "Employee-Month Attendance Expected",
            total_employees * len(months)
        ),
        (
            "Employee-Month Attendance Uploaded",
            int(period_total("Attendance Uploaded"))
        ),
        (
            "Employee-Month Attendance Missing",
            len(attendance_missing_df)
        ),
        (
            "Employee-Month Payroll Expected",
            total_employees * len(months)
        ),
        (
            "Employee-Month Payroll Processed",
            int(period_total("Payroll Processed"))
        ),
        (
            "Employee-Month Payroll Missing",
            len(payroll_missing_df)
        ),

        ("Total Working Days", period_total("Working Days", 2)),
        ("Total Present Days", period_total("Present Days", 2)),
        ("Total Weekly Off", period_total("Weekly Off", 2)),
        (
            "Total Attendance Paid Leave",
            period_total("Attendance Paid Leave", 2)
        ),
        ("Total Holiday", period_total("Holiday", 2)),
        (
            "Total Attendance LOP Days",
            period_total("Attendance LOP Days", 2)
        ),
        (
            "Total Attendance Paid Days",
            period_total("Attendance Paid Days", 2)
        ),
        (
            "Total Overtime Hours",
            period_total("Attendance OT Hours", 2)
        ),
        (
            "Total Attendance Bonus",
            period_total("Attendance Bonus")
        ),
        (
            "Total Attendance Manual Deduction",
            period_total("Attendance Manual Deduction")
        ),

        (
            "Total Final Payable Days",
            period_total("Payroll Payable Days", 2)
        ),
        (
            "Total Final Paid Leave Days",
            period_total("Payroll Paid Leave Days", 2)
        ),
        (
            "Total Final LWP Days",
            period_total("Payroll LWP Days", 2)
        ),
        (
            "Total LWP Deduction",
            period_total("LWP Deduction")
        ),
        (
            "Employees With LWP",
            int(period_total("Employees With LWP"))
        ),

        ("Period Gross Salary", period_total("Gross")),
        (
            "Period Overtime Amount",
            period_total("Overtime Amount")
        ),
        (
            "Period Festival Bonus Paid",
            period_total("Festival Bonus")
        ),
        ("Period Bonus CTC", period_total("Bonus CTC")),

        ("Total PF Employee", period_total("PF Employee")),
        ("Total ESIC Employee", period_total("ESIC Employee")),
        (
            "Total Professional Tax",
            period_total("Professional Tax")
        ),
        ("Total LWF Employee", period_total("LWF Employee")),
        ("Total TDS", period_total("TDS")),
        (
            "Total Manual Deduction",
            period_total("Payroll Manual Deduction")
        ),
        (
            "Period Total Deductions",
            period_total("Total Deductions")
        ),

        ("Total PF Employer", period_total("PF Employer")),
        ("Total ESIC Employer", period_total("ESIC Employer")),
        ("Total Gratuity", period_total("Gratuity")),
        ("Total LWF Employer", period_total("LWF Employer")),
        (
            "Total Employer Contribution",
            period_total("Employer Contribution")
        ),
        (
            "Total Contributions",
            period_total("Total Contributions")
        ),

        ("Period Net Pay", period_total("Net Pay")),
        ("Period CTC", period_total("Monthly CTC")),
        ("Bank Payment Total Amount", bank_payment_total),

        (
            "Approved Leave Requests",
            int(period_total("Approved Leave Requests"))
        ),
        (
            "Rejected Leave Requests",
            int(period_total("Rejected Leave Requests"))
        ),
        (
            "Pending Leave Requests",
            int(period_total("Pending Leave Requests"))
        ),

        (
            "Employees Missing / Invalid Mobile",
            int((employees_df["mobile_status"] != "OK").sum())
        ),
        (
            "Employees Missing / Invalid Email",
            int((employees_df["email_status"] != "OK").sum())
        ),
        (
            "Employees Missing / Invalid Aadhaar",
            int((employees_df["aadhaar_status"] != "OK").sum())
        ),
        (
            "Employees Missing / Invalid PAN",
            int((employees_df["pan_status"] != "OK").sum())
        ),
        (
            "Employees Missing Address",
            int((employees_df["address_status"] != "OK").sum())
        ),
        (
            "Employees Missing UAN",
            int((employees_df["uan_status"] != "OK").sum())
        ),
        (
            "Employees Missing ESIC No.",
            int((employees_df["esic_status"] != "OK").sum())
        ),
        (
            "Employees Missing Bank Name",
            int((employees_df["bank_status"] != "OK").sum())
        ),
        (
            "Employees Missing Account No.",
            int((employees_df["account_status"] != "OK").sum())
        ),
        (
            "Employees Missing IFSC",
            int((employees_df["ifsc_status"] != "OK").sum())
        ),
        (
            "Employees Missing Department",
            int((employees_df["department_status"] != "OK").sum())
        ),
        (
            "Employees Missing Gender",
            int((employees_df["gender_status"] != "OK").sum())
        ),
        (
            "Employees With Invalid Salary",
            int((employees_df["salary_status"] == "Invalid").sum())
        )
    ]
    period_summary_df = pd.DataFrame(
        period_summary_items,
        columns=["Audit Item", "Value"]
    )

    employee_master_columns = [
        "emp_code",
        "employee_name",
        "role",
        "department",
        "gender",
        "monthly_salary",
        "tax_regime",

        "mobile_no",
        "email_id",
        "aadhaar_no",
        "pan_no",
        "address",

        "uan_no",
        "esic_no",
        "bank_name",
        "account_no",
        "ifsc_code",

        "mobile_status",
        "email_status",
        "aadhaar_status",
        "pan_status",
        "address_status",
        "uan_status",
        "esic_status",
        "bank_status",
        "account_status",
        "ifsc_status",
        "department_status",
        "gender_status",
        "salary_status",
        "whatsapp_status"
    ]

    employee_master_audit_df = employees_df[
        employee_master_columns
    ].copy()

    employee_master_audit_df.rename(
        columns={
            "emp_code": "Employee Code",
            "employee_name": "Employee Name",
            "role": "Role",
            "department": "Department",
            "gender": "Gender",
            "monthly_salary": "Monthly Salary",
            "tax_regime": "Tax Regime",
            "mobile_no": "Mobile No.",
            "email_id": "Email ID",
            "aadhaar_no": "Aadhaar No. (Masked)",
            "pan_no": "PAN No. (Masked)",
            "address": "Address",
            "uan_no": "UAN No.",
            "esic_no": "ESIC No.",
            "bank_name": "Bank Name",
            "account_no": "Account No.",
            "ifsc_code": "IFSC",
            "mobile_status": "Mobile Status",
            "email_status": "Email Status",
            "aadhaar_status": "Aadhaar Status",
            "pan_status": "PAN Status",
            "address_status": "Address Status",
            "uan_status": "UAN Status",
            "esic_status": "ESIC Status",
            "bank_status": "Bank Status",
            "account_status": "Account Status",
            "ifsc_status": "IFSC Status",
            "department_status": "Department Status",
            "gender_status": "Gender Status",
            "salary_status": "Salary Status",
            "whatsapp_status": "WhatsApp Status"
        },
        inplace=True
    )

    file_name = (
        f"hr_audit_{start_month}_to_{end_month}.xlsx"
    )
    file_path = os.path.join(
        UPLOAD_FOLDER,
        file_name
    )

    with pd.ExcelWriter(
        file_path,
        engine="openpyxl"
    ) as writer:
        period_summary_df.to_excel(
            writer,
            index=False,
            sheet_name="Period Summary"
        )

        month_summary_df.to_excel(
            writer,
            index=False,
            sheet_name="Month Wise Summary"
        )

        employee_period_df.to_excel(
            writer,
            index=False,
            sheet_name="Employee Period Audit"
        )

        employee_master_audit_df.to_excel(
            writer,
            index=False,
            sheet_name="Employee Master Audit"
        )

        attendance_df.to_excel(
            writer,
            index=False,
            sheet_name="Attendance Audit"
        )

        payroll_df.to_excel(
            writer,
            index=False,
            sheet_name="Payroll Audit"
        )

        leave_df.to_excel(
            writer,
            index=False,
            sheet_name="Leave Audit"
        )

        bank_payment_df.to_excel(
            writer,
            index=False,
            sheet_name="Bank Payment Sheet"
        )

        attendance_missing_df.to_excel(
            writer,
            index=False,
            sheet_name="Attendance Missing"
        )

        payroll_missing_df.to_excel(
            writer,
            index=False,
            sheet_name="Payroll Missing"
        )

        workbook = writer.book

        header_fill = PatternFill(
            start_color="2563EB",
            end_color="2563EB",
            fill_type="solid"
        )

        missing_fill = PatternFill(
            start_color="FEE2E2",
            end_color="FEE2E2",
            fill_type="solid"
        )

        warning_fill = PatternFill(
            start_color="FEF3C7",
            end_color="FEF3C7",
            fill_type="solid"
        )

        ok_fill = PatternFill(
            start_color="DCFCE7",
            end_color="DCFCE7",
            fill_type="solid"
        )

        total_fill = PatternFill(
            start_color="DBEAFE",
            end_color="DBEAFE",
            fill_type="solid"
        )

        thin_border = Border(
            left=Side(
                style="thin",
                color="E5E7EB"
            ),
            right=Side(
                style="thin",
                color="E5E7EB"
            ),
            top=Side(
                style="thin",
                color="E5E7EB"
            ),
            bottom=Side(
                style="thin",
                color="E5E7EB"
            )
        )

        currency_headers = {
            "Monthly Salary",
            "LWP Deduction",
            "Gross",
            "Overtime Amount",
            "Festival Bonus",
            "Total Deductions",
            "Net Pay",
            "Employer Contributions",
            "Monthly CTC",
            "Bonus CTC",
            "Attendance Bonus",
            "Attendance Manual Deduction",
            "PF Employee",
            "ESIC Employee",
            "Professional Tax",
            "LWF Employee",
            "TDS",
            "Payroll Manual Deduction",
            "PF Employer",
            "ESIC Employer",
            "Gratuity",
            "LWF Employer",
            "Employer Contribution",
            "Total Contributions",
            "Bank Payment Total",
            "basic",
            "da",
            "hra",
            "special_allowance",
            "other_allowance",
            "gross",
            "overtime_amount",
            "festival_bonus",
            "pf_employee",
            "esi_employee",
            "professional_tax",
            "lwf_employee",
            "tds",
            "manual_deduction",
            "total_deductions",
            "net_pay",
            "pf_employer",
            "esi_employer",
            "gratuity",
            "lwf_employer",
            "total_contributions",
            "bonus_ctc",
            "monthly_ctc",
            "annual_ctc"
        }

        text_headers = {
            "Employee Code",
            "emp_code",
            "Mobile No.",
            "mobile_no",
            "Aadhaar No. (Masked)",
            "PAN No. (Masked)",
            "UAN No.",
            "uan_no",
            "ESIC No.",
            "esic_no",
            "Account No.",
            "account_no",
            "IFSC",
            "IFSC Code",
            "ifsc_code",
            "Emp Code"
        }

        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]

            ws.freeze_panes = "A2"
            ws.sheet_view.showGridLines = False

            if ws.max_row >= 1 and ws.max_column >= 1:
                ws.auto_filter.ref = ws.dimensions

            for cell in ws[1]:
                cell.font = Font(
                    bold=True,
                    color="FFFFFF"
                )
                cell.fill = header_fill
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )
                cell.border = thin_border

            ws.row_dimensions[1].height = 34

            header_map = {
                cell.column: str(cell.value or "")
                for cell in ws[1]
            }

            for row_cells in ws.iter_rows(
                min_row=2,
                max_row=ws.max_row
            ):
                for cell in row_cells:
                    cell.border = thin_border
                    cell.alignment = Alignment(
                        vertical="top",
                        wrap_text=True
                    )

                    cell_text = str(
                        cell.value
                        if cell.value is not None
                        else ""
                    )

                    if cell_text in [
                        "Missing",
                        "Missing / Invalid",
                        "Attendance Missing",
                        "Payroll Missing",
                        "Mobile Missing",
                        "Invalid",
                        "Pending"
                    ]:
                        cell.fill = missing_fill

                    elif cell_text == "OK":
                        cell.fill = ok_fill

                    elif cell_text == "Ready":
                        cell.fill = ok_fill

                    elif cell_text == "Rejected":
                        cell.fill = warning_fill

                    header_name = header_map.get(
                        cell.column,
                        ""
                    )

                    if header_name in currency_headers:
                        cell.number_format = (
                            '#,##0.00;[Red]-#,##0.00'
                        )

                    if header_name in text_headers:
                        cell.number_format = "@"

            if sheet_name == "Period Summary":
                for row_number in range(2, ws.max_row + 1):
                    item_name = str(
                        ws.cell(
                            row=row_number,
                            column=1
                        ).value or ""
                    )

                    if any(keyword in item_name for keyword in [
                        "Amount",
                        "Salary",
                        "Pay",
                        "CTC",
                        "Deduction",
                        "Bonus",
                        "Tax",
                        "Contribution",
                        "Gratuity",
                        "PF ",
                        "ESIC ",
                        "LWF ",
                        "TDS"
                    ]):
                        ws.cell(
                            row=row_number,
                            column=2
                        ).number_format = '₹#,##0'

            if sheet_name == "Bank Payment Sheet":
                header_positions = {
                    str(cell.value or ""): cell.column
                    for cell in ws[1]
                }

                net_pay_col = header_positions.get("Net Pay")
                account_col = header_positions.get("Account No")
                ifsc_col = header_positions.get("IFSC Code")

                if net_pay_col:
                    for row_number in range(2, ws.max_row + 1):
                        ws.cell(
                            row=row_number,
                            column=net_pay_col
                        ).number_format = '₹#,##0'

                for column_number in [account_col, ifsc_col]:
                    if column_number:
                        for row_number in range(2, ws.max_row + 1):
                            ws.cell(
                                row=row_number,
                                column=column_number
                            ).number_format = "@"

                if ws.max_row >= 2:
                    total_row_number = ws.max_row

                    for column_number in range(1, ws.max_column + 1):
                        total_cell = ws.cell(
                            row=total_row_number,
                            column=column_number
                        )
                        total_cell.font = Font(
                            bold=True,
                            color="166534"
                        )
                        total_cell.fill = ok_fill

            for column_cells in ws.columns:
                max_length = 0
                column_letter = (
                    column_cells[0].column_letter
                )

                for cell in column_cells:
                    cell_value = str(
                        cell.value
                        if cell.value is not None
                        else ""
                    )

                    max_length = max(
                        max_length,
                        len(cell_value)
                    )

                width = min(
                    max(max_length + 3, 12),
                    38
                )

                ws.column_dimensions[
                    column_letter
                ].width = width

        # Highlight period summary values.
        period_ws = workbook["Period Summary"]

        for row_number in range(
            2,
            period_ws.max_row + 1
        ):
            period_ws.cell(
                row=row_number,
                column=1
            ).font = Font(bold=True)

            period_ws.cell(
                row=row_number,
                column=2
            ).fill = total_fill

        # Keep the period summary as the first visible sheet.
        workbook.active = workbook.sheetnames.index(
            "Period Summary"
        )

    return send_file(
        file_path,
        as_attachment=True,
        download_name=file_name
    )



@app.route("/download-bank-payment-sheet")
@login_required
def download_bank_payment_sheet():
    if not require_pro_feature("Upgrade to PRO to download Bank Payment Sheet."):
        return redirect(url_for("pricing"))

    month = request.args.get("month", "").strip()

    if not month:
        flash("Please select month to download Bank Payment Sheet.", "warning")
        return redirect(url_for("payroll_history"))

    company_id = current_company_id()

    if not company_id:
        flash("Company not found. Please login again.", "danger")
        return redirect(url_for("login"))

    conn = get_db()

    payroll_df = pd.read_sql_query("""
        SELECT
            p.emp_code,
            p.employee_name,
            p.department,
            COALESCE(p.net_pay, 0) AS net_pay,

            COALESCE(e.bank_name, '') AS bank_name,
            COALESCE(e.account_no, '') AS account_no,
            COALESCE(e.ifsc_code, '') AS ifsc_code

        FROM payroll_history p
        LEFT JOIN employees e
          ON p.company_id = e.company_id
         AND p.emp_code = e.emp_code

        WHERE p.company_id = ?
          AND p.month = ?
          AND p.is_current = 1

        ORDER BY p.emp_code
    """, conn, params=(company_id, month))

    conn.close()

    if payroll_df.empty:
        flash("No payroll data found for selected month.", "warning")
        return redirect(url_for("payroll_history", month=month))

    bank_payment_df = payroll_df.copy()

    bank_payment_df.insert(0, "Sr No", range(1, len(bank_payment_df) + 1))
    bank_payment_df.insert(1, "Month", month)

    bank_payment_df.rename(columns={
        "emp_code": "Emp Code",
        "employee_name": "Employee Name",
        "department": "Department",
        "bank_name": "Bank Name",
        "account_no": "Account No",
        "ifsc_code": "IFSC Code",
        "net_pay": "Net Pay"
    }, inplace=True)

    bank_payment_df["Payment Mode"] = "Bank Transfer"
    bank_payment_df["Payment Status"] = "Pending"
    bank_payment_df["Remarks"] = ""

    bank_payment_df = bank_payment_df[[
        "Sr No",
        "Month",
        "Emp Code",
        "Employee Name",
        "Department",
        "Bank Name",
        "Account No",
        "IFSC Code",
        "Net Pay",
        "Payment Mode",
        "Payment Status",
        "Remarks"
    ]]

    total_payment = round(float(bank_payment_df["Net Pay"].fillna(0).sum()))

    total_row = pd.DataFrame([{
        "Sr No": "",
        "Month": "",
        "Emp Code": "",
        "Employee Name": "TOTAL PAYMENT AMOUNT",
        "Department": "",
        "Bank Name": "",
        "Account No": "",
        "IFSC Code": "",
        "Net Pay": total_payment,
        "Payment Mode": "",
        "Payment Status": "",
        "Remarks": ""
    }])

    bank_payment_df = pd.concat([bank_payment_df, total_row], ignore_index=True)

    file_name = f"bank_payment_sheet_{month}.xlsx"
    file_path = os.path.join(UPLOAD_FOLDER, file_name)

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        bank_payment_df.to_excel(writer, index=False, sheet_name="Bank Payment Sheet")

        workbook = writer.book
        ws = workbook["Bank Payment Sheet"]

        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        total_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        warning_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

        thin_border = Border(
            left=Side(style="thin", color="E2E8F0"),
            right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin", color="E2E8F0"),
            bottom=Side(style="thin", color="E2E8F0")
        )

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row_cells:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

                if str(cell.value or "").strip() in ["Pending", "Missing", ""]:
                    pass

        # Net Pay amount formatting
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=9).number_format = '₹#,##0'

        # Keep Account No and IFSC as text
        for col in ["G", "H"]:
            for cell in ws[col]:
                cell.number_format = "@"

        # Highlight missing bank details
        for row_idx in range(2, ws.max_row):
            bank_name = str(ws.cell(row=row_idx, column=6).value or "").strip()
            account_no = str(ws.cell(row=row_idx, column=7).value or "").strip()
            ifsc_code = str(ws.cell(row=row_idx, column=8).value or "").strip()

            if bank_name == "" or account_no == "" or ifsc_code == "":
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = warning_fill

                ws.cell(row=row_idx, column=12).value = "Bank details missing"

        # Highlight total row
        total_row_idx = ws.max_row

        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=total_row_idx, column=col_idx)
            cell.font = Font(bold=True, color="166534")
            cell.fill = total_fill
            cell.border = thin_border

        ws.cell(row=total_row_idx, column=4).value = "TOTAL PAYMENT AMOUNT"
        ws.cell(row=total_row_idx, column=9).number_format = '₹#,##0'

        # Auto width
        for column_cells in ws.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter

            for cell in column_cells:
                cell_value = str(cell.value) if cell.value is not None else ""
                max_length = max(max_length, len(cell_value))

            ws.column_dimensions[column_letter].width = max_length + 4

    return send_file(
        file_path,
        as_attachment=True,
        download_name=file_name
    )


# ============================================================
# SMART HIRE PAYROLL - FULL & FINAL SETTLEMENT COMPLETE UPDATE
#
# Replace the existing Full & Final routes with this complete block.
#
# Required imports in app.py (add only when missing):
#
# import os
# import calendar
# import datetime
# import pandas as pd
#
# from flask import (
#     request, redirect, url_for, flash, render_template,
#     send_file, jsonify
# )
#
# from openpyxl.styles import (
#     Font, PatternFill, Alignment, Border, Side
# )
#
# from reportlab.pdfgen import canvas
# from reportlab.lib.pagesizes import A4
#
# Existing application helpers used:
# - get_db()
# - current_company_id()
# - get_compliance_settings(company_id)
# - require_pro_feature(...)
# - login_required
# - UPLOAD_FOLDER
# - PAYSLIP_FOLDER
# ============================================================


def ensure_full_final_columns():
    """
    Safe migration for existing databases.
    It never deletes existing settlement data.
    """
    conn = get_db()
    cur = conn.cursor()

    cur.execute("PRAGMA table_info(full_final_settlements)")
    existing_columns = {
        row["name"] if hasattr(row, "keys") else row[1]
        for row in cur.fetchall()
    }

    required_columns = {
        "salary_days_policy": "TEXT DEFAULT 'fixed_30'",
        "salary_days_basis": "REAL DEFAULT 30",
        "attendance_source": "TEXT DEFAULT '-'",
        "leave_encashment_basis": "TEXT DEFAULT 'salary'",
        "leave_encashment_rate": "REAL DEFAULT 0",

        "pf_employee": "REAL DEFAULT 0",
        "esi_employee": "REAL DEFAULT 0",
        "professional_tax": "REAL DEFAULT 0",
        "lwf_employee": "REAL DEFAULT 0",
        "tds": "REAL DEFAULT 0",
        "statutory_deductions": "REAL DEFAULT 0",

        "payment_status": "TEXT DEFAULT 'Pending'",
        "payment_date": "TEXT DEFAULT ''"
    }

    for column_name, column_definition in required_columns.items():
        if column_name not in existing_columns:
            cur.execute(
                f"""
                ALTER TABLE full_final_settlements
                ADD COLUMN {column_name} {column_definition}
                """
            )

    cur.execute("PRAGMA table_info(employees)")
    employee_columns = {
        row["name"] if hasattr(row, "keys") else row[1]
        for row in cur.fetchall()
    }

    if "date_of_exit" not in employee_columns:
        cur.execute("""
            ALTER TABLE employees
            ADD COLUMN date_of_exit TEXT
        """)

    conn.commit()
    conn.close()


def _fnf_to_float(value, default=0):
    try:
        if value is None or str(value).strip() == "":
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def _fnf_money_round(value):
    try:
        return round(float(value or 0))
    except (TypeError, ValueError):
        return 0


def _fnf_clean_text(value, default="-"):
    value = str(value or "").strip()

    if not value or value.lower() in {"nan", "none", "null"}:
        return default

    return value


def _fnf_clean_reason(value):
    value = _fnf_clean_text(value, "-")

    if value == "-":
        return "-"

    return value[:1].upper() + value[1:]


def _fnf_safe_filename(value, default="file"):
    value = _fnf_clean_text(value, default)

    for character in [
        " ", "/", "\\", ":", "*", "?", '"', "<", ">", "|"
    ]:
        value = value.replace(character, "_")

    return value


def _fnf_row_value(row, key, default=None):
    try:
        return row[key]
    except Exception:
        return default


def _fnf_salary_policy_label(policy):
    labels = {
        "attendance": "Attendance Based",
        "fixed_26": "Fixed 26 Days",
        "fixed_30": "Fixed 30 Days",
        "calendar": "Calendar Days",
        "custom": "Custom Salary Days"
    }

    return labels.get(policy, "Fixed 30 Days")


def _fnf_get_calculation_context(
    company_id,
    emp_code,
    settlement_month,
    last_working_day,
    paid_days_override=None
):
    if not emp_code:
        raise ValueError("Employee is required.")

    try:
        last_working_date = datetime.datetime.strptime(
            last_working_day,
            "%Y-%m-%d"
        )
    except (TypeError, ValueError):
        raise ValueError("Last Working Day must be a valid date.")

    try:
        settlement_date = datetime.datetime.strptime(
            settlement_month,
            "%Y-%m"
        )
    except (TypeError, ValueError):
        raise ValueError("Settlement Month must be in YYYY-MM format.")

    if last_working_date.strftime("%Y-%m") != settlement_month:
        raise ValueError(
            "Settlement Month must match the Last Working Day month."
        )

    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            emp_code,
            employee_name,
            role,
            department,
            monthly_salary,
            COALESCE(date_of_exit, '') AS date_of_exit
        FROM employees
        WHERE company_id = ?
          AND emp_code = ?
        LIMIT 1
    """, (company_id, emp_code))

    employee = cur.fetchone()

    if not employee:
        conn.close()
        raise ValueError("Employee not found.")

    monthly_salary = _fnf_to_float(
        employee["monthly_salary"],
        0
    )

    if monthly_salary <= 0:
        conn.close()
        raise ValueError(
            "Employee monthly salary is invalid. "
            "Please update Employee Master."
        )

    cur.execute("""
        SELECT COALESCE(paid_leave, 0) AS paid_leave
        FROM leave_balances
        WHERE company_id = ?
          AND emp_code = ?
        LIMIT 1
    """, (company_id, emp_code))

    leave_row = cur.fetchone()
    leave_balance = max(
        0,
        _fnf_to_float(
            leave_row["paid_leave"] if leave_row else 0,
            0
        )
    )

    cur.execute("""
        SELECT
            COALESCE(working_days, 0) AS working_days,
            COALESCE(paid_days, 0) AS paid_days,
            COALESCE(lop_days, 0) AS lop_days,
            COALESCE(paid_leave, 0) AS paid_leave,
            COALESCE(weekly_off, 0) AS weekly_off,
            COALESCE(holiday, 0) AS holiday
        FROM attendance
        WHERE company_id = ?
          AND emp_code = ?
          AND month = ?
        LIMIT 1
    """, (company_id, emp_code, settlement_month))

    attendance = cur.fetchone()

    cur.execute("""
        SELECT
            COALESCE(pf_employee, 0) AS pf_employee,
            COALESCE(esi_employee, 0) AS esi_employee,
            COALESCE(professional_tax, 0) AS professional_tax,
            COALESCE(lwf_employee, 0) AS lwf_employee,
            COALESCE(tds, 0) AS tds
        FROM payroll_history
        WHERE company_id = ?
          AND emp_code = ?
          AND month = ?
          AND is_current = 1
        LIMIT 1
    """, (company_id, emp_code, settlement_month))

    payroll = cur.fetchone()
    conn.close()

    settings = get_compliance_settings(company_id)

    salary_days_policy = (
        str(
            _fnf_row_value(
                settings,
                "salary_days_policy",
                "fixed_30"
            )
            or "fixed_30"
        )
        .strip()
        .lower()
    )

    custom_salary_days = _fnf_to_float(
        _fnf_row_value(
            settings,
            "custom_salary_days",
            30
        ),
        30
    )

    year = settlement_date.year
    month = settlement_date.month
    calendar_days = calendar.monthrange(year, month)[1]

    if salary_days_policy == "fixed_26":
        salary_days_basis = 26

    elif salary_days_policy == "fixed_30":
        salary_days_basis = 30

    elif salary_days_policy == "calendar":
        salary_days_basis = calendar_days

    elif salary_days_policy == "custom":
        salary_days_basis = (
            custom_salary_days
            if custom_salary_days > 0
            else 30
        )

    elif salary_days_policy == "attendance":
        attendance_working_days = _fnf_to_float(
            attendance["working_days"]
            if attendance
            else 0,
            0
        )

        if not attendance or attendance_working_days <= 0:
            raise ValueError(
                "Attendance-based salary policy is active. "
                "Please upload attendance for this employee and "
                "settlement month before creating Full & Final."
            )

        salary_days_basis = attendance_working_days

    else:
        salary_days_policy = "fixed_30"
        salary_days_basis = 30

    attendance_available = attendance is not None

    if attendance_available:
        paid_days = max(
            0,
            _fnf_to_float(attendance["paid_days"], 0)
        )
        attendance_source = "Attendance Register"

    else:
        paid_days = min(
            float(last_working_date.day),
            float(salary_days_basis)
        )
        attendance_source = "Last Working Day Fallback"

    if paid_days_override not in [None, ""]:
        override_value = _fnf_to_float(
            paid_days_override,
            -1
        )

        if override_value < 0:
            raise ValueError(
                "Paid Days Override cannot be negative."
            )

        if override_value > float(salary_days_basis):
            raise ValueError(
                "Paid Days Override cannot exceed "
                f"Salary Days Basis ({salary_days_basis:g})."
            )

        paid_days = override_value
        attendance_source = "Manual HR Override"

    paid_days = max(
        0,
        min(float(paid_days), float(salary_days_basis))
    )

    per_day_salary = (
        monthly_salary / salary_days_basis
        if salary_days_basis > 0
        else 0
    )

    earned_salary = _fnf_money_round(
        per_day_salary * paid_days
    )

    suggested_deductions = {
        "pf_employee": _fnf_money_round(
            payroll["pf_employee"] if payroll else 0
        ),
        "esi_employee": _fnf_money_round(
            payroll["esi_employee"] if payroll else 0
        ),
        "professional_tax": _fnf_money_round(
            payroll["professional_tax"] if payroll else 0
        ),
        "lwf_employee": _fnf_money_round(
            payroll["lwf_employee"] if payroll else 0
        ),
        "tds": _fnf_money_round(
            payroll["tds"] if payroll else 0
        )
    }

    return {
        "employee": employee,
        "last_working_date": last_working_date,
        "monthly_salary": monthly_salary,
        "leave_balance": leave_balance,

        "salary_days_policy": salary_days_policy,
        "salary_policy_label": _fnf_salary_policy_label(
            salary_days_policy
        ),
        "salary_days_basis": float(salary_days_basis),
        "attendance_available": attendance_available,
        "attendance_source": attendance_source,
        "paid_days": float(paid_days),

        "per_day_salary": per_day_salary,
        "earned_salary": earned_salary,

        "suggested_deductions": suggested_deductions,
        "payroll_deduction_source": (
            "Payroll History"
            if payroll
            else "No Payroll Record"
        )
    }


@app.route("/fnf-preview-data")
@login_required
def fnf_preview_data():
    if not require_pro_feature(
        "Upgrade to PRO to use Full & Final Settlement."
    ):
        return jsonify({
            "ok": False,
            "message": "Full & Final Settlement is not available."
        }), 403

    company_id = current_company_id()

    if not company_id:
        return jsonify({
            "ok": False,
            "message": "Company not found. Please login again."
        }), 401

    ensure_full_final_columns()

    try:
        context = _fnf_get_calculation_context(
            company_id=company_id,
            emp_code=request.args.get(
                "emp_code",
                ""
            ).strip(),
            settlement_month=request.args.get(
                "settlement_month",
                ""
            ).strip(),
            last_working_day=request.args.get(
                "last_working_day",
                ""
            ).strip(),
            paid_days_override=request.args.get(
                "paid_days_override",
                ""
            ).strip()
        )

        return jsonify({
            "ok": True,
            "employee_name": context[
                "employee"
            ]["employee_name"],
            "department": context[
                "employee"
            ]["department"],
            "monthly_salary": context["monthly_salary"],
            "leave_balance": context["leave_balance"],

            "salary_days_policy": context[
                "salary_days_policy"
            ],
            "salary_policy_label": context[
                "salary_policy_label"
            ],
            "salary_days_basis": context[
                "salary_days_basis"
            ],
            "attendance_available": context[
                "attendance_available"
            ],
            "attendance_source": context[
                "attendance_source"
            ],
            "paid_days": context["paid_days"],

            "per_day_salary": context["per_day_salary"],
            "earned_salary": context["earned_salary"],

            "suggested_deductions": context[
                "suggested_deductions"
            ],
            "payroll_deduction_source": context[
                "payroll_deduction_source"
            ]
        })

    except ValueError as error:
        return jsonify({
            "ok": False,
            "message": str(error)
        }), 400

    except Exception as error:
        return jsonify({
            "ok": False,
            "message": (
                "Unable to calculate settlement preview: "
                + str(error)
            )
        }), 500


@app.route("/full-and-final", methods=["GET", "POST"])
@login_required
def full_and_final():
    if not require_pro_feature(
        "Upgrade to PRO to use Full & Final Settlement."
    ):
        return redirect(url_for("pricing"))

    company_id = current_company_id()

    if not company_id:
        flash(
            "Company not found. Please login again.",
            "danger"
        )
        return redirect(url_for("login"))

    ensure_full_final_columns()

    if request.method == "POST":
        conn = get_db()
        cur = conn.cursor()

        try:
            emp_code = request.form.get(
                "emp_code",
                ""
            ).strip()

            last_working_day = request.form.get(
                "last_working_day",
                ""
            ).strip()

            settlement_month = request.form.get(
                "settlement_month",
                ""
            ).strip()

            paid_days_override = request.form.get(
                "paid_days_override",
                ""
            ).strip()

            reason = _fnf_clean_reason(
                request.form.get("reason", "")
            )

            remarks = _fnf_clean_text(
                request.form.get("remarks", ""),
                "-"
            )

            apply_leave_encashment = (
                request.form.get(
                    "apply_leave_encashment",
                    "no"
                )
                .strip()
                .lower()
            )

            leave_encashment_basis = (
                request.form.get(
                    "leave_encashment_basis",
                    "salary"
                )
                .strip()
                .lower()
            )

            manual_leave_rate = _fnf_to_float(
                request.form.get(
                    "manual_leave_rate"
                ),
                0
            )

            bonus_payable = _fnf_to_float(
                request.form.get("bonus_payable"),
                0
            )

            gratuity_payable = _fnf_to_float(
                request.form.get("gratuity_payable"),
                0
            )

            other_earnings = _fnf_to_float(
                request.form.get("other_earnings"),
                0
            )

            notice_recovery = _fnf_to_float(
                request.form.get("notice_recovery"),
                0
            )

            loan_recovery = _fnf_to_float(
                request.form.get("loan_recovery"),
                0
            )

            advance_recovery = _fnf_to_float(
                request.form.get("advance_recovery"),
                0
            )

            other_deductions = _fnf_to_float(
                request.form.get("other_deductions"),
                0
            )

            pf_employee = _fnf_to_float(
                request.form.get("pf_employee"),
                0
            )

            esi_employee = _fnf_to_float(
                request.form.get("esi_employee"),
                0
            )

            professional_tax = _fnf_to_float(
                request.form.get("professional_tax"),
                0
            )

            lwf_employee = _fnf_to_float(
                request.form.get("lwf_employee"),
                0
            )

            tds = _fnf_to_float(
                request.form.get("tds"),
                0
            )

            payment_status = (
                request.form.get(
                    "payment_status",
                    "Pending"
                )
                .strip()
                .title()
            )

            payment_date = request.form.get(
                "payment_date",
                ""
            ).strip()

            errors = []

            if apply_leave_encashment not in {
                "yes",
                "no"
            }:
                apply_leave_encashment = "no"

            if leave_encashment_basis not in {
                "salary",
                "manual"
            }:
                leave_encashment_basis = "salary"

            if payment_status not in {
                "Pending",
                "Paid",
                "Hold"
            }:
                errors.append(
                    "Invalid payment status selected."
                )

            monetary_values = {
                "Manual Leave Rate": manual_leave_rate,
                "Bonus Payable": bonus_payable,
                "Gratuity Payable": gratuity_payable,
                "Other Earnings": other_earnings,
                "Notice Recovery": notice_recovery,
                "Loan Recovery": loan_recovery,
                "Advance Recovery": advance_recovery,
                "Other Deductions": other_deductions,
                "PF Employee": pf_employee,
                "ESIC Employee": esi_employee,
                "Professional Tax": professional_tax,
                "LWF Employee": lwf_employee,
                "TDS": tds
            }

            for label, value in monetary_values.items():
                if value < 0:
                    errors.append(
                        f"{label} cannot be negative."
                    )

                if value > 100000000:
                    errors.append(
                        f"{label} amount is too high. "
                        "Please check."
                    )

            if (
                apply_leave_encashment == "yes"
                and leave_encashment_basis == "manual"
                and manual_leave_rate <= 0
            ):
                errors.append(
                    "Enter a valid Manual Leave Encashment "
                    "Per-Day Rate."
                )

            if payment_status == "Paid":
                if not payment_date:
                    payment_date = (
                        datetime.datetime.now()
                        .strftime("%Y-%m-%d")
                    )

                try:
                    datetime.datetime.strptime(
                        payment_date,
                        "%Y-%m-%d"
                    )
                except ValueError:
                    errors.append(
                        "Payment Date must be a valid date."
                    )

            else:
                payment_date = ""

            if errors:
                flash(" ".join(errors), "danger")
                conn.close()
                return redirect(url_for("full_and_final"))

            context = _fnf_get_calculation_context(
                company_id=company_id,
                emp_code=emp_code,
                settlement_month=settlement_month,
                last_working_day=last_working_day,
                paid_days_override=paid_days_override
            )

            cur.execute("""
                SELECT id
                FROM full_final_settlements
                WHERE company_id = ?
                  AND emp_code = ?
                  AND settlement_month = ?
                LIMIT 1
            """, (
                company_id,
                emp_code,
                settlement_month
            ))

            if cur.fetchone():
                conn.close()
                flash(
                    "Full & Final settlement already exists "
                    "for this employee and settlement month. "
                    "Delete the existing settlement before "
                    "creating a new one.",
                    "warning"
                )
                return redirect(url_for("full_and_final"))

            leave_balance = context["leave_balance"]

            if apply_leave_encashment == "yes":
                if leave_encashment_basis == "manual":
                    leave_encashment_rate = (
                        manual_leave_rate
                    )
                else:
                    leave_encashment_rate = context[
                        "per_day_salary"
                    ]

                leave_encashment = _fnf_money_round(
                    leave_encashment_rate
                    * leave_balance
                )

            else:
                leave_encashment_rate = 0
                leave_encashment = 0

            statutory_deductions = _fnf_money_round(
                pf_employee
                + esi_employee
                + professional_tax
                + lwf_employee
                + tds
            )

            total_earnings = _fnf_money_round(
                context["earned_salary"]
                + leave_encashment
                + bonus_payable
                + gratuity_payable
                + other_earnings
            )

            total_deductions = _fnf_money_round(
                notice_recovery
                + loan_recovery
                + advance_recovery
                + other_deductions
                + statutory_deductions
            )

            final_payable = _fnf_money_round(
                total_earnings - total_deductions
            )

            now = (
                datetime.datetime.now()
                .strftime("%Y-%m-%d %H:%M:%S")
            )

            employee = context["employee"]

            cur.execute("""
                INSERT INTO full_final_settlements (
                    company_id,
                    emp_code,
                    employee_name,
                    role,
                    department,
                    last_working_day,
                    settlement_month,

                    monthly_salary,
                    paid_days,
                    earned_salary,

                    leave_balance,
                    leave_encashment,
                    bonus_payable,
                    gratuity_payable,
                    other_earnings,

                    notice_recovery,
                    loan_recovery,
                    advance_recovery,
                    other_deductions,

                    total_earnings,
                    total_deductions,
                    final_payable,

                    reason,
                    remarks,
                    created_at,

                    salary_days_policy,
                    salary_days_basis,
                    attendance_source,
                    leave_encashment_basis,
                    leave_encashment_rate,

                    pf_employee,
                    esi_employee,
                    professional_tax,
                    lwf_employee,
                    tds,
                    statutory_deductions,

                    payment_status,
                    payment_date
                )
                VALUES (
                    ?, ?, ?, ?, ?, ?, ?,
                    ?, ?, ?,
                    ?, ?, ?, ?, ?,
                    ?, ?, ?, ?,
                    ?, ?, ?,
                    ?, ?, ?,
                    ?, ?, ?, ?, ?,
                    ?, ?, ?, ?, ?, ?,
                    ?, ?
                )
            """, (
                company_id,
                employee["emp_code"],
                employee["employee_name"],
                employee["role"],
                employee["department"],
                last_working_day,
                settlement_month,

                context["monthly_salary"],
                context["paid_days"],
                context["earned_salary"],

                leave_balance,
                leave_encashment,
                bonus_payable,
                gratuity_payable,
                other_earnings,

                notice_recovery,
                loan_recovery,
                advance_recovery,
                other_deductions,

                total_earnings,
                total_deductions,
                final_payable,

                reason,
                remarks,
                now,

                context["salary_days_policy"],
                context["salary_days_basis"],
                context["attendance_source"],
                leave_encashment_basis,
                leave_encashment_rate,

                pf_employee,
                esi_employee,
                professional_tax,
                lwf_employee,
                tds,
                statutory_deductions,

                payment_status,
                payment_date
            ))

            cur.execute("""
                UPDATE employees
                SET date_of_exit = ?
                WHERE company_id = ?
                  AND emp_code = ?
            """, (
                last_working_day,
                company_id,
                emp_code
            ))

            conn.commit()

            flash(
                f"Full & Final settlement created for "
                f"{employee['employee_name']}. "
                f"Final Payable: ₹{final_payable}",
                "success"
            )

            return redirect(url_for("full_and_final"))

        except ValueError as error:
            conn.rollback()
            flash(str(error), "danger")
            return redirect(url_for("full_and_final"))

        except Exception as error:
            conn.rollback()
            flash(
                "Error while creating Full & Final "
                f"settlement: {str(error)}",
                "danger"
            )
            return redirect(url_for("full_and_final"))

        finally:
            try:
                conn.close()
            except Exception:
                pass

    conn = get_db()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT
                e.emp_code,
                e.employee_name,
                e.role,
                e.department,
                e.monthly_salary,
                COALESCE(e.date_of_exit, '') AS date_of_exit,
                COALESCE(lb.paid_leave, 0) AS paid_leave
            FROM employees AS e

            LEFT JOIN leave_balances AS lb
                ON e.company_id = lb.company_id
               AND e.emp_code = lb.emp_code

            WHERE e.company_id = ?
              AND (
                    e.date_of_exit IS NULL
                    OR TRIM(e.date_of_exit) = ''
              )

            ORDER BY e.employee_name
        """, (company_id,))

        employees = cur.fetchall()

        cur.execute("""
            SELECT *
            FROM full_final_settlements
            WHERE company_id = ?
            ORDER BY id DESC
        """, (company_id,))

        settlements = cur.fetchall()

        total_settlements = len(settlements)

        total_final_payable = _fnf_money_round(
            sum(
                _fnf_to_float(
                    row["final_payable"],
                    0
                )
                for row in settlements
            )
        )

        total_earnings = _fnf_money_round(
            sum(
                _fnf_to_float(
                    row["total_earnings"],
                    0
                )
                for row in settlements
            )
        )

        total_deductions = _fnf_money_round(
            sum(
                _fnf_to_float(
                    row["total_deductions"],
                    0
                )
                for row in settlements
            )
        )

        total_paid = _fnf_money_round(
            sum(
                _fnf_to_float(
                    row["final_payable"],
                    0
                )
                for row in settlements
                if str(
                    row["payment_status"]
                    or ""
                ).strip().lower() == "paid"
            )
        )

    except Exception as error:
        flash(
            "Error while loading Full & Final data: "
            + str(error),
            "danger"
        )
        return redirect(url_for("dashboard"))

    finally:
        conn.close()

    return render_template(
        "full_and_final.html",
        employees=employees,
        settlements=settlements,
        total_settlements=total_settlements,
        total_final_payable=total_final_payable,
        total_earnings=total_earnings,
        total_deductions=total_deductions,
        total_paid=total_paid,
        today=datetime.datetime.now().strftime(
            "%Y-%m-%d"
        )
    )


@app.route(
    "/update-fnf-payment/<int:settlement_id>",
    methods=["POST"]
)
@login_required
def update_fnf_payment(settlement_id):
    if not require_pro_feature(
        "Upgrade to PRO to update Full & Final payment."
    ):
        return redirect(url_for("pricing"))

    company_id = current_company_id()

    if not company_id:
        flash(
            "Company not found. Please login again.",
            "danger"
        )
        return redirect(url_for("login"))

    ensure_full_final_columns()

    payment_status = (
        request.form.get(
            "payment_status",
            "Pending"
        )
        .strip()
        .title()
    )

    payment_date = request.form.get(
        "payment_date",
        ""
    ).strip()

    if payment_status not in {
        "Pending",
        "Paid",
        "Hold"
    }:
        flash("Invalid payment status.", "danger")
        return redirect(url_for("full_and_final"))

    if payment_status == "Paid":
        if not payment_date:
            payment_date = (
                datetime.datetime.now()
                .strftime("%Y-%m-%d")
            )

        try:
            datetime.datetime.strptime(
                payment_date,
                "%Y-%m-%d"
            )
        except ValueError:
            flash(
                "Payment Date must be a valid date.",
                "danger"
            )
            return redirect(url_for("full_and_final"))

    else:
        payment_date = ""

    conn = get_db()
    cur = conn.cursor()

    try:
        cur.execute("""
            UPDATE full_final_settlements
            SET payment_status = ?,
                payment_date = ?
            WHERE id = ?
              AND company_id = ?
        """, (
            payment_status,
            payment_date,
            settlement_id,
            company_id
        ))

        if cur.rowcount == 0:
            flash(
                "Full & Final settlement not found.",
                "warning"
            )
        else:
            conn.commit()
            flash(
                "Payment status updated successfully.",
                "success"
            )

    except Exception as error:
        conn.rollback()
        flash(
            "Unable to update payment status: "
            + str(error),
            "danger"
        )

    finally:
        conn.close()

    return redirect(url_for("full_and_final"))


@app.route(
    "/download-fnf-excel/<int:settlement_id>"
)
@login_required
def download_fnf_excel(settlement_id):
    if not require_pro_feature(
        "Upgrade to PRO to download Full & Final Excel."
    ):
        return redirect(url_for("pricing"))

    company_id = current_company_id()

    if not company_id:
        flash(
            "Company not found. Please login again.",
            "danger"
        )
        return redirect(url_for("login"))

    ensure_full_final_columns()

    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            f.*,
            COALESCE(c.company_name, '') AS company_name,
            COALESCE(c.address, '') AS company_address,
            COALESCE(c.email, '') AS company_email,
            COALESCE(c.phone, '') AS company_phone
        FROM full_final_settlements AS f

        JOIN companies AS c
          ON f.company_id = c.id

        WHERE f.id = ?
          AND f.company_id = ?
    """, (settlement_id, company_id))

    row = cur.fetchone()
    conn.close()

    if not row:
        flash(
            "Full & Final settlement not found.",
            "warning"
        )
        return redirect(url_for("full_and_final"))

    os.makedirs(UPLOAD_FOLDER, exist_ok=True)

    emp_code = _fnf_safe_filename(
        row["emp_code"],
        "employee"
    )

    settlement_month = _fnf_safe_filename(
        row["settlement_month"],
        "month"
    )

    file_name = (
        f"fnf_settlement_{emp_code}_"
        f"{settlement_month}.xlsx"
    )

    file_path = os.path.join(
        UPLOAD_FOLDER,
        file_name
    )

    def money(value):
        return _fnf_money_round(value)

    def number_value(value):
        number = _fnf_to_float(value, 0)

        if float(number).is_integer():
            return int(number)

        return round(number, 2)

    company_phone = _fnf_clean_text(
        row["company_phone"],
        ""
    )

    data = [
        ["FULL & FINAL SETTLEMENT STATEMENT", ""],
        ["", ""],

        ["Company Details", ""],
        [
            "Company Name",
            _fnf_clean_text(
                row["company_name"],
                "SMARTHIRE AI"
            )
        ],
        [
            "Company Address",
            _fnf_clean_text(row["company_address"])
        ],
        [
            "Company Email",
            _fnf_clean_text(row["company_email"])
        ],
        [
            "Company Phone",
            company_phone or "-"
        ],
        ["", ""],

        ["Employee & Exit Details", ""],
        ["Employee Code", _fnf_clean_text(row["emp_code"])],
        ["Employee Name", _fnf_clean_text(row["employee_name"])],
        ["Designation", _fnf_clean_text(row["role"])],
        ["Department", _fnf_clean_text(row["department"])],
        [
            "Last Working Day",
            _fnf_clean_text(row["last_working_day"])
        ],
        [
            "Settlement Month",
            _fnf_clean_text(row["settlement_month"])
        ],
        [
            "Reason for Leaving",
            _fnf_clean_reason(row["reason"])
        ],
        ["", ""],

        ["Salary Calculation", ""],
        [
            "Salary Days Policy",
            _fnf_salary_policy_label(
                row["salary_days_policy"]
            )
        ],
        [
            "Salary Days Basis",
            number_value(row["salary_days_basis"])
        ],
        [
            "Paid Days",
            number_value(row["paid_days"])
        ],
        [
            "Attendance Source",
            _fnf_clean_text(row["attendance_source"])
        ],
        ["Monthly Salary", money(row["monthly_salary"])],
        ["Earned Salary", money(row["earned_salary"])],
        ["", ""],

        ["Earnings", "Amount"],
        [
            "Leave Balance",
            number_value(row["leave_balance"])
        ],
        [
            "Leave Encashment Basis",
            (
                "Manual Per-Day Rate"
                if row["leave_encashment_basis"] == "manual"
                else "Salary Per-Day Rate"
            )
        ],
        [
            "Leave Encashment Rate",
            money(row["leave_encashment_rate"])
        ],
        [
            "Leave Encashment",
            money(row["leave_encashment"])
        ],
        ["Bonus Payable", money(row["bonus_payable"])],
        [
            "Gratuity Payable",
            money(row["gratuity_payable"])
        ],
        ["Other Earnings", money(row["other_earnings"])],
        ["Total Earnings", money(row["total_earnings"])],
        ["", ""],

        ["Statutory Deductions", "Amount"],
        ["PF Employee", money(row["pf_employee"])],
        ["ESIC Employee", money(row["esi_employee"])],
        [
            "Professional Tax",
            money(row["professional_tax"])
        ],
        ["LWF Employee", money(row["lwf_employee"])],
        ["TDS", money(row["tds"])],
        [
            "Total Statutory Deductions",
            money(row["statutory_deductions"])
        ],
        ["", ""],

        ["Other Recoveries / Deductions", "Amount"],
        [
            "Notice Recovery",
            money(row["notice_recovery"])
        ],
        ["Loan Recovery", money(row["loan_recovery"])],
        [
            "Advance Recovery",
            money(row["advance_recovery"])
        ],
        [
            "Other Deductions",
            money(row["other_deductions"])
        ],
        ["Total Deductions", money(row["total_deductions"])],
        ["", ""],

        ["Final Settlement", ""],
        ["Final Payable", money(row["final_payable"])],
        [
            "Payment Status",
            _fnf_clean_text(row["payment_status"])
        ],
        [
            "Payment Date",
            _fnf_clean_text(row["payment_date"])
        ],
        ["Remarks", _fnf_clean_text(row["remarks"])],
        ["Created At", _fnf_clean_text(row["created_at"])],
        ["", ""],

        ["Approvals", ""],
        ["Prepared By", "________________________"],
        ["Checked By", "________________________"],
        [
            "HR / Authorized Signatory",
            "________________________"
        ]
    ]

    dataframe = pd.DataFrame(
        data,
        columns=["Particulars", "Details"]
    )

    with pd.ExcelWriter(
        file_path,
        engine="openpyxl"
    ) as writer:
        dataframe.to_excel(
            writer,
            index=False,
            sheet_name="F&F Settlement"
        )

        workbook = writer.book
        worksheet = workbook["F&F Settlement"]

        black_fill = PatternFill(
            start_color="000000",
            end_color="000000",
            fill_type="solid"
        )

        white_fill = PatternFill(
            start_color="FFFFFF",
            end_color="FFFFFF",
            fill_type="solid"
        )

        light_fill = PatternFill(
            start_color="F2F2F2",
            end_color="F2F2F2",
            fill_type="solid"
        )

        white_font = Font(
            bold=True,
            color="FFFFFF"
        )

        title_font = Font(
            bold=True,
            size=14,
            color="000000"
        )

        bold_font = Font(
            bold=True,
            color="000000"
        )

        normal_font = Font(color="000000")

        thin_border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000")
        )

        center = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        left = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True
        )

        right = Alignment(
            horizontal="right",
            vertical="center",
            wrap_text=True
        )

        for cell in worksheet[1]:
            cell.fill = black_fill
            cell.font = white_font
            cell.alignment = center
            cell.border = thin_border

        worksheet.merge_cells("A2:B2")
        worksheet["A2"].font = title_font
        worksheet["A2"].alignment = center

        section_labels = {
            "Company Details",
            "Employee & Exit Details",
            "Salary Calculation",
            "Earnings",
            "Statutory Deductions",
            "Other Recoveries / Deductions",
            "Final Settlement",
            "Approvals"
        }

        amount_labels = {
            "Monthly Salary",
            "Earned Salary",
            "Leave Encashment Rate",
            "Leave Encashment",
            "Bonus Payable",
            "Gratuity Payable",
            "Other Earnings",
            "Total Earnings",
            "PF Employee",
            "ESIC Employee",
            "Professional Tax",
            "LWF Employee",
            "TDS",
            "Total Statutory Deductions",
            "Notice Recovery",
            "Loan Recovery",
            "Advance Recovery",
            "Other Deductions",
            "Total Deductions",
            "Final Payable"
        }

        bold_rows = {
            "Total Earnings",
            "Total Statutory Deductions",
            "Total Deductions",
            "Final Payable"
        }

        for row_cells in worksheet.iter_rows(
            min_row=2,
            max_row=worksheet.max_row
        ):
            label = row_cells[0].value

            for cell in row_cells:
                cell.border = thin_border
                cell.fill = white_fill
                cell.font = normal_font
                cell.alignment = left

            if label in section_labels:
                worksheet.merge_cells(
                    start_row=row_cells[0].row,
                    start_column=1,
                    end_row=row_cells[0].row,
                    end_column=2
                )

                section_cell = worksheet.cell(
                    row=row_cells[0].row,
                    column=1
                )

                section_cell.font = bold_font
                section_cell.fill = light_fill
                section_cell.alignment = center

            if label in bold_rows:
                for cell in row_cells:
                    cell.font = bold_font
                    cell.fill = light_fill

            if label in amount_labels:
                amount_cell = worksheet.cell(
                    row=row_cells[0].row,
                    column=2
                )
                amount_cell.number_format = '₹#,##0'
                amount_cell.alignment = right

        worksheet.column_dimensions["A"].width = 34
        worksheet.column_dimensions["B"].width = 48
        worksheet.freeze_panes = "A4"

        for row_number in range(
            1,
            worksheet.max_row + 1
        ):
            worksheet.row_dimensions[
                row_number
            ].height = 20

        worksheet.page_setup.orientation = "portrait"
        worksheet.page_setup.paperSize = (
            worksheet.PAPERSIZE_A4
        )
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True

    return send_file(
        file_path,
        as_attachment=True,
        download_name=file_name
    )


@app.route(
    "/download-fnf-pdf/<int:settlement_id>"
)
@login_required
def download_fnf_pdf(settlement_id):
    if not require_pro_feature(
        "Upgrade to PRO to download Full & Final PDF."
    ):
        return redirect(url_for("pricing"))

    company_id = current_company_id()

    if not company_id:
        flash(
            "Company not found. Please login again.",
            "danger"
        )
        return redirect(url_for("login"))

    ensure_full_final_columns()

    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            f.*,
            COALESCE(c.company_name, '') AS company_name,
            COALESCE(c.address, '') AS company_address,
            COALESCE(c.email, '') AS company_email,
            COALESCE(c.phone, '') AS company_phone
        FROM full_final_settlements AS f

        JOIN companies AS c
          ON f.company_id = c.id

        WHERE f.id = ?
          AND f.company_id = ?
    """, (settlement_id, company_id))

    row = cur.fetchone()
    conn.close()

    if not row:
        flash(
            "Full & Final settlement not found.",
            "warning"
        )
        return redirect(url_for("full_and_final"))

    os.makedirs(PAYSLIP_FOLDER, exist_ok=True)

    emp_code = _fnf_safe_filename(
        row["emp_code"],
        "employee"
    )

    settlement_month = _fnf_safe_filename(
        row["settlement_month"],
        "month"
    )

    file_name = (
        f"fnf_settlement_{emp_code}_"
        f"{settlement_month}.pdf"
    )

    file_path = os.path.join(
        PAYSLIP_FOLDER,
        file_name
    )

    pdf = canvas.Canvas(
        file_path,
        pagesize=A4
    )

    width, height = A4

    def money_text(value):
        return "Rs. {:,}".format(
            _fnf_money_round(value)
        )

    def number_text(value):
        number = _fnf_to_float(value, 0)

        if float(number).is_integer():
            return str(int(number))

        return str(round(number, 2))

    def draw_line_item(
        label,
        value,
        y_value,
        left_x=52,
        value_x=None,
        bold=False
    ):
        if value_x is None:
            value_x = width - 52

        pdf.setFont(
            "Helvetica-Bold" if bold else "Helvetica",
            8.2
        )

        pdf.drawString(left_x, y_value, label)
        pdf.drawRightString(
            value_x,
            y_value,
            str(value)
        )

    company_name = _fnf_clean_text(
        row["company_name"],
        "SMARTHIRE AI"
    )

    company_address = _fnf_clean_text(
        row["company_address"],
        ""
    )

    company_email = _fnf_clean_text(
        row["company_email"],
        ""
    )

    company_phone = _fnf_clean_text(
        row["company_phone"],
        ""
    )

    pdf.setFillColorRGB(1, 1, 1)
    pdf.rect(
        0,
        0,
        width,
        height,
        fill=1,
        stroke=0
    )

    pdf.setFillColorRGB(0, 0, 0)
    pdf.setStrokeColorRGB(0, 0, 0)

    top_y = height - 42

    pdf.rect(
        36,
        top_y - 70,
        width - 72,
        70,
        fill=0,
        stroke=1
    )

    pdf.setFont("Helvetica-Bold", 15)
    pdf.drawCentredString(
        width / 2,
        top_y - 20,
        company_name.upper()[:65]
    )

    pdf.setFont("Helvetica-Bold", 9)
    pdf.drawCentredString(
        width / 2,
        top_y - 38,
        "FULL & FINAL SETTLEMENT STATEMENT"
    )

    contact_parts = [
        part
        for part in [
            company_address,
            company_email,
            company_phone
        ]
        if part
    ]

    pdf.setFont("Helvetica", 7.5)
    pdf.drawCentredString(
        width / 2,
        top_y - 56,
        " | ".join(contact_parts)[:120]
    )

    y = top_y - 92

    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(42, y, "Employee & Exit Details")

    y -= 10
    box_height = 112

    pdf.rect(
        42,
        y - box_height,
        width - 84,
        box_height
    )

    middle_x = width / 2
    pdf.line(
        middle_x,
        y,
        middle_x,
        y - box_height
    )

    left_details = [
        ("Employee Code", row["emp_code"]),
        ("Employee Name", row["employee_name"]),
        ("Designation", row["role"]),
        ("Department", row["department"]),
        (
            "Monthly Salary",
            money_text(row["monthly_salary"])
        )
    ]

    right_details = [
        (
            "Last Working Day",
            row["last_working_day"]
        ),
        (
            "Settlement Month",
            row["settlement_month"]
        ),
        (
            "Salary Policy",
            _fnf_salary_policy_label(
                row["salary_days_policy"]
            )
        ),
        (
            "Salary Days Basis",
            number_text(row["salary_days_basis"])
        ),
        (
            "Paid Days",
            number_text(row["paid_days"])
        )
    ]

    left_y = y - 20

    for label, value in left_details:
        pdf.setFont("Helvetica-Bold", 7.8)
        pdf.drawString(
            55,
            left_y,
            label + ":"
        )

        pdf.setFont("Helvetica", 7.8)
        pdf.drawString(
            150,
            left_y,
            _fnf_clean_text(value)[:34]
        )

        left_y -= 19

    right_y = y - 20

    for label, value in right_details:
        pdf.setFont("Helvetica-Bold", 7.8)
        pdf.drawString(
            middle_x + 12,
            right_y,
            label + ":"
        )

        pdf.setFont("Helvetica", 7.8)
        pdf.drawString(
            middle_x + 110,
            right_y,
            _fnf_clean_text(value)[:28]
        )

        right_y -= 19

    y -= box_height + 24

    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(42, y, "Settlement Calculation")

    y -= 12

    table_x = 42
    table_width = width - 84
    row_height = 18

    earnings = [
        ("Earned Salary", money_text(row["earned_salary"])),
        (
            "Leave Encashment",
            money_text(row["leave_encashment"])
        ),
        ("Bonus Payable", money_text(row["bonus_payable"])),
        (
            "Gratuity Payable",
            money_text(row["gratuity_payable"])
        ),
        ("Other Earnings", money_text(row["other_earnings"])),
        ("Total Earnings", money_text(row["total_earnings"]))
    ]

    deductions = [
        ("PF Employee", money_text(row["pf_employee"])),
        ("ESIC Employee", money_text(row["esi_employee"])),
        (
            "Professional Tax",
            money_text(row["professional_tax"])
        ),
        ("LWF Employee", money_text(row["lwf_employee"])),
        ("TDS", money_text(row["tds"])),
        (
            "Other Recoveries",
            money_text(
                _fnf_to_float(row["notice_recovery"])
                + _fnf_to_float(row["loan_recovery"])
                + _fnf_to_float(row["advance_recovery"])
                + _fnf_to_float(row["other_deductions"])
            )
        ),
        (
            "Total Deductions",
            money_text(row["total_deductions"])
        )
    ]

    table_rows = max(
        len(earnings),
        len(deductions)
    ) + 1

    table_height = row_height * table_rows

    pdf.rect(
        table_x,
        y - table_height,
        table_width,
        table_height
    )

    mid = table_x + table_width / 2

    pdf.line(
        mid,
        y,
        mid,
        y - table_height
    )

    left_amount_x = (
        table_x + table_width / 2 - 10
    )

    right_amount_x = table_x + table_width - 10

    pdf.setFont("Helvetica-Bold", 8.3)
    pdf.drawString(
        table_x + 8,
        y - 13,
        "Earnings"
    )
    pdf.drawRightString(
        left_amount_x,
        y - 13,
        "Amount"
    )
    pdf.drawString(
        mid + 8,
        y - 13,
        "Deductions"
    )
    pdf.drawRightString(
        right_amount_x,
        y - 13,
        "Amount"
    )

    for index in range(table_rows + 1):
        line_y = y - row_height * index
        pdf.line(
            table_x,
            line_y,
            table_x + table_width,
            line_y
        )

    for index in range(table_rows - 1):
        item_y = (
            y
            - row_height
            - 13
            - index * row_height
        )

        if index < len(earnings):
            label, value = earnings[index]
            draw_line_item(
                label,
                value,
                item_y,
                table_x + 8,
                left_amount_x,
                bold=label == "Total Earnings"
            )

        if index < len(deductions):
            label, value = deductions[index]
            draw_line_item(
                label,
                value,
                item_y,
                mid + 8,
                right_amount_x,
                bold=label == "Total Deductions"
            )

    y -= table_height + 20

    pdf.rect(
        42,
        y - 34,
        width - 84,
        34
    )

    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(
        55,
        y - 22,
        "FINAL PAYABLE: "
        + money_text(row["final_payable"])
    )

    pdf.setFont("Helvetica-Bold", 9)
    pdf.drawRightString(
        width - 55,
        y - 22,
        (
            "STATUS: "
            + _fnf_clean_text(
                row["payment_status"]
            ).upper()
        )
    )

    y -= 56

    pdf.setFont("Helvetica-Bold", 8.8)
    pdf.drawString(42, y, "Reason / Remarks")

    y -= 10

    pdf.rect(
        42,
        y - 38,
        width - 84,
        38
    )

    reason_remarks = (
        "Reason: "
        + _fnf_clean_reason(row["reason"])
        + " | Remarks: "
        + _fnf_clean_text(row["remarks"])
    )

    pdf.setFont("Helvetica", 7.8)
    pdf.drawString(
        54,
        y - 22,
        reason_remarks[:125]
    )

    y -= 68

    pdf.setFont("Helvetica", 8)

    signature_positions = [
        (42, "Prepared By"),
        (width / 2 - 50, "Checked By"),
        (width - 150, "HR / Authorized Signatory")
    ]

    for x_position, label in signature_positions:
        pdf.line(
            x_position,
            y + 14,
            x_position + 105,
            y + 14
        )
        pdf.drawString(
            x_position,
            y,
            label
        )

    pdf.setFont("Helvetica-Oblique", 6.8)
    pdf.drawCentredString(
        width / 2,
        30,
        (
            "Computer-generated Full & Final settlement. "
            "Signature is not required when digitally approved."
        )
    )

    pdf.save()

    return send_file(
        file_path,
        as_attachment=True,
        download_name=file_name
    )


@app.route(
    "/delete-fnf/<int:settlement_id>",
    methods=["POST"]
)
@login_required
def delete_fnf(settlement_id):
    if not require_pro_feature(
        "Upgrade to PRO to delete Full & Final settlements."
    ):
        return redirect(url_for("pricing"))

    company_id = current_company_id()

    if not company_id:
        flash(
            "Company not found. Please login again.",
            "danger"
        )
        return redirect(url_for("login"))

    ensure_full_final_columns()

    conn = get_db()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT
                id,
                emp_code,
                employee_name,
                last_working_day
            FROM full_final_settlements
            WHERE id = ?
              AND company_id = ?
        """, (settlement_id, company_id))

        row = cur.fetchone()

        if not row:
            flash(
                "Full & Final settlement not found.",
                "warning"
            )
            return redirect(url_for("full_and_final"))

        cur.execute("""
            DELETE FROM full_final_settlements
            WHERE id = ?
              AND company_id = ?
        """, (settlement_id, company_id))

        deleted_count = cur.rowcount

        if deleted_count > 0:
            cur.execute("""
                UPDATE employees
                SET date_of_exit = NULL
                WHERE company_id = ?
                  AND emp_code = ?
                  AND COALESCE(date_of_exit, '') = ?
            """, (
                company_id,
                row["emp_code"],
                row["last_working_day"]
            ))

            conn.commit()

            flash(
                "Full & Final settlement deleted for "
                f"{row['employee_name']}. "
                "Employee exit date was restored when it "
                "matched this settlement.",
                "success"
            )

        else:
            flash(
                "Settlement not found or already deleted.",
                "warning"
            )

    except Exception as error:
        conn.rollback()
        flash(
            "Error while deleting Full & Final "
            f"settlement: {str(error)}",
            "danger"
        )

    finally:
        conn.close()

    return redirect(url_for("full_and_final"))


@app.route("/delete-payroll/<month>", methods=["POST"])
@login_required
def delete_payroll(month):
    company_id = current_company_id()

    if not company_id:
        flash("Company not found. Please login again.", "danger")
        return redirect(url_for("login"))

    # Month validation: expected format YYYY-MM
    try:
        datetime.datetime.strptime(month, "%Y-%m")
    except ValueError:
        flash("Invalid payroll month selected.", "danger")
        return redirect(url_for("payroll_history"))

    conn = get_db()
    cur = conn.cursor()

    try:
        # Check records before delete
        cur.execute("""
            SELECT COUNT(*) AS total_records
            FROM payroll_history
            WHERE company_id = ?
              AND month = ?
        """, (company_id, month))

        result = cur.fetchone()
        total_records = result["total_records"] if result else 0

        if total_records <= 0:
            flash(f"No payroll records found for {month}.", "warning")
            return redirect(url_for("payroll_history", month=month))

        # Delete payroll records for selected company + selected month only.
        # Leave requests and leave balances are intentionally untouched.
        cur.execute("""
            DELETE FROM payroll_history
            WHERE company_id = ?
              AND month = ?
        """, (company_id, month))

        deleted_count = cur.rowcount
        conn.commit()

        flash(
            f"Payroll deleted successfully for {month}. "
            f"{deleted_count} record(s) removed. "
            "Leave records and leave balances were not changed.",
            "success"
        )

    except Exception as e:
        conn.rollback()
        flash(f"Error while deleting payroll: {str(e)}", "danger")

    finally:
        conn.close()

    return redirect(url_for("payroll_history", month=month))


# ---------------------------
# PAYSLIP PDF
# ---------------------------


def draw_payslip_branding_overlay(c, row, width, height):
    """
    Final overlay: payslip ke top header me sirf company logo draw karta hai.

    Authorized signature generate_payslip() ke footer section me already
    draw hoti hai, isliye signature ko yahan dobara draw nahi karna hai.
    """
    def clean(value, default=""):
        value = "" if value is None else str(value).strip()
        if not value or value.lower() in ["nan", "none", "null"]:
            return default
        return value

    try:
        company_id = row["company_id"]
    except Exception:
        company_id = current_company_id()

    assets = get_payslip_company_assets(company_id)
    logo_path = clean(assets.get("logo_path"), "")

    # Logo: top blue header ke left side me white background ke sath
    if logo_path and os.path.exists(logo_path):
        try:
            logo = ImageReader(logo_path)

            c.setFillColorRGB(1, 1, 1)
            c.roundRect(52, height - 101, 86, 48, 5, fill=1, stroke=0)

            c.drawImage(
                logo,
                57,
                height - 96,
                width=76,
                height=38,
                preserveAspectRatio=True,
                mask="auto"
            )

            c.setFillColorRGB(0, 0, 0)

        except Exception as e:
            print("Final logo overlay failed:", e)


def generate_payslip(row):
    os.makedirs(PAYSLIP_FOLDER, exist_ok=True)

    def clean_value(value, default="-"):
        if value is None:
            return default

        value = str(value).strip()

        if value == "":
            return default

        if value.lower() in ["nan", "none", "null"]:
            return default

        return value

    def safe_filename(value, default="file"):
        value = clean_value(value, default)

        for ch in [" ", "/", "\\", ":", "*", "?", '"', "<", ">", "|"]:
            value = value.replace(ch, "_")

        return value

    def to_float(value, default=0):
        try:
            if value is None or str(value).strip() == "":
                return default
            return float(value)
        except Exception:
            return default

    def money(value):
        try:
            return f"Rs. {int(round(float(value or 0)))}"
        except Exception:
            return "Rs. 0"

    def number_to_indian_words(value):
        """
        Integer amount ko Indian numbering system me words me convert karta hai.
        Example: 23483 -> Twenty Three Thousand Four Hundred Eighty Three
        """
        try:
            number = int(round(float(value or 0)))
        except Exception:
            number = 0

        if number == 0:
            return "Zero"

        if number < 0:
            return "Minus " + number_to_indian_words(abs(number))

        ones = [
            "", "One", "Two", "Three", "Four", "Five", "Six", "Seven",
            "Eight", "Nine", "Ten", "Eleven", "Twelve", "Thirteen",
            "Fourteen", "Fifteen", "Sixteen", "Seventeen", "Eighteen",
            "Nineteen"
        ]

        tens = [
            "", "", "Twenty", "Thirty", "Forty",
            "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"
        ]

        def below_hundred(n):
            if n < 20:
                return ones[n]
            return tens[n // 10] + (f" {ones[n % 10]}" if n % 10 else "")

        def below_thousand(n):
            words = []

            if n >= 100:
                words.append(f"{ones[n // 100]} Hundred")
                n %= 100

            if n:
                words.append(below_hundred(n))

            return " ".join(words)

        parts = []

        crore = number // 10000000
        number %= 10000000

        lakh = number // 100000
        number %= 100000

        thousand = number // 1000
        number %= 1000

        if crore:
            parts.append(f"{number_to_indian_words(crore)} Crore")

        if lakh:
            parts.append(f"{below_hundred(lakh)} Lakh")

        if thousand:
            parts.append(f"{below_hundred(thousand)} Thousand")

        if number:
            parts.append(below_thousand(number))

        return " ".join(parts)

    def wrap_words(value, max_chars=82):
        """
        Amount-in-words text ko maximum do readable lines me divide karta hai.
        """
        words = str(value).split()
        lines = []
        current = ""

        for word in words:
            candidate = word if not current else f"{current} {word}"

            if len(candidate) <= max_chars:
                current = candidate
            else:
                if current:
                    lines.append(current)
                current = word

        if current:
            lines.append(current)

        return lines[:2]

    def number_text(value, default="0"):
        try:
            if value is None or str(value).strip() == "":
                return default

            num = float(value)

            if num.is_integer():
                return str(int(num))

            return str(round(num, 2))

        except Exception:
            return default

    def row_value(key, default=""):
        try:
            return row[key]
        except Exception:
            return default

    def short_text(value, max_len=55):
        value = clean_value(value, "")

        if len(value) > max_len:
            return value[:max_len - 3] + "..."

        return value

    emp_code_for_file = safe_filename(row_value("emp_code"), "employee")
    month_for_file = safe_filename(row_value("month"), "month")
    employee_name_for_file = safe_filename(row_value("employee_name"), "employee")

    file_name = f"{emp_code_for_file}_{month_for_file}.pdf"
    file_path = os.path.join(PAYSLIP_FOLDER, file_name)

    c = canvas.Canvas(file_path, pagesize=letter)
    width, height = letter

    # Force white page background for official print-friendly PDF
    c.setFillColorRGB(1, 1, 1)
    c.rect(0, 0, width, height, fill=1, stroke=0)

    # Default text and border color black
    c.setFillColorRGB(0, 0, 0)
    c.setStrokeColorRGB(0, 0, 0)

    company_name = clean_value(row_value("company_name"), "SMART HIRE AI PAYROLL")
    company_address = clean_value(row_value("company_address"), "")
    company_email = clean_value(row_value("company_email"), "")
    company_phone = clean_value(row_value("company_phone"), "")

    emp_code = clean_value(row_value("emp_code"))
    employee_name = clean_value(row_value("employee_name"))
    role = clean_value(row_value("role"))
    department = clean_value(row_value("department"))
    gender = clean_value(row_value("gender"))

    uan_no = clean_value(row_value("uan_no"))
    esic_no = clean_value(row_value("esic_no"))
    bank_name = clean_value(row_value("bank_name"))
    account_no = clean_value(row_value("account_no"))
    ifsc_code = clean_value(row_value("ifsc_code"))

    working_days = row_value("attendance_working_days", row_value("working_days", 0))
    present_days = row_value("attendance_present_days", row_value("present_days", 0))
    weekly_off = row_value("attendance_weekly_off", row_value("weekly_off", 0))
    attendance_paid_leave = row_value("attendance_paid_leave", row_value("paid_leave_days", 0))
    holiday = row_value("attendance_holiday", row_value("holiday", 0))
    lop_days = row_value("attendance_lop_days", row_value("lwp_days", 0))
    attendance_paid_days = row_value("attendance_paid_days", row_value("payable_days", 0))
    overtime_hours = row_value("attendance_overtime_hours", row_value("overtime_hours", 0))

    payable_days = row_value("payable_days", attendance_paid_days)
    lwp_deduction = row_value("lwp_deduction", 0)

    calculated_paid_days = (
        to_float(present_days)
        + to_float(weekly_off)
        + to_float(attendance_paid_leave)
        + to_float(holiday)
        - to_float(lop_days)
    )

    if to_float(payable_days) <= 0:
        payable_days = calculated_paid_days

    absent_days = (
        to_float(working_days)
        - to_float(present_days)
        - to_float(weekly_off)
        - to_float(attendance_paid_leave)
        - to_float(holiday)
    )

    if absent_days < 0:
        absent_days = 0

    y = height - 30

    # Header
    c.setFillColorRGB(0.10, 0.17, 0.28)
    c.rect(35, y - 66, width - 70, 66, fill=1, stroke=0)

    c.setFillColorRGB(1, 1, 1)
    c.setFont("Helvetica-Bold", 16)
    c.drawCentredString(width / 2, y - 18, short_text(company_name.upper(), 60))

    c.setFont("Helvetica", 8.5)
    c.drawCentredString(width / 2, y - 34, "Salary Slip")

    address_line = short_text(company_address, 82) if company_address not in ["", "-"] else ""

    email_phone_parts = []

    if company_email not in ["", "-"]:
        email_phone_parts.append(company_email)

    if company_phone not in ["", "-"]:
        email_phone_parts.append(company_phone)

    email_phone_line = " | ".join(email_phone_parts)

    if address_line:
        c.drawCentredString(width / 2, y - 49, address_line)

    if email_phone_line:
        c.drawCentredString(width / 2, y - 61, short_text(email_phone_line, 82))

    y -= 88

    # Title
    c.setFillColorRGB(0, 0, 0)
    c.setFont("Helvetica-Bold", 13)
    c.drawCentredString(width / 2, y, f"SALARY SLIP - {clean_value(row_value('month'))}")

    y -= 26

    # Employee Info
    c.setFont("Helvetica-Bold", 10.5)
    c.drawString(40, y, "Employee Information")

    y -= 9

    emp_box_height = 118
    c.rect(40, y - emp_box_height, width - 80, emp_box_height)

    # Center divider
    c.line(width / 2, y, width / 2, y - emp_box_height)

    left = [
        ("Employee Code", emp_code),
        ("Employee Name", employee_name),
        ("Designation", role),
        ("Department", department),
        ("Gender", gender),
    ]

    right = [
        ("UAN No.", uan_no),
        ("ESIC No.", esic_no),
        ("Bank Name", bank_name),
        ("Account No.", account_no),
        ("IFSC Code", ifsc_code),
    ]

    yy = y - 22

    for label, value in left:
        c.setFont("Helvetica-Bold", 8.3)
        c.drawString(55, yy, f"{label}:")
        c.setFont("Helvetica", 8.3)
        c.drawString(150, yy, short_text(clean_value(value), 30))
        yy -= 20

    yy = y - 22

    for label, value in right:
        c.setFont("Helvetica-Bold", 8.3)
        c.drawString(320, yy, f"{label}:")
        c.setFont("Helvetica", 8.3)
        c.drawString(405, yy, short_text(clean_value(value), 28))
        yy -= 20

    y -= 140

    # Attendance Summary
    c.setFont("Helvetica-Bold", 10.5)
    c.drawString(40, y, "Attendance & Leave Summary")

    y -= 9

    att_box_height = 66
    c.rect(40, y - att_box_height, width - 80, att_box_height)

    c.setFont("Helvetica-Bold", 7.2)
    c.drawString(50, y - 14, "Working")
    c.drawString(110, y - 14, "Present")
    c.drawString(170, y - 14, "Weekly Off")
    c.drawString(245, y - 14, "Paid Leave")
    c.drawString(320, y - 14, "Holiday")
    c.drawString(380, y - 14, "LOP")
    c.drawString(430, y - 14, "Absent")
    c.drawString(490, y - 14, "Paid Days")

    c.setFont("Helvetica", 7.8)
    c.drawString(50, y - 29, number_text(working_days))
    c.drawString(110, y - 29, number_text(present_days))
    c.drawString(170, y - 29, number_text(weekly_off))
    c.drawString(245, y - 29, number_text(attendance_paid_leave))
    c.drawString(320, y - 29, number_text(holiday))
    c.drawString(380, y - 29, number_text(lop_days))
    c.drawString(430, y - 29, number_text(absent_days))
    c.drawString(490, y - 29, number_text(payable_days))

    c.setFont("Helvetica-Bold", 7.8)
    c.drawString(55, y - 51, "Overtime Hours")
    c.drawString(230, y - 51, "LOP Deduction")
    c.drawString(420, y - 51, "Pay Month")

    c.setFont("Helvetica", 7.8)
    c.drawString(155, y - 51, number_text(overtime_hours))
    c.drawString(330, y - 51, money(lwp_deduction))
    c.drawString(495, y - 51, clean_value(row_value("month")))

    y -= 84

    # Salary Details
    c.setFont("Helvetica-Bold", 10.5)
    c.drawString(40, y, "Salary Details")

    y -= 13

    x = 40
    table_w = width - 80
    row_h = 18
    table_y = y
    rows_count = 9

    c.rect(x, table_y - (row_h * rows_count), table_w, row_h * rows_count)

    c.setFillColorRGB(0.12, 0.36, 0.85)
    c.rect(x, table_y - row_h, table_w, row_h, fill=1, stroke=0)

    c.setFillColorRGB(0, 0, 0)
    c.line(x + table_w * 0.28, table_y, x + table_w * 0.28, table_y - row_h * rows_count)
    c.line(x + table_w * 0.50, table_y, x + table_w * 0.50, table_y - row_h * rows_count)
    c.line(x + table_w * 0.78, table_y, x + table_w * 0.78, table_y - row_h * rows_count)

    for i in range(rows_count + 1):
        c.line(x, table_y - row_h * i, x + table_w, table_y - row_h * i)

    c.setFillColorRGB(1, 1, 1)
    c.setFont("Helvetica-Bold", 8.5)
    c.drawString(x + 8, table_y - 12.5, "Earnings")
    c.drawString(x + table_w * 0.28 + 8, table_y - 12.5, "Amount")
    c.drawString(x + table_w * 0.50 + 8, table_y - 12.5, "Deductions")
    c.drawString(x + table_w * 0.78 + 8, table_y - 12.5, "Amount")

    c.setFillColorRGB(0, 0, 0)

    earnings = [
        ("Basic", row_value("basic", 0)),
        ("DA", row_value("da", 0)),
        ("HRA", row_value("hra", 0)),
        ("Special Allowance", row_value("special_allowance", 0)),
        ("Other Allowance", row_value("other_allowance", 0)),
        ("Overtime Amount", row_value("overtime_amount", 0)),
        ("Festival Bonus", row_value("festival_bonus", 0)),
    ]

    deductions = [
        ("PF Employee", row_value("pf_employee", 0)),
        ("ESIC Employee", row_value("esi_employee", 0)),
        ("Professional Tax", row_value("professional_tax", 0)),
        ("TDS", row_value("tds", 0)),
        ("Manual Deduction", row_value("manual_deduction", 0)),
        ("LWF Employee", row_value("lwf_employee", 0)),
        ("Total Deductions", row_value("total_deductions", 0)),
    ]

    c.setFont("Helvetica", 8.1)
    start_y = table_y - row_h - 12.5

    for i in range(7):
        yy = start_y - row_h * i

        c.drawString(x + 8, yy, earnings[i][0])
        c.drawRightString(x + table_w * 0.50 - 8, yy, money(earnings[i][1]))

        c.drawString(x + table_w * 0.50 + 8, yy, deductions[i][0])
        c.drawRightString(x + table_w - 8, yy, money(deductions[i][1]))

    total_y = table_y - row_h * 8 - 12.5

    c.setFont("Helvetica-Bold", 8.5)
    c.drawString(x + 8, total_y, "Gross Earnings")
    c.drawRightString(x + table_w * 0.50 - 8, total_y, money(row_value("gross", 0)))

    c.drawString(x + table_w * 0.50 + 8, total_y, "Net Pay")
    c.drawRightString(x + table_w - 8, total_y, money(row_value("net_pay", 0)))

    y = table_y - row_h * rows_count - 22

    # Employer Contributions
    c.setFont("Helvetica-Bold", 10.5)
    c.drawString(40, y, "Employer Contributions / CTC")

    y -= 9

    emp_ctc_box_height = 64
    c.rect(40, y - emp_ctc_box_height, width - 80, emp_ctc_box_height)

    pf_employer = to_float(row_value("pf_employer", 0))
    esi_employer = to_float(row_value("esi_employer", 0))
    gratuity = to_float(row_value("gratuity", 0))
    lwf_employer = to_float(row_value("lwf_employer", 0))

    employer_total = pf_employer + esi_employer + gratuity + lwf_employer

    c.setFont("Helvetica", 8.1)
    c.drawString(55, y - 14, f"PF Employer: {money(pf_employer)}")
    c.drawString(55, y - 29, f"ESIC Employer: {money(esi_employer)}")
    c.drawString(55, y - 44, f"Gratuity: {money(gratuity)}")
    c.drawString(55, y - 59, f"LWF Employer: {money(lwf_employer)}")

    c.drawString(330, y - 14, f"Monthly CTC: {money(row_value('monthly_ctc', 0))}")
    c.drawString(330, y - 29, f"Annual CTC: {money(row_value('annual_ctc', 0))}")
    c.drawString(330, y - 44, f"Bonus CTC: {money(row_value('bonus_ctc', 0))}")

    c.setFont("Helvetica-Bold", 8.1)
    c.drawString(330, y - 59, f"Employer Total: {money(employer_total)}")

    y -= 78

    # Net Pay Highlight + Amount in Words
    net_pay_value = to_float(row_value("net_pay", 0))
    net_pay_words = number_to_indian_words(net_pay_value)
    amount_in_words = f"AMOUNT IN WORDS: RUPEES {net_pay_words.upper()} ONLY"

    net_box_height = 52

    c.setFillColorRGB(0.86, 0.96, 0.89)
    c.rect(40, y - net_box_height, width - 80, net_box_height, fill=1, stroke=0)

    c.setFillColorRGB(0, 0.35, 0.15)
    c.setFont("Helvetica-Bold", 12)
    c.drawString(55, y - 18, f"NET PAYABLE: {money(net_pay_value)}")

    c.setFont("Helvetica-Bold", 7.2)
    amount_lines = wrap_words(amount_in_words, 88)

    for index, line in enumerate(amount_lines):
        c.drawString(55, y - 34 - (index * 9), line)

    c.setFillColorRGB(0, 0, 0)

    # Authorized Signature
    # Prepared By aur Checked By remove kiye gaye hain.
    y -= 72

    try:
        company_id_for_assets = row_value("company_id")
    except Exception:
        company_id_for_assets = current_company_id()

    company_assets = get_payslip_company_assets(company_id_for_assets)

    signature_path = clean_value(company_assets.get("authorized_signature_path"), "")
    authorized_signatory = clean_value(company_assets.get("authorized_signatory"), "")
    authorized_designation = clean_value(
        company_assets.get("authorized_designation"),
        "HR / Authorized Signatory"
    )

    sign_x = 445
    sign_y = y + 7
    sign_w = 115
    sign_h = 36

    if signature_path and os.path.exists(signature_path):
        try:
            img = ImageReader(signature_path)
            c.drawImage(
                img,
                sign_x,
                sign_y,
                width=sign_w,
                height=sign_h,
                preserveAspectRatio=True,
                mask="auto"
            )
        except Exception as e:
            print("Signature draw failed:", e)
            c.line(sign_x, sign_y, sign_x + sign_w, sign_y)
    else:
        c.line(sign_x, sign_y, sign_x + sign_w, sign_y)

    label = (
        authorized_signatory
        if authorized_signatory not in ["", "-"]
        else authorized_designation
    )

    c.setFont("Helvetica", 8)
    c.drawCentredString(sign_x + (sign_w / 2), y, short_text(label, 32))

    footer_y = y - 22

    if (
        authorized_signatory not in ["", "-"]
        and authorized_designation not in ["", "-"]
    ):
        c.setFont("Helvetica", 7)
        c.drawCentredString(
            sign_x + (sign_w / 2),
            y - 10,
            short_text(authorized_designation, 34)
        )
        footer_y = y - 30

    c.setFont("Helvetica-Oblique", 7)
    c.drawCentredString(
        width / 2,
        footer_y,
        "Computer-generated payslip. Signature not required if digitally approved."
    )

    draw_payslip_branding_overlay(c, row, width, height)

    c.save()

    return file_path


# ============================================================
# PAYROLL PRO - WHATSAPP PAYSLIP INTEGRATION
#
# Paste this complete block in app.py.
# Delete the old /download-payslip/<int:payroll_id> route first,
# because an updated route is included below.
#
# Required imports at the top of app.py:
# import os
# import time
# import datetime
# import requests
#
# Required environment variables:
# WHATSAPP_ACCESS_TOKEN
# WHATSAPP_PHONE_NUMBER_ID
# WHATSAPP_API_VERSION
# WHATSAPP_SEND_MODE=template
# WHATSAPP_TEMPLATE_NAME=salary_payslip_document
# WHATSAPP_TEMPLATE_LANGUAGE=en_US
# ============================================================


def ensure_whatsapp_payslip_log_table():
    """Create the WhatsApp payslip audit table safely."""
    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS whatsapp_payslip_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_id INTEGER NOT NULL,
            payroll_id INTEGER NOT NULL,
            emp_code TEXT,
            month TEXT,
            mobile_no TEXT,
            status TEXT NOT NULL,
            message_id TEXT,
            error_message TEXT,
            sent_at TEXT,
            created_at TEXT NOT NULL
        )
    """)

    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_whatsapp_payslip_logs_lookup
        ON whatsapp_payslip_logs (company_id, payroll_id, id)
    """)

    conn.commit()
    conn.close()


def clean_whatsapp_number_for_send(value):
    """Return an Indian WhatsApp number in 91XXXXXXXXXX format."""
    digits = "".join(
        ch for ch in str(value or "")
        if ch.isdigit()
    )

    if len(digits) == 10 and digits[0] in "6789":
        return "91" + digits

    if (
        len(digits) == 12
        and digits.startswith("91")
        and digits[2] in "6789"
    ):
        return digits

    if (
        len(digits) == 11
        and digits.startswith("0")
        and digits[1] in "6789"
    ):
        return "91" + digits[1:]

    return ""


def get_whatsapp_cloud_config():
    """Read WhatsApp Cloud API settings without hardcoding secrets."""
    config = {
        "access_token": os.getenv(
            "WHATSAPP_ACCESS_TOKEN",
            ""
        ).strip(),
        "phone_number_id": os.getenv(
            "WHATSAPP_PHONE_NUMBER_ID",
            ""
        ).strip(),
        "api_version": os.getenv(
            "WHATSAPP_API_VERSION",
            ""
        ).strip(),
        "send_mode": os.getenv(
            "WHATSAPP_SEND_MODE",
            "template"
        ).strip().lower(),
        "template_name": os.getenv(
            "WHATSAPP_TEMPLATE_NAME",
            "salary_payslip_document"
        ).strip(),
        "template_language": os.getenv(
            "WHATSAPP_TEMPLATE_LANGUAGE",
            "en_US"
        ).strip()
    }

    missing = []

    if not config["access_token"]:
        missing.append("WHATSAPP_ACCESS_TOKEN")

    if not config["phone_number_id"]:
        missing.append("WHATSAPP_PHONE_NUMBER_ID")

    if not config["api_version"]:
        missing.append("WHATSAPP_API_VERSION")

    if config["send_mode"] not in ["template", "document"]:
        raise RuntimeError(
            "WHATSAPP_SEND_MODE must be template or document."
        )

    if (
        config["send_mode"] == "template"
        and not config["template_name"]
    ):
        missing.append("WHATSAPP_TEMPLATE_NAME")

    if missing:
        raise RuntimeError(
            "Missing WhatsApp configuration: "
            + ", ".join(missing)
        )

    return config


def whatsapp_api_error(response):
    """Return a readable Meta API error without exposing the token."""
    try:
        payload = response.json()
    except Exception:
        return (
            f"HTTP {response.status_code}: "
            f"{response.text[:500]}"
        )

    error = payload.get("error", {})

    if error:
        message = str(error.get("message") or "WhatsApp API error")
        details = str(
            error.get("error_data", {}).get("details") or ""
        )
        code = error.get("code")

        result = message

        if code:
            result += f" (Code {code})"

        if details:
            result += f" - {details}"

        return result

    return str(payload)[:500]


def upload_whatsapp_pdf_media(pdf_path):
    """Upload one PDF to WhatsApp Cloud API and return its media ID."""
    config = get_whatsapp_cloud_config()

    if not pdf_path or not os.path.exists(pdf_path):
        raise FileNotFoundError("Payslip PDF file was not generated.")

    upload_url = (
        f"https://graph.facebook.com/"
        f"{config['api_version']}/"
        f"{config['phone_number_id']}/media"
    )

    headers = {
        "Authorization": f"Bearer {config['access_token']}"
    }

    with open(pdf_path, "rb") as pdf_file:
        response = requests.post(
            upload_url,
            headers=headers,
            data={
                "messaging_product": "whatsapp",
                "type": "application/pdf"
            },
            files={
                "file": (
                    os.path.basename(pdf_path),
                    pdf_file,
                    "application/pdf"
                )
            },
            timeout=90
        )

    if not response.ok:
        raise RuntimeError(whatsapp_api_error(response))

    payload = response.json()
    media_id = str(payload.get("id") or "").strip()

    if not media_id:
        raise RuntimeError(
            "WhatsApp media upload succeeded but media ID was missing."
        )

    return media_id


def send_whatsapp_pdf_document(
    mobile_no,
    media_id,
    filename,
    caption=""
):
    """
    Send a normal document message.
    Use this mode only when a customer-service conversation window is open.
    """
    config = get_whatsapp_cloud_config()
    whatsapp_number = clean_whatsapp_number_for_send(mobile_no)

    if not whatsapp_number:
        raise ValueError("Valid employee mobile number is missing.")

    messages_url = (
        f"https://graph.facebook.com/"
        f"{config['api_version']}/"
        f"{config['phone_number_id']}/messages"
    )

    payload = {
        "messaging_product": "whatsapp",
        "recipient_type": "individual",
        "to": whatsapp_number,
        "type": "document",
        "document": {
            "id": media_id,
            "filename": filename
        }
    }

    if caption:
        payload["document"]["caption"] = caption[:1024]

    response = requests.post(
        messages_url,
        headers={
            "Authorization": f"Bearer {config['access_token']}",
            "Content-Type": "application/json"
        },
        json=payload,
        timeout=60
    )

    if not response.ok:
        raise RuntimeError(whatsapp_api_error(response))

    return response.json()


def send_whatsapp_pdf_template(
    mobile_no,
    media_id,
    filename,
    employee_name,
    pay_month,
    company_name
):
    """
    Send an approved utility template containing a PDF document header.

    Expected approved template:
    Name: salary_payslip_document
    Category: Utility
    Header: Document
    Body: Dear {{1}}, your salary payslip for {{2}} from {{3}} is attached.
    """
    config = get_whatsapp_cloud_config()
    whatsapp_number = clean_whatsapp_number_for_send(mobile_no)

    if not whatsapp_number:
        raise ValueError("Valid employee mobile number is missing.")

    messages_url = (
        f"https://graph.facebook.com/"
        f"{config['api_version']}/"
        f"{config['phone_number_id']}/messages"
    )

    payload = {
        "messaging_product": "whatsapp",
        "recipient_type": "individual",
        "to": whatsapp_number,
        "type": "template",
        "template": {
            "name": config["template_name"],
            "language": {
                "code": config["template_language"]
            },
            "components": [
                {
                    "type": "header",
                    "parameters": [
                        {
                            "type": "document",
                            "document": {
                                "id": media_id,
                                "filename": filename
                            }
                        }
                    ]
                },
                {
                    "type": "body",
                    "parameters": [
                        {
                            "type": "text",
                            "text": str(employee_name or "Employee")[:60]
                        },
                        {
                            "type": "text",
                            "text": str(pay_month or "-")[:30]
                        },
                        {
                            "type": "text",
                            "text": str(company_name or "Company")[:100]
                        }
                    ]
                }
            ]
        }
    }

    response = requests.post(
        messages_url,
        headers={
            "Authorization": f"Bearer {config['access_token']}",
            "Content-Type": "application/json"
        },
        json=payload,
        timeout=60
    )

    if not response.ok:
        raise RuntimeError(whatsapp_api_error(response))

    return response.json()


def send_whatsapp_payslip_document(row, pdf_path):
    """Upload and send one generated payslip using configured mode."""
    config = get_whatsapp_cloud_config()

    mobile_no = row["mobile_no"]
    employee_name = row["employee_name"]
    pay_month = row["month"]
    company_name = row["company_name"]
    filename = os.path.basename(pdf_path)

    media_id = upload_whatsapp_pdf_media(pdf_path)

    if config["send_mode"] == "document":
        caption = (
            f"Dear {employee_name}, your salary payslip "
            f"for {pay_month} from {company_name} is attached."
        )

        return send_whatsapp_pdf_document(
            mobile_no=mobile_no,
            media_id=media_id,
            filename=filename,
            caption=caption
        )

    return send_whatsapp_pdf_template(
        mobile_no=mobile_no,
        media_id=media_id,
        filename=filename,
        employee_name=employee_name,
        pay_month=pay_month,
        company_name=company_name
    )


def get_payroll_payslip_row(payroll_id, company_id):
    """Fetch the complete payroll row used by PDF download and WhatsApp."""
    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            p.*,

            COALESCE(p.paid_leave_days, 0) AS paid_leave_days,
            COALESCE(p.lwp_days, 0) AS lwp_days,
            COALESCE(p.lwp_deduction, 0) AS lwp_deduction,
            COALESCE(p.payable_days, 0) AS payable_days,

            COALESCE(c.company_name, '') AS company_name,
            COALESCE(c.address, '') AS company_address,
            COALESCE(c.email, '') AS company_email,
            COALESCE(c.phone, '') AS company_phone,
            COALESCE(c.logo_path, '') AS company_logo_path,
            COALESCE(
                c.authorized_signature_path,
                ''
            ) AS authorized_signature_path,
            COALESCE(
                c.authorized_signatory,
                ''
            ) AS authorized_signatory,
            COALESCE(
                c.authorized_designation,
                ''
            ) AS authorized_designation,

            COALESCE(e.uan_no, '') AS uan_no,
            COALESCE(e.esic_no, '') AS esic_no,
            COALESCE(e.bank_name, '') AS bank_name,
            COALESCE(e.account_no, '') AS account_no,
            COALESCE(e.ifsc_code, '') AS ifsc_code,
            COALESCE(e.mobile_no, '') AS mobile_no,
            COALESCE(e.email_id, '') AS email_id,

            COALESCE(a.working_days, 0) AS attendance_working_days,
            COALESCE(a.present_days, 0) AS attendance_present_days,
            COALESCE(a.weekly_off, 0) AS attendance_weekly_off,
            COALESCE(a.paid_leave, 0) AS attendance_paid_leave,
            COALESCE(a.holiday, 0) AS attendance_holiday,
            COALESCE(a.lop_days, 0) AS attendance_lop_days,
            COALESCE(a.paid_days, 0) AS attendance_paid_days,
            COALESCE(
                a.overtime_hours,
                0
            ) AS attendance_overtime_hours

        FROM payroll_history AS p

        JOIN companies AS c
          ON p.company_id = c.id

        JOIN employees AS e
          ON p.company_id = e.company_id
         AND p.emp_code = e.emp_code

        LEFT JOIN attendance AS a
          ON p.company_id = a.company_id
         AND p.emp_code = a.emp_code
         AND p.month = a.month

        WHERE p.id = ?
          AND p.company_id = ?
          AND p.is_current = 1

        ORDER BY a.id DESC
        LIMIT 1
    """, (payroll_id, company_id))

    row = cur.fetchone()
    conn.close()

    return row


def log_whatsapp_payslip_result(
    company_id,
    payroll_id,
    emp_code,
    pay_month,
    mobile_no,
    status,
    message_id="",
    error_message=""
):
    """Save one WhatsApp payslip attempt for admin tracking."""
    ensure_whatsapp_payslip_log_table()

    now_text = datetime.datetime.now().strftime(
        "%Y-%m-%d %H:%M:%S"
    )

    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
        INSERT INTO whatsapp_payslip_logs (
            company_id,
            payroll_id,
            emp_code,
            month,
            mobile_no,
            status,
            message_id,
            error_message,
            sent_at,
            created_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        company_id,
        payroll_id,
        emp_code,
        pay_month,
        mobile_no,
        status,
        message_id,
        error_message[:1000],
        now_text if status == "Sent" else "",
        now_text
    ))

    conn.commit()
    conn.close()


def send_one_payroll_payslip_whatsapp(payroll_id, company_id):
    """Generate and send one employee payslip; never expose secrets."""
    row = get_payroll_payslip_row(payroll_id, company_id)

    if not row:
        return {
            "status": "Failed",
            "payroll_id": payroll_id,
            "month": "",
            "employee_name": "Employee",
            "error": "Payslip record not found."
        }

    emp_code = str(row["emp_code"] or "")
    pay_month = str(row["month"] or "")
    employee_name = str(row["employee_name"] or emp_code)
    mobile_no = str(row["mobile_no"] or "")
    normalized_mobile = clean_whatsapp_number_for_send(mobile_no)

    if not normalized_mobile:
        log_whatsapp_payslip_result(
            company_id=company_id,
            payroll_id=payroll_id,
            emp_code=emp_code,
            pay_month=pay_month,
            mobile_no=mobile_no,
            status="Mobile Missing",
            error_message="Valid 10-digit mobile number is missing."
        )

        return {
            "status": "Mobile Missing",
            "payroll_id": payroll_id,
            "month": pay_month,
            "employee_name": employee_name,
            "error": "Valid mobile number is missing."
        }

    try:
        pdf_path = generate_payslip(row)
        api_result = send_whatsapp_payslip_document(row, pdf_path)

        messages = api_result.get("messages") or []
        message_id = ""

        if messages and isinstance(messages[0], dict):
            message_id = str(messages[0].get("id") or "")

        log_whatsapp_payslip_result(
            company_id=company_id,
            payroll_id=payroll_id,
            emp_code=emp_code,
            pay_month=pay_month,
            mobile_no=normalized_mobile,
            status="Sent",
            message_id=message_id
        )

        return {
            "status": "Sent",
            "payroll_id": payroll_id,
            "month": pay_month,
            "employee_name": employee_name,
            "message_id": message_id,
            "error": ""
        }

    except Exception as e:
        error_text = str(e)

        log_whatsapp_payslip_result(
            company_id=company_id,
            payroll_id=payroll_id,
            emp_code=emp_code,
            pay_month=pay_month,
            mobile_no=normalized_mobile,
            status="Failed",
            error_message=error_text
        )

        print(
            "WhatsApp payslip send failed:",
            payroll_id,
            error_text
        )

        return {
            "status": "Failed",
            "payroll_id": payroll_id,
            "month": pay_month,
            "employee_name": employee_name,
            "error": error_text
        }


@app.route("/download-payslip/<int:payroll_id>")
@login_required
def download_payslip(payroll_id):
    if not require_pro_feature(
        "Upgrade to PRO to download PDF payslips."
    ):
        return redirect(url_for("pricing"))

    row = get_payroll_payslip_row(
        payroll_id,
        current_company_id()
    )

    if not row:
        flash("Payslip not found.", "danger")
        return redirect(url_for("payroll_history"))

    pdf_path = generate_payslip(row)

    return send_file(
        pdf_path,
        as_attachment=True,
        download_name=os.path.basename(pdf_path)
    )


@app.route(
    "/send-payslip-whatsapp/<int:payroll_id>",
    methods=["POST"]
)
@login_required
def send_payslip_whatsapp(payroll_id):
    if not require_pro_feature(
        "Upgrade to PRO to send WhatsApp payslips."
    ):
        return redirect(url_for("pricing"))

    company_id = current_company_id()

    if not company_id:
        flash("Company not found. Please login again.", "danger")
        return redirect(url_for("login"))

    result = send_one_payroll_payslip_whatsapp(
        payroll_id,
        company_id
    )

    if result["status"] == "Sent":
        flash(
            f"Payslip sent to {result['employee_name']} on WhatsApp.",
            "success"
        )

    elif result["status"] == "Mobile Missing":
        flash(
            f"WhatsApp not sent: mobile number missing for "
            f"{result['employee_name']}.",
            "warning"
        )

    else:
        flash(
            f"WhatsApp send failed for {result['employee_name']}: "
            f"{result['error']}",
            "danger"
        )

    selected_month = (
        request.form.get("month", "").strip()
        or result.get("month", "")
    )

    return redirect(
        url_for(
            "payroll_history",
            month=selected_month
        )
    )


@app.route(
    "/send-all-payslips-whatsapp",
    methods=["POST"]
)
@login_required
def send_all_payslips_whatsapp():
    if not require_pro_feature(
        "Upgrade to PRO to send bulk WhatsApp payslips."
    ):
        return redirect(url_for("pricing"))

    company_id = current_company_id()
    selected_month = request.form.get("month", "").strip()
    selected_department = request.form.get(
        "department",
        ""
    ).strip()

    if not company_id:
        flash("Company not found. Please login again.", "danger")
        return redirect(url_for("login"))

    if not selected_month:
        flash(
            "Please select a payroll month before bulk WhatsApp sending.",
            "warning"
        )
        return redirect(url_for("payroll_history"))

    ensure_whatsapp_payslip_log_table()

    conn = get_db()
    cur = conn.cursor()

    query = """
        SELECT id
        FROM payroll_history
        WHERE company_id = ?
          AND month = ?
          AND is_current = 1
    """

    params = [company_id, selected_month]

    if selected_department:
        query += " AND department = ?"
        params.append(selected_department)

    query += " ORDER BY id ASC"

    cur.execute(query, tuple(params))
    payroll_ids = [row["id"] for row in cur.fetchall()]
    conn.close()

    if not payroll_ids:
        flash(
            "No payroll records found for the selected month.",
            "warning"
        )
        return redirect(
            url_for(
                "payroll_history",
                month=selected_month,
                department=selected_department
            )
        )

    sent_count = 0
    failed_count = 0
    mobile_missing_count = 0

    for payroll_id in payroll_ids:
        result = send_one_payroll_payslip_whatsapp(
            payroll_id,
            company_id
        )

        if result["status"] == "Sent":
            sent_count += 1
        elif result["status"] == "Mobile Missing":
            mobile_missing_count += 1
        else:
            failed_count += 1

        # Gentle pause between requests during local/small-batch use.
        time.sleep(0.20)

    flash_category = (
        "success"
        if failed_count == 0
        else "warning"
    )

    flash(
        f"WhatsApp payslip summary — Sent: {sent_count}, "
        f"Mobile Missing: {mobile_missing_count}, "
        f"Failed: {failed_count}.",
        flash_category
    )

    return redirect(
        url_for(
            "payroll_history",
            month=selected_month,
            department=selected_department
        )
    )


@app.route("/download-all-payslips")
@login_required
def download_all_payslips():
    if not require_pro_feature("Upgrade to PRO to download all payslips."):
        return redirect(url_for("pricing"))

    month = request.args.get("month", "").strip()

    if not month:
        flash("Please select month to download payslips.", "warning")
        return redirect(url_for("payroll_history"))

    # Basic month validation: expected YYYY-MM
    try:
        datetime.datetime.strptime(month, "%Y-%m")
    except ValueError:
        flash("Invalid payroll month selected.", "danger")
        return redirect(url_for("payroll_history"))

    company_id = current_company_id()

    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            p.*,

            COALESCE(p.paid_leave_days, 0) AS paid_leave_days,
            COALESCE(p.lwp_days, 0) AS lwp_days,
            COALESCE(p.lwp_deduction, 0) AS lwp_deduction,
            COALESCE(p.payable_days, 0) AS payable_days,

            COALESCE(c.company_name, '') AS company_name,
            COALESCE(c.address, '') AS company_address,
            COALESCE(c.email, '') AS company_email,
            COALESCE(c.phone, '') AS company_phone,
            COALESCE(c.logo_path, '') AS company_logo_path,
            COALESCE(c.authorized_signature_path, '') AS authorized_signature_path,
            COALESCE(c.authorized_signatory, '') AS authorized_signatory,
            COALESCE(c.authorized_designation, '') AS authorized_designation,
            COALESCE(c.logo_path, '') AS company_logo_path,
            COALESCE(c.authorized_signature_path, '') AS authorized_signature_path,
            COALESCE(c.authorized_signatory, '') AS authorized_signatory,
            COALESCE(c.authorized_designation, '') AS authorized_designation,

            COALESCE(e.uan_no, '') AS uan_no,
            COALESCE(e.esic_no, '') AS esic_no,
            COALESCE(e.bank_name, '') AS bank_name,
            COALESCE(e.account_no, '') AS account_no,
            COALESCE(e.ifsc_code, '') AS ifsc_code,

            COALESCE(a.working_days, 0) AS attendance_working_days,
            COALESCE(a.present_days, 0) AS attendance_present_days,
            COALESCE(a.weekly_off, 0) AS attendance_weekly_off,
            COALESCE(a.paid_leave, 0) AS attendance_paid_leave,
            COALESCE(a.holiday, 0) AS attendance_holiday,
            COALESCE(a.lop_days, 0) AS attendance_lop_days,
            COALESCE(a.paid_days, 0) AS attendance_paid_days,
            COALESCE(a.overtime_hours, 0) AS attendance_overtime_hours

        FROM payroll_history p

        JOIN companies c
          ON p.company_id = c.id

        JOIN employees e
          ON p.company_id = e.company_id
         AND p.emp_code = e.emp_code

        LEFT JOIN attendance a
          ON p.company_id = a.company_id
         AND p.emp_code = a.emp_code
         AND p.month = a.month

        WHERE p.company_id = ?
          AND p.month = ?
          AND p.is_current = 1

        ORDER BY p.emp_code
    """, (company_id, month))

    rows = cur.fetchall()
    conn.close()

    if not rows:
        flash("No payroll data found for selected month.", "warning")
        return redirect(url_for("payroll_history", month=month))

    zip_buffer = BytesIO()

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for row in rows:
            pdf_path = generate_payslip(row)

            emp_code = str(row["emp_code"] or "EMP").strip()
            employee_name = str(row["employee_name"] or "Employee").strip()

            safe_employee_name = (
                employee_name
                .replace(" ", "_")
                .replace("/", "_")
                .replace("\\", "_")
                .replace(":", "_")
                .replace("*", "_")
                .replace("?", "_")
                .replace('"', "_")
                .replace("<", "_")
                .replace(">", "_")
                .replace("|", "_")
            )

            pdf_name = f"{emp_code}_{safe_employee_name}_{month}.pdf"
            zip_file.write(pdf_path, pdf_name)

    zip_buffer.seek(0)

    return send_file(
        zip_buffer,
        as_attachment=True,
        download_name=f"payslips_{month}.zip",
        mimetype="application/zip"
    )


@app.route("/faq")
@login_required
def faq():
    return render_template("faq.html")


# ---------------------------
# START APP
# ---------------------------
def setup_database():
    init_db()
    ensure_leave_tables()
    add_leave_payroll_columns()
    add_payment_order_id_column()


# Render / Gunicorn ke liye database setup
setup_database()


@app.route("/logout")
@login_required
def logout():
    session.clear()
    flash("Logged out successfully.", "success")
    return redirect(url_for("login"))


if __name__ == "__main__":
    app.run(debug=True)
